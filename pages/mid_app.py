import hashlib
import html
import json
import re
import uuid
from datetime import datetime
from io import BytesIO
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from streamlit_sortables import sort_items
except Exception:
    sort_items = None


# =========================
# 중학교 간편 생기부 v20
# =========================
st.set_page_config(
    page_title="중학교 간편 생기부",
    page_icon="🍯",
    layout="wide",
)

MID_APP_TITLE = "🍯 중학교 간편 생기부 v20"
MID_APP_SUBTITLE = "수행평가/관찰내용·평가 요소 기반 중학교 생기부 간편 작성 도우미 · patched-20260624-mid-v20"
MID_APP_VERSION = "patched-20260624-mid-v20"

MID_DEFAULT_RULES = """- 중학교 학교생활기록부 교과 세부능력 및 특기사항 문체로 작성한다.
- 학생 이름, 학년, 반, 번호, 학교명 등 개인정보를 쓰지 않는다.
- 첫 문장을 '학생은', '이 학생은', '해당 학생은'으로 시작하지 않는다.
- 활동 결과에 없는 내용을 추측하거나 과장하지 않는다.
- '깊은 이해', '창의융합', '혁신적', '흥미와 전문성 심화', '본인은', '의지를 밝힘' 같은 표현을 피한다.
- 한 문장 또는 짧은 한 문단으로 작성한다.
- 명사형 종결을 사용한다. 예: 수행함, 설명함, 정리함, 제시함, 해석함, 이해한 것으로 보임."""

MASTER_PROMPT = """너는 중학교 학교생활기록부 교과 세부능력 및 특기사항을 작성하는 교사 보조 도구이다.

[기본 원칙]
- 제공된 수행평가/관찰내용과 평가 요소별 성취수준 자료만 근거로 사용한다.
- 학생 이름, 학년, 반, 번호, 학교명 등 개인정보는 포함하지 않는다.
- 학생은, 이 학생은, 해당 학생은으로 문장을 시작하지 않는다.
- 성취수준 코드를 그대로 나열하지 않고 교사의 평가 문구를 자연스럽게 바꾸어 작성한다.
- 근거 없는 인성 평가, 성격 판단, 진로 추정, 태도 평가는 작성하지 않는다.
- 부정적 표현을 직접 쓰지 않고 현재 수행한 내용 중심으로 서술한다.
- 한 문단으로 작성한다.
- 문장은 명사형 종결 어미로 마무리한다.
- 제목, 번호, 설명, 따옴표, 안내 문구 없이 생기부 문장만 출력한다."""

AI_PROVIDER_OPTIONS = ["ChatGPT", "Gemini", "Claude"]
AI_DEFAULT_MODELS = {
    "ChatGPT": "gpt-5.5",
    "Gemini": "gemini-3.1-pro-preview",
    "Claude": "claude-fable-5",
}
AI_SECRET_KEY_NAMES = {
    "ChatGPT": "OPENAI_API_KEY",
    "Gemini": "GEMINI_API_KEY",
    "Claude": "ANTHROPIC_API_KEY",
}

VARIATION_GUIDES = {
    "낮음": "평가 내용은 그대로 유지하고 어휘와 문장 순서만 조금씩 바꾼다.",
    "보통": "평가 내용은 유지하되 문장 구조와 표현을 적당히 다르게 만든다.",
    "높음": "평가 내용은 벗어나지 않으면서 문장 흐름과 표현을 다양하게 구성한다.",
}


# =========================
# 공통 유틸
# =========================
def make_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def byte_count(text: str) -> int:
    return len(clean_text(text).encode("utf-8"))


def loading_elapsed_seconds(started_at=None) -> float:
    """전체 생성 중 진행 카드가 갱신되어도 로딩 문구 순서가 처음으로 돌아가지 않게 경과 시간을 계산한다."""
    if started_at is None:
        return 0.0
    if isinstance(started_at, datetime):
        base = started_at
    else:
        text = clean_text(started_at)
        if not text:
            return 0.0
        base = None
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"]:
            try:
                base = datetime.strptime(text[:19], fmt)
                break
            except Exception:
                pass
        if base is None:
            return 0.0
    try:
        return max(0.0, (datetime.now() - base).total_seconds())
    except Exception:
        return 0.0


GENERATION_LOADING_MESSAGES = [
    '칭찬 쥐어짜는 중...',
    '학생의 작은 성취를 확대 해석하지 않게 조심하는 중...',
    '복붙 냄새를 환기시키는 중...',
    '생기부 문체로 점잖게 번역하는 중...',
    '과장 표현은 빼고 근거만 남기는 중...',
    '‘깊은 이해’를 몰래 검열하는 중...',
    '교사 양심과 업무 효율 사이에서 줄타기 중...',
    '비슷하지만 안 비슷한 문장 찾는 중...',
    '성취수준을 생활기록부 말투로 바꾸는 중...',
    '문장 끝을 ‘함’으로 착지시키는 중...',
    '학생의 장점을 현미경으로 찾는 중...',
    '루브릭을 문장으로 반죽하는 중...',
    '너무 멋있어 보이지 않게 톤을 낮추는 중...',
    '평가 문구를 세특 문장으로 숙성시키는 중...',
    '행정 문체에 감성을 한 방울 넣는 중...',
    '학생 이름이 들어가지 않았는지 몰래 확인하는 중...',
    '적당히 구체적이고 적당히 담백한 문장 찾는 중...',
    '똑같은 말인데 다르게 보이게 다듬는 중...',
    '평가 자료 밖으로 상상력이 탈출하지 못하게 막는 중...',
    '근거 없는 칭찬을 얌전히 돌려보내는 중...',
    '문장을 한 문단 안에 얌전히 앉히는 중...',
    '성취수준 A를 자랑스럽지만 과하지 않게 포장하는 중...',
    '성취수준 C도 품격 있게 표현하는 중...',
    '‘보완 필요’라는 말을 예쁘게 우회하는 중...',
    '생기부에 들어가도 안 민망한 표현 찾는 중...',
    '너무 AI 같지 않게 숨 고르는 중...',
    '문장 사이 중복을 살살 덜어내는 중...',
    '활동 내용과 평가 문구를 자연스럽게 이어 붙이는 중...',
    '칭찬은 하되 무리수는 두지 않는 중...',
    '관찰 기록을 세특 문장으로 승진시키는 중...',
    '‘우수함’을 세 번 말하지 않으려고 노력하는 중...',
    '학생별 차이를 티 안 나게 살리는 중...',
    '생기부 문장에 교사 말투를 입히는 중...',
    '과학 활동을 행정 문장으로 변환하는 중...',
    '평가 요소들을 한 문장 안에서 화해시키는 중...',
    '문장이 너무 신나지 않게 진정시키는 중...',
    '교사의 관찰처럼 보이도록 문장을 다듬는 중...',
    '문장 끝맺음을 공손하게 정렬하는 중...',
    '‘탐구함’과 ‘분석함’ 사이에서 고민하는 중...',
    '생활기록부에 어울리는 온도로 데우는 중...',
    '같은 활동, 다른 표현을 열심히 짜내는 중...',
    '학생의 수행 내용을 근거 중심으로 압축하는 중...',
    '표현은 다양하게, 의미는 안전하게 맞추는 중...',
    '너무 칭찬 같지도, 너무 평범하지도 않게 조절하는 중...',
    '마지막 문장까지 명사형 종결로 착지 준비 중...',
]

def generation_loading_ticker_html(loading_offset_seconds=0) -> str:
    """생성 대기 중에도 화면에서 7초마다 바뀌는 짧은 안내 문구를 만든다."""
    if not GENERATION_LOADING_MESSAGES:
        return ""

    display_messages = GENERATION_LOADING_MESSAGES + [GENERATION_LOADING_MESSAGES[0]]
    line_height = 30
    interval_seconds = 7
    total_seconds = len(GENERATION_LOADING_MESSAGES) * interval_seconds
    try:
        offset_seconds = float(loading_offset_seconds or 0) % total_seconds
    except Exception:
        offset_seconds = 0.0
    message_lines = "".join([
        f'<div class="generation-loading-line">{html.escape(message)}</div>'
        for message in display_messages
    ])
    return f"""
    <div class="generation-loading-box">
<div class="generation-loading-window" style="height:{line_height}px;">
            <div class="generation-loading-track" style="animation: generationLoadingTicker {total_seconds}s steps({len(GENERATION_LOADING_MESSAGES)}) infinite; animation-delay:-{offset_seconds:.2f}s;">
                {message_lines}
            </div>
        </div>
    </div>
    """

def show_generation_overlay(slot, title, detail, progress_ratio=None, step_lines=None, recent_items=None, loading_offset_seconds=0):
    """AI 생성 중 진행 상황을 화면 안쪽에 표시하고, 직전 생성 완료 문장 1개만 보여준다."""
    if slot is None:
        slot = st.empty()

    safe_title = html.escape(clean_text(title))
    safe_detail = html.escape(clean_text(detail))
    loading_html = generation_loading_ticker_html(loading_offset_seconds=loading_offset_seconds)

    recent_items = recent_items or []
    previous_html = ""
    previous_item = None
    for item in reversed(recent_items):
        if isinstance(item, dict) and clean_text(item.get("text", "")):
            previous_item = item
            break

    if previous_item:
        label = html.escape(clean_text(previous_item.get("label", "직전 생성 학생")))
        text = html.escape(clean_text(previous_item.get("text", "")))
        previous_html = f"""
        <div class="generation-inline-previous">
            <div class="generation-inline-previous-title">바로 이전 학생 생성 결과</div>
            <div class="generation-inline-previous-label">{label}</div>
            <div class="generation-inline-previous-text">{text}</div>
        </div>
        """
    else:
        previous_html = """
        <div class="generation-inline-empty">
            아직 바로 이전 생성 결과가 없습니다. 첫 번째 학생이 생성되면 다음 학생 생성 중 이곳에 표시됩니다.
        </div>
        """

    if progress_ratio is None:
        progress_html = ""
    else:
        try:
            progress_value = max(0, min(100, int(float(progress_ratio) * 100)))
        except Exception:
            progress_value = 0
        progress_html = f"""
        <div class="generation-inline-progress-wrap">
            <div class="generation-inline-progress-bar" style="width:{progress_value}%;"></div>
        </div>
        <div class="generation-inline-percent">{progress_value}%</div>
        """

    card_html = f"""
    <!doctype html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
    html, body {{ margin: 0; padding: 0; background: transparent; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; }}
    .generation-inline-card {{
        width: calc(100% - 2px);
        background: #FFFFFF;
        border: 1px solid #CBD5E1;
        border-left: 8px solid #D92D20;
        border-radius: 16px;
        box-shadow: 0 3px 10px rgba(15, 23, 42, 0.08);
        padding: 15px 17px 14px 17px;
        color: #111827;
        box-sizing: border-box;
        margin: 0;
    }}
    .generation-inline-title {{ font-size: 1.02rem; font-weight: 900; margin-bottom: 5px; }}
    .generation-inline-detail {{ font-size: 0.93rem; font-weight: 700; color: #334155; margin-bottom: 9px; }}
    .generation-inline-progress-wrap {{ width: 100%; height: 10px; border-radius: 999px; background: #E5E7EB; overflow: hidden; margin-top: 8px; }}
    .generation-inline-progress-bar {{ height: 10px; border-radius: 999px; background: linear-gradient(90deg, #EF4444 0%, #D92D20 100%); transition: width 0.25s ease; }}
    .generation-inline-percent {{ margin-top: 4px; text-align: right; color: #64748B; font-size: 0.78rem; font-weight: 800; }}
    .generation-loading-box {{ margin-top: 10px; background: #FFF7ED; border: 1px solid #FED7AA; border-radius: 12px; padding: 8px 11px; display: flex; align-items: center; overflow: hidden; }}
    .generation-loading-label {{ display: none; }}
    .generation-loading-window {{ flex: 1 1 auto; overflow: hidden; min-width: 0; }}
    .generation-loading-track {{ will-change: transform; }}
    .generation-loading-line {{ height: 30px; line-height: 30px; color: #7C2D12; font-size: 0.91rem; font-weight: 900; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    @keyframes generationLoadingTicker {{ from {{ transform: translateY(0); }} to {{ transform: translateY(-1350px); }} }}
    .generation-inline-previous {{ margin-top: 12px; background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 12px; padding: 11px 13px; }}
    .generation-inline-previous-title {{ font-weight: 900; color: #0F172A; font-size: 0.9rem; margin-bottom: 6px; }}
    .generation-inline-previous-label {{ font-weight: 900; color: #1E3A8A; font-size: 0.84rem; margin-bottom: 5px; }}
    .generation-inline-previous-text {{ color: #111827; font-size: 0.9rem; line-height: 1.6; white-space: pre-wrap; word-break: keep-all; overflow-wrap: anywhere; }}
    .generation-inline-empty {{ margin-top: 10px; color: #64748B; background: #F8FAFC; border: 1px dashed #CBD5E1; border-radius: 12px; padding: 10px 12px; font-size: 0.88rem; font-weight: 700; }}
    </style>
    </head>
    <body>
    <div class="generation-inline-card">
        <div class="generation-inline-title">{safe_title}</div>
        <div class="generation-inline-detail">{safe_detail}</div>
        {loading_html}
        {progress_html}
        {previous_html}
    </div>
    </body>
    </html>
    """
    slot.empty()
    with slot.container():
        components.html(card_html, height=360, scrolling=True)
    return slot

def generation_preview_items_from_results(students_df, results, exclude_sid=""):
    """선택/전체 생성 중 바로 직전 생성 완료 문장 1개를 진행 상황 창에 표시하기 위한 자료로 변환한다."""
    if not isinstance(results, dict) or students_df is None or getattr(students_df, "empty", True):
        return []

    student_map = {clean_text(row.get("student_id", "")): row for _, row in students_df.iterrows()}
    entries = []
    for sid, result in results.items():
        sid = clean_text(sid)
        if not sid or sid == clean_text(exclude_sid) or not isinstance(result, dict):
            continue
        text = clean_text(result.get("edited", result.get("generated", "")))
        if not text:
            continue
        row = student_map.get(sid, {})
        label = (
            f"{clean_text(row.get('학년', ''))}학년 "
            f"{clean_text(row.get('반', ''))}반 "
            f"{clean_text(row.get('번호', ''))}번 "
            f"{clean_text(row.get('성명', ''))}"
        ).strip()
        entries.append(
            {
                "created_at": clean_text(result.get("created_at", "")),
                "label": label or "이전 생성 문장",
                "text": text,
            }
        )

    entries = sorted(entries, key=lambda x: x.get("created_at", ""))
    return [{"label": e["label"], "text": e["text"]} for e in entries[-1:]]

def to_int_or_big(value):
    text = re.sub(r"\D", "", clean_text(value))
    if text == "":
        return 999999
    return int(text)


def json_safe(obj):
    if isinstance(obj, dict):
        return {str(k): json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [json_safe(v) for v in obj]
    if isinstance(obj, tuple):
        return [json_safe(v) for v in obj]
    try:
        if pd.isna(obj):
            return ""
    except Exception:
        pass
    if hasattr(obj, "item"):
        try:
            return obj.item()
        except Exception:
            return str(obj)
    return obj


def default_level_code(index: int) -> str:
    default_codes = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    if 0 <= index < len(default_codes):
        return default_codes[index]
    return str(index + 1)


def as_dict(value):
    """이전 버전 세션/JSON에서 list 등으로 남은 값을 안전하게 dict로 바꾼다."""
    return value if isinstance(value, dict) else {}

def render_rubric_input_block(prefix, current_levels=None, current_rubrics=None):
    """고등학교용 app.py와 같은 방식으로 성취수준 코드와 평가 문구를 입력받는다."""
    current_levels = current_levels if isinstance(current_levels, list) else []
    current_rubrics = current_rubrics if isinstance(current_rubrics, dict) else {}

    default_count = len(current_levels) if current_levels else 5
    default_count = max(1, min(10, int(default_count)))

    level_count = st.selectbox(
        "성취수준 개수",
        options=list(range(1, 11)),
        index=default_count - 1,
        key=f"{prefix}_level_count",
    )
    st.markdown("**성취수준 코드와 성취 수준별 교사의 평가 문구**")

    levels = []
    rubrics = {}
    for idx in range(level_count):
        existing_code = current_levels[idx] if idx < len(current_levels) else default_level_code(idx)
        existing_text = current_rubrics.get(existing_code, "")
        col_code, col_text = st.columns([1, 5])
        with col_code:
            code = st.text_input(
                f"{idx + 1}번 코드",
                value=existing_code,
                key=f"{prefix}_code_{idx}",
                label_visibility="collapsed",
                placeholder="A",
            )
        with col_text:
            comment = st.text_area(
                f"{idx + 1}번 평가 문구",
                value=existing_text,
                key=f"{prefix}_comment_{idx}",
                height=80,
                label_visibility="collapsed",
                placeholder="내용을 입력하세요. 교사가 평가한 내용을 입력해야 합니다.",
            )
        code = clean_text(code)
        if code:
            levels.append(code)
            rubrics[code] = clean_text(comment)
    return levels, rubrics



def get_default_ai_model(provider: str) -> str:
    provider = provider if provider in AI_DEFAULT_MODELS else "ChatGPT"
    return AI_DEFAULT_MODELS[provider]


def get_default_ai_key(provider: str) -> str:
    secret_name = AI_SECRET_KEY_NAMES.get(provider, "OPENAI_API_KEY")
    try:
        return st.secrets.get(secret_name, "")
    except Exception:
        return ""


def sort_students_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    result = df.copy()
    for col in ["student_id", "학년", "반", "번호", "성명"]:
        if col not in result.columns:
            result[col] = ""
    result["_학년정렬"] = result["학년"].map(to_int_or_big)
    result["_반정렬"] = result["반"].map(to_int_or_big)
    result["_번호정렬"] = result["번호"].map(to_int_or_big)
    result = result.sort_values(
        by=["_학년정렬", "_반정렬", "_번호정렬", "성명"],
        kind="stable",
    ).drop(columns=["_학년정렬", "_반정렬", "_번호정렬"])
    return result.reset_index(drop=True)


def sort_assessments():
    return sorted(
        st.session_state.get("mid_assessments", []),
        key=lambda x: (
            int(x.get("order", 999) or 999),
            clean_text(x.get("name", "")),
            clean_text(x.get("assessment_id", "")),
        ),
    )


def get_items_for_assessment(assessment_id):
    return [
        item for item in st.session_state.get("mid_items", [])
        if isinstance(item, dict) and item.get("assessment_id", "") == assessment_id
    ]


def sort_items_for_assessment(assessment_id):
    return sorted(
        get_items_for_assessment(assessment_id),
        key=lambda x: (
            int(x.get("order", 999) or 999),
            clean_text(x.get("name", "")),
            clean_text(x.get("item_id", "")),
        ),
    )


def normalize_assessment_orders():
    """수행평가/관찰내용 순서를 1, 2, 3...으로 정리한다."""
    assessments = sort_assessments()
    for idx, assessment in enumerate(assessments, start=1):
        assessment["order"] = idx


def normalize_item_orders(assessment_id):
    """한 수행평가/관찰내용 안의 평가 요소 순서를 1, 2, 3...으로 정리한다."""
    items = sort_items_for_assessment(assessment_id)
    for idx, item in enumerate(items, start=1):
        item["order"] = idx


def apply_assessment_drag_order(sorted_labels, label_to_assessment_id):
    id_to_assessment = {
        assessment.get("assessment_id", ""): assessment
        for assessment in st.session_state.get("mid_assessments", [])
    }
    for idx, label in enumerate(sorted_labels, start=1):
        assessment_id = label_to_assessment_id.get(label)
        if assessment_id in id_to_assessment:
            id_to_assessment[assessment_id]["order"] = idx


def apply_item_drag_order(assessment_id, sorted_labels, label_to_item_id):
    id_to_item = {item.get("item_id", ""): item for item in get_items_for_assessment(assessment_id)}
    for idx, label in enumerate(sorted_labels, start=1):
        item_id = label_to_item_id.get(label)
        if item_id in id_to_item:
            id_to_item[item_id]["order"] = idx


def sortable_style():
    return """
    .sortable-component { width: 100% !important; border: 1px solid #D1D5DB !important; border-radius: 12px !important; padding: 10px !important; background-color: #F9FAFB !important; box-sizing: border-box !important; overflow: visible !important; }
    .sortable-container { width: 100% !important; background-color: #F9FAFB !important; border-radius: 10px !important; padding: 6px !important; box-sizing: border-box !important; overflow: visible !important; }
    .sortable-container-header { width: 100% !important; background-color: #F3F4F6 !important; color: #111827 !important; font-weight: 700 !important; padding: 8px 12px !important; border-radius: 8px !important; border: 1px solid #E5E7EB !important; box-sizing: border-box !important; }
    .sortable-container-body { width: 100% !important; display: flex !important; flex-direction: column !important; flex-wrap: nowrap !important; align-items: stretch !important; gap: 8px !important; background-color: #F9FAFB !important; padding-top: 8px !important; box-sizing: border-box !important; overflow: visible !important; }
    .sortable-item, .sortable-item:hover, .sortable-container-body > div, .sortable-container-body > div:hover { display: block !important; width: 100% !important; min-height: 44px !important; background-color: #FFFFFF !important; color: #111827 !important; font-weight: 700 !important; border: 1px solid #D1D5DB !important; border-radius: 10px !important; padding: 11px 14px !important; margin: 0 !important; box-shadow: 0 1px 2px rgba(17, 24, 39, 0.06) !important; box-sizing: border-box !important; white-space: normal !important; line-height: 1.35 !important; overflow-wrap: anywhere !important; }
    .sortable-item:hover, .sortable-container-body > div:hover { background-color: #F3F4F6 !important; border-color: #9CA3AF !important; }
    .sortable-item::before { content: "☰ " !important; color: #6B7280 !important; font-weight: 700 !important; }
    """


def sortable_key(base_key, labels):
    raw = "||".join([clean_text(label) for label in labels])
    digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:10]
    return f"{base_key}_{len(labels)}_{digest}"


def sort_labels_with_gray_box(labels, key, header="정렬"):
    if sort_items is None:
        st.warning("드래그 정렬 기능을 사용하려면 requirements.txt에 streamlit-sortables를 추가해야 합니다.")
        return labels
    if len(labels) < 2:
        return labels
    component_key = sortable_key(key, labels)
    try:
        return sort_items(labels, header=header, custom_style=sortable_style(), key=component_key)
    except TypeError:
        try:
            return sort_items(labels, key=f"{component_key}_basic")
        except Exception as e:
            st.error(f"드래그 정렬 컴포넌트 오류: {e}")
            return labels
    except Exception as e:
        st.error(f"드래그 정렬 컴포넌트 오류: {e}")
        return labels


def item_type_to_kor(item_type):
    return {
        "rubric": "성취도 선택형",
        "comment": "개별 코멘트형",
        "rubric_plus": "성취도 + 추가 코멘트형",
    }.get(item_type, "성취도 선택형")


def item_type_from_kor(label):
    return {
        "성취도 선택형": "rubric",
        "개별 코멘트형": "comment",
        "성취도 + 추가 코멘트형": "rubric_plus",
    }.get(label, "rubric")


def get_assessment_name(assessment_id):
    for assessment in st.session_state.get("mid_assessments", []):
        if assessment.get("assessment_id", "") == assessment_id:
            return assessment.get("name", "")
    return ""


def get_item_by_id(item_id):
    for item in st.session_state.get("mid_items", []):
        if item.get("item_id", "") == item_id:
            return item
    return None


def all_used_items():
    rows = []
    for assessment in sort_assessments():
        if not assessment.get("use", True):
            continue
        for item in sort_items_for_assessment(assessment.get("assessment_id", "")):
            rows.append(item)
    return rows


def record_key(student_id, item_id):
    return f"{student_id}::{item_id}"


def normalize_record_value(value):
    """이전 버전의 문자열 기록과 새 dict 기록을 모두 지원한다."""
    if isinstance(value, dict):
        return {
            "level": clean_text(value.get("level", "")),
            "comment": clean_text(value.get("comment", "")),
        }
    return {"level": clean_text(value), "comment": ""}


def get_record(student_id, item_id):
    raw = st.session_state.get("mid_records", {}).get(record_key(student_id, item_id), {})
    return normalize_record_value(raw)


def get_level(student_id, item_id):
    return get_record(student_id, item_id).get("level", "")


def get_comment(student_id, item_id):
    return get_record(student_id, item_id).get("comment", "")


def set_record(student_id, item_id, level="", comment=""):
    st.session_state.mid_records[record_key(student_id, item_id)] = {
        "level": clean_text(level),
        "comment": clean_text(comment),
    }


def set_level(student_id, item_id, level):
    current = get_record(student_id, item_id)
    set_record(student_id, item_id, level=level, comment=current.get("comment", ""))


def has_api_error_text(text: str) -> bool:
    text = clean_text(text)
    lowered = text.lower()
    markers = [
        "error code: 429",
        "insufficient_quota",
        "exceeded your current quota",
        "check your plan and billing details",
        "api 생성 중 오류",
        "api error",
    ]
    return any(marker in lowered for marker in markers)


def safe_result_text(result: dict) -> str:
    if not isinstance(result, dict):
        return ""
    text = clean_text(result.get("edited", result.get("generated", "")))
    if not has_api_error_text(text):
        return text
    fallback = clean_text(result.get("fallback", ""))
    return fallback if not has_api_error_text(fallback) else ""


def normalize_sentence(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"^[0-9]+[\.)]\s*", "", text)
    return text.strip(' \n\t-•·"\'“”‘’')


# =========================
# 세션 상태
# =========================
def init_mid_state():
    if "mid_settings" not in st.session_state:
        st.session_state.mid_settings = {
            "school_year": "2026",
            "semester": "1학기",
            "school_level": "중학교",
            "grade": "2",
            "subject": "과학",
            "target_bytes_min": 250,
            "target_bytes_max": 450,
            "custom_rules": MID_DEFAULT_RULES,
        }

    if "mid_assessments" not in st.session_state:
        st.session_state.mid_assessments = []

    if "mid_items" not in st.session_state:
        st.session_state.mid_items = []

    if "mid_students" not in st.session_state:
        st.session_state.mid_students = pd.DataFrame(
            columns=["student_id", "학년", "반", "번호", "성명"]
        )

    if "mid_records" not in st.session_state:
        st.session_state.mid_records = {}

    if "mid_results" not in st.session_state:
        st.session_state.mid_results = {}

    if "mid_selected_result_student_id" not in st.session_state:
        st.session_state.mid_selected_result_student_id = ""

    if "mid_generation_job" not in st.session_state:
        st.session_state.mid_generation_job = {
            "active": False,
            "stop_requested": False,
            "student_ids": [],
            "index": 0,
            "log": [],
            "ai_provider": "ChatGPT",
            "api_key": "",
            "model": "",
            "variation_level": "보통",
            "started_at": "",
            "loading_started_at": "",
            "finished_at": "",
        }


def sanitize_mid_state():
    settings = st.session_state.get("mid_settings", {})
    if "custom_rules" not in settings:
        settings["custom_rules"] = MID_DEFAULT_RULES
    if "target_bytes_min" not in settings:
        settings["target_bytes_min"] = 250
    if "target_bytes_max" not in settings:
        settings["target_bytes_max"] = 450
    st.session_state.mid_settings = settings

    if not isinstance(st.session_state.get("mid_students"), pd.DataFrame):
        st.session_state.mid_students = pd.DataFrame(st.session_state.get("mid_students", []))
    for col in ["student_id", "학년", "반", "번호", "성명"]:
        if col not in st.session_state.mid_students.columns:
            st.session_state.mid_students[col] = ""
    if not st.session_state.mid_students.empty:
        st.session_state.mid_students["student_id"] = st.session_state.mid_students["student_id"].apply(
            lambda x: clean_text(x) if clean_text(x) else make_id("mid_stu")
        )
        st.session_state.mid_students = sort_students_df(st.session_state.mid_students)

    clean_assessments = []
    for idx, assessment in enumerate(st.session_state.get("mid_assessments", []), start=1):
        if not isinstance(assessment, dict):
            continue
        if not clean_text(assessment.get("assessment_id", "")):
            assessment["assessment_id"] = make_id("mid_assess")
        assessment["name"] = clean_text(assessment.get("name", "")) or "이름 없는 수행평가/관찰내용"
        # 고등학교용 app.py와 호환되도록 area/unit 둘 다 보존한다.
        area_value = clean_text(assessment.get("area", assessment.get("unit", "")))
        assessment["area"] = area_value
        assessment["unit"] = area_value
        assessment["description"] = clean_text(assessment.get("description", ""))
        assessment["order"] = int(assessment.get("order", idx) or idx)
        assessment["use"] = bool(assessment.get("use", True))
        clean_assessments.append(assessment)
    clean_assessments = sorted(clean_assessments, key=lambda x: int(x.get("order", 999) or 999))
    for idx, assessment in enumerate(clean_assessments, start=1):
        assessment["order"] = idx
    st.session_state.mid_assessments = clean_assessments
    valid_assessment_ids = {a["assessment_id"] for a in clean_assessments}

    clean_items = []
    for idx, item in enumerate(st.session_state.get("mid_items", []), start=1):
        if not isinstance(item, dict):
            continue
        if item.get("assessment_id", "") not in valid_assessment_ids:
            continue
        if not clean_text(item.get("item_id", "")):
            item["item_id"] = make_id("mid_item")
        item["name"] = clean_text(item.get("name", "")) or "이름 없는 평가 요소"
        if item.get("type") not in ["rubric", "comment", "rubric_plus"]:
            item["type"] = "rubric"
        item["order"] = int(item.get("order", idx) or idx)
        if not isinstance(item.get("levels", []), list):
            item["levels"] = []
        if not isinstance(item.get("rubrics", {}), dict):
            item["rubrics"] = {}
        if item["type"] == "comment":
            item["levels"] = []
            item["rubrics"] = {}
        clean_items.append(item)
    st.session_state.mid_items = clean_items
    for assessment in st.session_state.mid_assessments:
        aid = assessment.get("assessment_id", "")
        items = sort_items_for_assessment(aid)
        for idx, item in enumerate(items, start=1):
            item["order"] = idx

    valid_student_ids = set(st.session_state.mid_students["student_id"].astype(str).tolist()) if not st.session_state.mid_students.empty else set()
    valid_item_ids = {item.get("item_id", "") for item in st.session_state.mid_items}

    raw_records = as_dict(st.session_state.get("mid_records", {}))
    clean_records = {}
    for key, value in raw_records.items():
        key = str(key)
        if "::" not in key:
            continue
        sid, item_id = key.split("::", 1)
        if sid in valid_student_ids and item_id in valid_item_ids:
            clean_records[key] = normalize_record_value(value)
    st.session_state.mid_records = clean_records

    raw_results = st.session_state.get("mid_results", {})
    # v01~v02에서는 mid_results가 list였기 때문에 v03 진입 시 AttributeError가 날 수 있다.
    # v04부터는 dict가 아닐 경우 빈 결과로 안전하게 초기화한다.
    raw_results = raw_results if isinstance(raw_results, dict) else {}
    clean_results = {}
    for sid, result in raw_results.items():
        sid = clean_text(sid)
        if sid in valid_student_ids and isinstance(result, dict):
            clean_results[sid] = result
    st.session_state.mid_results = clean_results

    job_defaults = {
        "active": False,
        "stop_requested": False,
        "student_ids": [],
        "index": 0,
        "log": [],
        "ai_provider": "ChatGPT",
        "api_key": "",
        "model": "",
        "variation_level": "보통",
        "started_at": "",
        "loading_started_at": "",
        "finished_at": "",
    }
    raw_job = st.session_state.get("mid_generation_job", {})
    raw_job = raw_job if isinstance(raw_job, dict) else {}
    job = {**job_defaults, **raw_job}
    if not isinstance(job.get("student_ids"), list):
        job["student_ids"] = []
    if not isinstance(job.get("log"), list):
        job["log"] = []
    try:
        job["index"] = int(job.get("index", 0) or 0)
    except Exception:
        job["index"] = 0
    st.session_state.mid_generation_job = job


init_mid_state()
sanitize_mid_state()


# =========================
# 프로젝트 저장/불러오기
# =========================
def mid_project_to_json() -> str:
    data = {
        "settings": st.session_state.mid_settings,
        "students": st.session_state.mid_students.to_dict(orient="records"),
        "assessments": st.session_state.mid_assessments,
        "items": st.session_state.mid_items,
        "records": st.session_state.mid_records,
        "results": st.session_state.mid_results,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "app": "개꿀 생기부 - 중학교 간편",
        "version": MID_APP_VERSION,
    }
    return json.dumps(json_safe(data), ensure_ascii=False, indent=2, default=str)


def apply_mid_project_data(data):
    st.session_state.mid_settings = data.get("settings", st.session_state.mid_settings)
    st.session_state.mid_students = pd.DataFrame(data.get("students", []))
    st.session_state.mid_assessments = data.get("assessments", [])
    st.session_state.mid_items = data.get("items", [])
    st.session_state.mid_records = data.get("records", {})
    st.session_state.mid_results = data.get("results", {})
    st.session_state.mid_selected_result_student_id = ""
    sanitize_mid_state()


def load_mid_project_json(uploaded_file):
    data = json.load(uploaded_file)
    apply_mid_project_data(data)


def build_mid_sample_project_data():
    """사진 기반/업로드 샘플 구조를 유지한 중학교 샘플 프로젝트를 반환한다."""
    data = json.loads(r"""{
  "settings": {
    "school_year": "2026",
    "semester": "1학기",
    "school_level": "중학교",
    "grade": "2",
    "subject": "과학",
    "target_bytes_min": 250,
    "target_bytes_max": 450,
    "custom_rules": "- 중학교 학교생활기록부 교과 세부능력 및 특기사항 문체로 작성한다.\n- 학생 이름, 학년, 반, 번호, 학교명 등 개인정보를 쓰지 않는다.\n- 첫 문장을 '학생은', '이 학생은', '해당 학생은'으로 시작하지 않는다.\n- 활동 결과에 없는 내용을 추측하거나 과장하지 않는다.\n- '깊은 이해', '창의융합', '혁신적', '흥미와 전문성 심화', '본인은', '의지를 밝힘' 같은 표현을 피한다.\n- 한 문장 또는 짧은 한 문단으로 작성한다.\n- 명사형 종결을 사용한다. 예: 수행함, 설명함, 정리함, 제시함, 해석함, 이해한 것으로 보임."
  },
  "students": [
    {
      "student_id": "mid_sample_stu_1_01",
      "학년": "2",
      "반": "1",
      "번호": "1",
      "성명": "김민준"
    },
    {
      "student_id": "mid_sample_stu_1_02",
      "학년": "2",
      "반": "1",
      "번호": "2",
      "성명": "박서연"
    },
    {
      "student_id": "mid_sample_stu_1_03",
      "학년": "2",
      "반": "1",
      "번호": "3",
      "성명": "최도윤"
    },
    {
      "student_id": "mid_sample_stu_1_04",
      "학년": "2",
      "반": "1",
      "번호": "4",
      "성명": "이하은"
    },
    {
      "student_id": "mid_sample_stu_1_05",
      "학년": "2",
      "반": "1",
      "번호": "5",
      "성명": "정우진"
    },
    {
      "student_id": "mid_sample_stu_1_06",
      "학년": "2",
      "반": "1",
      "번호": "6",
      "성명": "한지우"
    },
    {
      "student_id": "mid_sample_stu_1_07",
      "학년": "2",
      "반": "1",
      "번호": "7",
      "성명": "오서준"
    },
    {
      "student_id": "mid_sample_stu_1_08",
      "학년": "2",
      "반": "1",
      "번호": "8",
      "성명": "윤채아"
    },
    {
      "student_id": "mid_sample_stu_1_09",
      "학년": "2",
      "반": "1",
      "번호": "9",
      "성명": "Kim Alex"
    },
    {
      "student_id": "mid_sample_stu_1_10",
      "학년": "2",
      "반": "1",
      "번호": "10",
      "성명": "임하준"
    },
    {
      "student_id": "mid_sample_stu_2_01",
      "학년": "2",
      "반": "2",
      "번호": "1",
      "성명": "강민서"
    },
    {
      "student_id": "mid_sample_stu_2_02",
      "학년": "2",
      "반": "2",
      "번호": "2",
      "성명": "송예린"
    },
    {
      "student_id": "mid_sample_stu_2_03",
      "학년": "2",
      "반": "2",
      "번호": "3",
      "성명": "장도현"
    },
    {
      "student_id": "mid_sample_stu_2_04",
      "학년": "2",
      "반": "2",
      "번호": "4",
      "성명": "유나경"
    },
    {
      "student_id": "mid_sample_stu_2_05",
      "학년": "2",
      "반": "2",
      "번호": "5",
      "성명": "조현우"
    },
    {
      "student_id": "mid_sample_stu_2_06",
      "학년": "2",
      "반": "2",
      "번호": "6",
      "성명": "신아린"
    },
    {
      "student_id": "mid_sample_stu_2_07",
      "학년": "2",
      "반": "2",
      "번호": "7",
      "성명": "배준서"
    },
    {
      "student_id": "mid_sample_stu_2_08",
      "학년": "2",
      "반": "2",
      "번호": "8",
      "성명": "Park Lina"
    },
    {
      "student_id": "mid_sample_stu_2_09",
      "학년": "2",
      "반": "2",
      "번호": "9",
      "성명": "문태오"
    },
    {
      "student_id": "mid_sample_stu_2_10",
      "학년": "2",
      "반": "2",
      "번호": "10",
      "성명": "서지민"
    },
    {
      "student_id": "mid_sample_stu_3_01",
      "학년": "2",
      "반": "3",
      "번호": "1",
      "성명": "홍시우"
    },
    {
      "student_id": "mid_sample_stu_3_02",
      "학년": "2",
      "반": "3",
      "번호": "2",
      "성명": "권예준"
    },
    {
      "student_id": "mid_sample_stu_3_03",
      "학년": "2",
      "반": "3",
      "번호": "3",
      "성명": "남서아"
    },
    {
      "student_id": "mid_sample_stu_3_04",
      "학년": "2",
      "반": "3",
      "번호": "4",
      "성명": "구민재"
    },
    {
      "student_id": "mid_sample_stu_3_05",
      "학년": "2",
      "반": "3",
      "번호": "5",
      "성명": "하은별"
    },
    {
      "student_id": "mid_sample_stu_3_06",
      "학년": "2",
      "반": "3",
      "번호": "6",
      "성명": "Choi Ryan"
    },
    {
      "student_id": "mid_sample_stu_3_07",
      "학년": "2",
      "반": "3",
      "번호": "7",
      "성명": "백다온"
    },
    {
      "student_id": "mid_sample_stu_3_08",
      "학년": "2",
      "반": "3",
      "번호": "8",
      "성명": "민서율"
    },
    {
      "student_id": "mid_sample_stu_3_09",
      "학년": "2",
      "반": "3",
      "번호": "9",
      "성명": "노하윤"
    },
    {
      "student_id": "mid_sample_stu_3_10",
      "학년": "2",
      "반": "3",
      "번호": "10",
      "성명": "차지호"
    }
  ],
  "assessments": [
    {
      "assessment_id": "mid_assess_digest",
      "name": "혼합물 분리하기",
      "unit": "물질 - (8) 물질의 특성",
      "description": "물질의 특성을 활용하여 혼합물 분리하기",
      "order": 1,
      "use": true,
      "area": "물질 - (8) 물질의 특성"
    },
    {
      "assessment_id": "mid_assess_photo",
      "name": "환경 요인과 광합성의 관계를 알아보는 실험 설계하기",
      "unit": "생명 - (12) 식물과 에너지",
      "description": "성취기준\n\n[9과12-01] 광합성 과정을 이해하고, 환경 요인과 광합성의 관계를 탐구하는 실험을 설계할 수 있다.\n\n평가요소\n- 환경요인과 광합성의 관계를 탐구하기 위한 적절한 변인 통제\n- 실험 과정\n- 자료 해석 및 결과 예측",
      "order": 2,
      "use": true,
      "area": "생명 - (12) 식물과 에너지"
    }
  ],
  "items": [
    {
      "item_id": "mid_item_digest_order",
      "assessment_id": "mid_assess_digest",
      "name": "혼합물 분리하기",
      "levels": [
        "A",
        "B",
        "C",
        "D",
        "E"
      ],
      "rubrics": {
        "A": "탐구 수행 단계별로 가장 효율적인 분리 방법을 선택해 수행하여 물질을 모두 분리함",
        "B": "탐구 수행 단계별로 분리 방법을 선택해 수행하여 물질을 모두 분리함",
        "C": "선택한 분리 방법과 절차에 따라 수행했으나 일부 물질만 분리함",
        "D": "선택한 분리 방법과 절차 중 일부만 수행하여 일부 물질만 분리함",
        "E": "혼합물의 분리를 수행하지 못함"
      },
      "order": 1,
      "type": "rubric"
    },
    {
      "item_id": "mid_item_photo_variable",
      "assessment_id": "mid_assess_photo",
      "name": "환경요인과 광합성의 관계를 탐구하기 위한 적절한 변인 통제",
      "levels": [
        "A",
        "B",
        "C",
        "D"
      ],
      "rubrics": {
        "A": "광합성과 관계된 환경요인 중 다르게 해야할 조건 1개와 같게 해야할 조건 3가지를 명확히 서술하였음",
        "B": "광합성과 관계된 환경요인 중 다르게 해야할 조건 1개와 같게 해야할 조건 2가지만 명확히 서술하였음",
        "C": "광합성과 관계된 환경요인 중 다르게 해야할 조건 1개와 같게 해야할 조건 1가지만 명확히 서술하였음.",
        "D": "광합성과 관계된 환경요인 중 다르게 해야할 조건과 같게 해야할 조건을 구분하지 못함."
      },
      "order": 1,
      "type": "rubric"
    },
    {
      "item_id": "mid_item_photo_predict",
      "assessment_id": "mid_assess_photo",
      "name": "자료 해석 및 결과 예측",
      "levels": [
        "A",
        "B",
        "C",
        "D"
      ],
      "rubrics": {
        "A": "변인에 따른 광합성 정도에 대한 결과를 자료를 바탕으로 올바르게 예측함.",
        "B": "변인에 따른 광합성 정도에 대한 결과를 자료를 바탕으로 대체로 잘 예측함.",
        "C": "변인에 따른 광합성 정도에 대한 결과를 자료를 바탕으로 에측을 시도했으나, 부족한 부분이 있음.",
        "D": "변인에 따른 광합성 정도에 대한 결과를 자료를 바탕으로 예측하지 못함."
      },
      "order": 2,
      "type": "rubric"
    },
    {
      "item_id": "mid_item_f14f2aeb",
      "assessment_id": "mid_assess_digest",
      "name": "물질의 특성을 혼합물 분리 과정과 연결 짓기",
      "type": "rubric",
      "levels": [
        "A",
        "B",
        "C",
        "D"
      ],
      "rubrics": {
        "A": "물질의 특성(밀도, 용해도, 끓는점, 녹는점)을 혼합물의 분리에 적용한 이유를 단계별로 과학적인 근거를 들어, 오류 없이 설명함",
        "B": "물질의 특성(밀도, 용해도, 끓는점, 녹는점)을 혼합물의 분리에 적용한 이유를 모든 단계에 설명하였으나, 일부 오류가 있음",
        "C": "물질의 특성(밀도, 용해도, 끓는점, 녹는점)을 혼합물의 분리에 적용한 이유를 일부 단계만 설명함",
        "D": "물질의 특성(밀도, 용해도, 끓는점, 녹는점)을 혼합물의 분리에 적용한 이유를 설명하지 못함"
      },
      "order": 2
    },
    {
      "item_id": "mid_item_aaad2743",
      "assessment_id": "mid_assess_digest",
      "name": "과학 탐구 보고서 작성하기",
      "type": "rubric",
      "levels": [
        "A",
        "B",
        "C"
      ],
      "rubrics": {
        "A": "실험 계획서와 결과 보고서를 빠진 내용이나 오류 없이 작성함",
        "B": "실험 계획서와 결과 보고서가 일부 누락 되었거나, 일부 오류가 있음",
        "C": "혼합물의 분리 실험 계획서와 결과 보고서를 작성하지 못함"
      },
      "order": 3
    },
    {
      "item_id": "mid_item_7cf025e5",
      "assessment_id": "mid_assess_digest",
      "name": "실험 과정의 한계나 오차 요인 찾고 개선하기",
      "type": "comment",
      "levels": [],
      "rubrics": {},
      "order": 4
    },
    {
      "item_id": "mid_item_efe2ebac",
      "assessment_id": "mid_assess_photo",
      "name": "총평",
      "type": "comment",
      "levels": [],
      "rubrics": {},
      "order": 3
    }
  ],
  "records": {
    "mid_sample_stu_1_01::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_01::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_01::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_01::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_01::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_01::mid_item_7cf025e5": {
      "level": "",
      "comment": "처음에는 분리 순서를 조금 망설였지만, 밀도 차이와 용해도 차이를 차례로 떠올리며 가장 안정적인 절차를 찾아감. 다음에는 여과 전 가라앉히는 시간을 더 충분히 두면 결과가 더 깔끔해질 것 같음."
    },
    "mid_sample_stu_1_01::mid_item_efe2ebac": {
      "level": "",
      "comment": "변인 통제 표를 빠르게 채웠고, 같게 해야 할 조건을 실험 결과의 신뢰도와 연결해 말함. 예측도 무난했지만 그래프 자료를 근거로 드는 표현을 더 늘리면 좋겠음."
    },
    "mid_sample_stu_1_02::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_02::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_02::mid_item_photo_predict": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_02::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_02::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_03::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_03::mid_item_photo_variable": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_03::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_03::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_03::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_03::mid_item_7cf025e5": {
      "level": "",
      "comment": "실험 도구를 다루는 손길이 조심스러웠고, 분리 과정에서 왜 이 방법을 먼저 써야 하는지 친구에게 설명하려는 모습이 보였음. 오차 원인을 말할 때 실제 실험 장면을 근거로 든 점이 좋았음."
    },
    "mid_sample_stu_1_03::mid_item_efe2ebac": {
      "level": "",
      "comment": "광합성 실험에서 바꿀 조건과 유지할 조건을 구분하려고 많이 고민함. 처음 답은 조금 흔들렸지만, 피드백 후에는 변인 통제의 이유를 스스로 정리함."
    },
    "mid_sample_stu_1_04::mid_item_digest_order": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_04::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_04::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_04::mid_item_f14f2aeb": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_04::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_05::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_05::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_05::mid_item_photo_predict": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_05::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_05::mid_item_aaad2743": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_05::mid_item_7cf025e5": {
      "level": "",
      "comment": "혼합물이 한 번에 분리되지 않았을 때 다시 조건을 확인하려는 태도가 좋았음. 특히 물질이 섞인 상태를 보고 무작정 진행하기보다, 어떤 특성을 이용할 수 있을지 먼저 정리하려고 했음."
    },
    "mid_sample_stu_1_05::mid_item_efe2ebac": {
      "level": "",
      "comment": "빛의 세기와 광합성 정도를 연결하는 설명이 자연스러웠음. 실험 설계에서는 통제 변인을 더 꼼꼼히 적으려는 노력이 보였고, 결과 예측도 비교적 설득력 있었음."
    },
    "mid_sample_stu_1_06::mid_item_digest_order": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_1_06::mid_item_photo_variable": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_06::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_06::mid_item_f14f2aeb": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_1_06::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_07::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_07::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_07::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_07::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_07::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_07::mid_item_7cf025e5": {
      "level": "",
      "comment": "보고서에 실험 순서를 비교적 꼼꼼히 남겼고, 결과가 예상과 달랐던 이유를 도구 사용이나 가열 정도와 연결해 생각함. 다음에는 결과 표의 단위를 더 신경 쓰면 좋겠음."
    },
    "mid_sample_stu_1_07::mid_item_efe2ebac": {
      "level": "",
      "comment": "자료를 읽을 때 숫자 변화보다 전체 경향을 먼저 보려고 했음. 실험 과정 설명은 간단했지만, 결과가 왜 그렇게 나올지 자기 말로 풀어보려는 점이 좋았음."
    },
    "mid_sample_stu_1_08::mid_item_digest_order": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_08::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_08::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_08::mid_item_f14f2aeb": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_08::mid_item_aaad2743": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_1_09::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_09::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_09::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_09::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_09::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_09::mid_item_7cf025e5": {
      "level": "",
      "comment": "분리 방법을 고를 때 친구 의견을 듣고도 자기 생각을 근거와 함께 말하려고 했음. 끓는점 차이를 이용하는 과정에서 안전에 대한 언급도 자연스럽게 나와 인상적이었음."
    },
    "mid_sample_stu_1_09::mid_item_efe2ebac": {
      "level": "",
      "comment": "변인 통제에서 헷갈리는 조건을 친구에게 질문하고 다시 고치는 모습이 좋았음. 총평으로 적자면, 실험 설계의 틀을 잡는 힘이 조금씩 좋아지고 있음."
    },
    "mid_sample_stu_1_10::mid_item_digest_order": {
      "level": "E",
      "comment": ""
    },
    "mid_sample_stu_1_10::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_10::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_1_10::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_1_10::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_01::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_01::mid_item_photo_variable": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_01::mid_item_photo_predict": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_01::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_01::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_01::mid_item_7cf025e5": {
      "level": "",
      "comment": "결과가 완전히 깔끔하지 않았을 때 포기하지 않고 한계점을 직접 적어보려 했음. 개선 방안으로 실험 도구의 세척과 시료 양 조절을 말한 점이 실제적이었음."
    },
    "mid_sample_stu_2_01::mid_item_efe2ebac": {
      "level": "",
      "comment": "광합성과 환경 요인의 관계를 너무 외워서 쓰기보다 실험 조건으로 바꿔 생각하려고 했음. 자료 해석은 아직 조심스럽지만, 예측의 방향은 잘 잡았음."
    },
    "mid_sample_stu_2_02::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_02::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_02::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_02::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_02::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_03::mid_item_digest_order": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_03::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_03::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_03::mid_item_f14f2aeb": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_03::mid_item_aaad2743": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_03::mid_item_7cf025e5": {
      "level": "",
      "comment": "혼합물 분리 순서를 정리하는 데 시간이 걸렸지만, 마지막에는 물질의 특성과 실험 과정을 연결해 설명하려고 노력함. 설명이 조금 짧아도 핵심 단어는 잘 잡아냈음."
    },
    "mid_sample_stu_2_03::mid_item_efe2ebac": {
      "level": "",
      "comment": "처음에는 조건을 많이 나열했지만, 활동이 진행되면서 조작 변인과 통제 변인을 구분하는 방식이 정리됨. 결과 예측에서 이유를 한 문장 더 붙이면 더 좋아질 듯함."
    },
    "mid_sample_stu_2_04::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_04::mid_item_photo_variable": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_2_04::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_04::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_04::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_05::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_05::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_05::mid_item_photo_predict": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_05::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_05::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_05::mid_item_7cf025e5": {
      "level": "",
      "comment": "실험 계획서 작성에서 빠뜨린 부분을 스스로 찾아 고쳐 넣는 모습이 좋았음. 특히 왜 특정 방법을 사용했는지 말로 설명하려고 해서 활동 흐름이 분명해졌음."
    },
    "mid_sample_stu_2_05::mid_item_efe2ebac": {
      "level": "",
      "comment": "실험 설계를 표로 정리하는 데 강점이 있었고, 변인을 구분하는 과정도 비교적 안정적이었음. 다만 자료 해석에서 근거를 조금 더 직접적으로 써주면 좋겠음."
    },
    "mid_sample_stu_2_06::mid_item_digest_order": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_2_06::mid_item_photo_variable": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_06::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_06::mid_item_f14f2aeb": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_06::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_07::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_07::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_07::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_07::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_07::mid_item_aaad2743": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_07::mid_item_7cf025e5": {
      "level": "",
      "comment": "친구들이 놓친 오차 요인을 잘 짚어냈고, 결과가 흐려진 이유를 관찰 과정과 연결해 말함. 다음에는 개선 방안을 조금 더 구체적인 절차로 적어보면 좋겠음."
    },
    "mid_sample_stu_2_07::mid_item_efe2ebac": {
      "level": "",
      "comment": "환경 요인을 하나씩 바꿔야 한다는 점을 잘 이해했고, 실험 과정의 공정성을 강조한 점이 좋았음. 결과 예측도 자료의 방향성과 잘 맞았음."
    },
    "mid_sample_stu_2_08::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_08::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_08::mid_item_photo_predict": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_08::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_08::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_09::mid_item_digest_order": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_09::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_09::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_09::mid_item_f14f2aeb": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_2_09::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_09::mid_item_7cf025e5": {
      "level": "",
      "comment": "처음에는 보고서 정리가 느렸지만, 실험 후반으로 갈수록 관찰 내용을 자기 말로 정리하는 모습이 좋아졌음. 실험 결과와 물질의 특성을 연결하려는 시도가 보였음."
    },
    "mid_sample_stu_2_09::mid_item_efe2ebac": {
      "level": "",
      "comment": "광합성 조건을 정할 때 생활 속 예를 떠올리며 접근해서 이해가 빨랐음. 답안은 짧은 편이지만 핵심 조건을 놓치지 않으려는 모습이 보였음."
    },
    "mid_sample_stu_2_10::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_10::mid_item_photo_variable": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_2_10::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_2_10::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_2_10::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_01::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_01::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_01::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_01::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_01::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_01::mid_item_7cf025e5": {
      "level": "",
      "comment": "분리 방법을 선택하는 근거를 차분히 설명했고, 중간에 생긴 실수를 숨기기보다 오차 요인으로 정리한 점이 좋았음. 실제 탐구 과정이 잘 드러나는 기록이었음."
    },
    "mid_sample_stu_3_01::mid_item_efe2ebac": {
      "level": "",
      "comment": "자료 해석에서 그래프의 증가·감소 경향을 먼저 보고 결론을 쓰려는 태도가 좋았음. 실험 과정은 조금 더 세밀하게 적으면 완성도가 높아질 것 같음."
    },
    "mid_sample_stu_3_02::mid_item_digest_order": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_02::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_02::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_02::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_02::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_03::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_03::mid_item_photo_variable": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_03::mid_item_photo_predict": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_03::mid_item_f14f2aeb": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_03::mid_item_aaad2743": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_03::mid_item_7cf025e5": {
      "level": "",
      "comment": "실험 과정에서 세부 절차를 잘 챙겼고, 보고서에는 분리 전후의 차이를 비교하려는 흔적이 보였음. 물질의 특성을 더 정확한 용어로 쓰면 훨씬 좋아질 것 같음."
    },
    "mid_sample_stu_3_03::mid_item_efe2ebac": {
      "level": "",
      "comment": "변인 통제의 필요성을 자기 말로 설명하려고 했고, 실험 결과가 달라질 수 있는 이유를 조건 변화와 연결함. 전반적으로 차분히 설계하는 편임."
    },
    "mid_sample_stu_3_04::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_04::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_04::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_04::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_04::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_05::mid_item_digest_order": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_3_05::mid_item_photo_variable": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_3_05::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_05::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_05::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_05::mid_item_7cf025e5": {
      "level": "",
      "comment": "분리 결과가 기대만큼 나오지 않았지만, 그 이유를 시료의 양과 분리 시간에서 찾으려는 점이 좋았음. 실패한 장면도 탐구 자료로 삼으려는 태도가 보였음."
    },
    "mid_sample_stu_3_05::mid_item_efe2ebac": {
      "level": "",
      "comment": "실험 조건을 구분하는 데 시간이 걸렸지만, 끝까지 표를 채우며 구조를 이해하려고 했음. 총평은 아직 자신감은 부족하지만 생각의 방향은 잘 잡혀 있음."
    },
    "mid_sample_stu_3_06::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_06::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_06::mid_item_photo_predict": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_06::mid_item_f14f2aeb": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_06::mid_item_aaad2743": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_07::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_07::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_07::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_07::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_07::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_07::mid_item_7cf025e5": {
      "level": "",
      "comment": "혼합물을 분리하는 과정에서 어떤 방법을 먼저 적용할지 스스로 판단하려고 했음. 설명은 조금 단순했지만 실제 실험 장면과 연결해 말하려는 점이 좋았음."
    },
    "mid_sample_stu_3_07::mid_item_efe2ebac": {
      "level": "",
      "comment": "자료를 보고 결과를 예측하는 과정에서 근거를 찾으려는 모습이 좋았음. 설명이 길지는 않지만 핵심 단어를 놓치지 않고 쓰려는 점이 보였음."
    },
    "mid_sample_stu_3_08::mid_item_digest_order": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_08::mid_item_photo_variable": {
      "level": "C",
      "comment": ""
    },
    "mid_sample_stu_3_08::mid_item_photo_predict": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_08::mid_item_f14f2aeb": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_08::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_09::mid_item_digest_order": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_09::mid_item_photo_variable": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_09::mid_item_photo_predict": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_3_09::mid_item_f14f2aeb": {
      "level": "D",
      "comment": ""
    },
    "mid_sample_stu_3_09::mid_item_aaad2743": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_09::mid_item_7cf025e5": {
      "level": "",
      "comment": "보고서 작성에서 관찰 내용을 비교적 솔직하게 남겼고, 오차 원인을 단순히 실수라고 넘기지 않으려 했음. 다음에는 개선 방법을 한 단계 더 구체화하면 좋겠음."
    },
    "mid_sample_stu_3_09::mid_item_efe2ebac": {
      "level": "",
      "comment": "광합성 실험 설계에서 변인 통제의 의미를 점차 이해해 감. 특히 결과 예측을 할 때 “왜 그렇게 될까”를 스스로 묻는 모습이 좋았음."
    },
    "mid_sample_stu_3_10::mid_item_digest_order": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_10::mid_item_photo_variable": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_10::mid_item_photo_predict": {
      "level": "B",
      "comment": ""
    },
    "mid_sample_stu_3_10::mid_item_f14f2aeb": {
      "level": "A",
      "comment": ""
    },
    "mid_sample_stu_3_10::mid_item_aaad2743": {
      "level": "A",
      "comment": ""
    }
  },
  "results": {},
  "saved_at": "__NOW__",
  "app": "개꿀 생기부 - 중학교 간편",
  "version": "sample-mid-v20"
}""")
    data["saved_at"] = datetime.now().isoformat(timespec="seconds")
    return data


def load_mid_sample_data():
    apply_mid_project_data(build_mid_sample_project_data())


# =========================
# API 생성
# =========================
def _post_json(url, headers, payload, timeout=90):
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib_request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib_request.urlopen(req, timeout=timeout) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else {}
    except urllib_error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"HTTP {e.code}: {detail[:700]}") from e


def generate_with_openai(prompt, api_key, model):
    if not api_key:
        return None
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        response = client.responses.create(model=model, input=prompt)
        return response.output_text.strip()
    except Exception as e:
        msg = clean_text(str(e))
        if has_api_error_text(msg):
            st.warning("ChatGPT API 할당량/결제 관련 오류가 발생해 내부 조합 방식으로 대신 생성합니다.")
        else:
            st.warning("ChatGPT API 생성 중 오류가 발생해 내부 조합 방식으로 대신 생성합니다.")
        return None


def generate_with_gemini(prompt, api_key, model):
    if not api_key:
        return None
    try:
        model_path = urllib_parse.quote(clean_text(model), safe="")
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_path}:generateContent?key={urllib_parse.quote(api_key)}"
        payload = {
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.35},
        }
        data = _post_json(url, headers={"Content-Type": "application/json"}, payload=payload)
        parts = data.get("candidates", [{}])[0].get("content", {}).get("parts", [])
        text = "".join([part.get("text", "") for part in parts if isinstance(part, dict)]).strip()
        return text or None
    except Exception:
        st.warning("Gemini API 생성 중 오류가 발생해 내부 조합 방식으로 대신 생성합니다.")
        return None


def generate_with_claude(prompt, api_key, model):
    if not api_key:
        return None
    try:
        payload = {
            "model": model,
            "max_tokens": 2000,
            "temperature": 0.35,
            "messages": [{"role": "user", "content": prompt}],
        }
        data = _post_json(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
            payload=payload,
        )
        blocks = data.get("content", [])
        text = "".join([
            block.get("text", "")
            for block in blocks
            if isinstance(block, dict) and block.get("type") == "text"
        ]).strip()
        return text or None
    except Exception:
        st.warning("Claude API 생성 중 오류가 발생해 내부 조합 방식으로 대신 생성합니다.")
        return None


def generate_with_ai(prompt, ai_provider, api_key, model):
    provider = ai_provider if ai_provider in AI_PROVIDER_OPTIONS else "ChatGPT"
    model = clean_text(model) or get_default_ai_model(provider)
    if provider == "Gemini":
        return generate_with_gemini(prompt, api_key, model)
    if provider == "Claude":
        return generate_with_claude(prompt, api_key, model)
    return generate_with_openai(prompt, api_key, model)


def build_student_material(student):
    sid = student.get("student_id", "")
    lines = []
    # 개인정보 보호를 위해 이름/학년/반/번호는 프롬프트에 넣지 않는다.
    for assessment in sort_assessments():
        if not assessment.get("use", True):
            continue
        chunks = []
        for item in sort_items_for_assessment(assessment.get("assessment_id", "")):
            item_id = item.get("item_id", "")
            item_type = item.get("type", "rubric")
            level = get_level(sid, item_id)
            comment = get_comment(sid, item_id)
            rubrics = item.get("rubrics", {}) if isinstance(item.get("rubrics", {}), dict) else {}
            teacher_comment = clean_text(rubrics.get(level, ""))
            levels = [clean_text(x) for x in item.get("levels", []) if clean_text(x)]

            if item_type == "comment":
                if comment:
                    chunks.append(f"- {item.get('name', '')}: 학생별 개별 코멘트 / 교사의 기록: {comment}")
                continue

            if level or teacher_comment or comment:
                if levels:
                    level_text = f"전체 성취수준 {', '.join(levels)} 중 {level or '미입력'}"
                else:
                    level_text = f"성취수준 {level or '미입력'}"
                line = f"- {item.get('name', '')}: {level_text} / 교사의 평가: {teacher_comment}"
                if item_type == "rubric_plus" and comment:
                    line += f" / 추가 코멘트: {comment}"
                chunks.append(line)
        if chunks:
            lines.append(f"{assessment.get('name', '')}")
            area_text = clean_text(assessment.get("area", assessment.get("unit", "")))
            if area_text:
                lines.append(f"- 영역/단원: {area_text}")
            if assessment.get("description"):
                lines.append(f"- 설명/관찰 기준/활동 내용: {assessment.get('description', '')}")
            lines.extend(chunks)
            lines.append("")
    return "\n".join(lines).strip()

def build_prompt(material, variation_level="보통", variant_no=1):
    settings = st.session_state.mid_settings
    custom_rules = clean_text(settings.get("custom_rules", MID_DEFAULT_RULES))
    variation_guide = VARIATION_GUIDES.get(variation_level, VARIATION_GUIDES["보통"])
    return f"""
{MASTER_PROMPT}

[선생님 추가 작성 규칙]
{custom_rules}

[작성 조건]
- 목표 분량: {settings.get('target_bytes_min', 250)}~{settings.get('target_bytes_max', 450)} byte
- 중학교 생기부에 맞게 간결하게 작성한다.
- 같은 성취수준 학생끼리도 결과 문장이 완전히 같지 않도록 자연스럽게 변주한다.
- 표현 변주 기준: {variation_guide}
- 문장 변주 번호: {variant_no}

[평가 자료]
{material}

[최종 출력]
세부능력 및 특기사항 문장만 출력하라.
""".strip()


def fallback_generate(material, variant_no=1):
    pieces = []
    for line in material.splitlines():
        line = line.strip()
        if not line.startswith("- "):
            continue
        if "교사의 평가:" in line:
            value = line.split("교사의 평가:", 1)[1].strip()
            if value:
                pieces.append(value.rstrip("."))
    unique = []
    seen = set()
    for piece in pieces:
        if piece and piece not in seen:
            unique.append(piece)
            seen.add(piece)
    if not unique:
        return "입력된 성취수준 자료가 부족하여 생기부 문장 생성이 어려움."

    if variant_no % 3 == 1:
        text = ", ".join(unique[:4])
        if not text.endswith(("함", "임", "음", "보임")):
            text += "함"
        return text
    if variant_no % 3 == 2:
        text = " 및 ".join(unique[:3])
        if not text.endswith(("함", "임", "음", "보임")):
            text += "을 바탕으로 활동을 수행함"
        return text
    text = ". ".join(unique[:3])
    if not text.endswith(("함", "임", "음", "보임", ".")):
        text += "함"
    return text


def generate_for_student(student, ai_provider, api_key, model, variation_level, variant_no):
    material = build_student_material(student)
    prompt = build_prompt(material, variation_level=variation_level, variant_no=variant_no)
    generated = None
    if api_key:
        generated = generate_with_ai(prompt, ai_provider, api_key, model)
    if not generated:
        generated = fallback_generate(material, variant_no=variant_no)
    generated = normalize_sentence(generated)
    return material, generated


# =========================
# 입력표 / 엑셀
# =========================
def build_input_fields(items):
    """평가 요소의 기록 방식에 따라 입력표 열 구조를 만든다."""
    fields = []
    for item in items:
        item_id = item.get("item_id", "")
        item_type = item.get("type", "rubric")
        if item_type == "comment":
            fields.append({"item": item, "field": "comment", "column": f"{item_id}__comment", "label": f"{item.get('name', '')}\n추가 코멘트"})
        elif item_type == "rubric_plus":
            fields.append({"item": item, "field": "level", "column": f"{item_id}__level", "label": f"{item.get('name', '')}\n성취수준"})
            fields.append({"item": item, "field": "comment", "column": f"{item_id}__comment", "label": f"{item.get('name', '')}\n추가 코멘트"})
        else:
            fields.append({"item": item, "field": "level", "column": f"{item_id}__level", "label": item.get("name", "")})
    return fields


def build_record_matrix_df():
    students = st.session_state.mid_students.copy()
    items = all_used_items()
    fields = build_input_fields(items)
    if students.empty:
        students = pd.DataFrame([
            {
                "student_id": make_id("mid_stu"),
                "학년": st.session_state.mid_settings.get("grade", "2"),
                "반": "1",
                "번호": "1",
                "성명": "",
            }
        ])

    rows = []
    for _, student in students.iterrows():
        sid = clean_text(student.get("student_id", "")) or make_id("mid_stu")
        row = {
            "student_id": sid,
            "학년": clean_text(student.get("학년", "")),
            "반": clean_text(student.get("반", "")),
            "번호": clean_text(student.get("번호", "")),
            "성명": clean_text(student.get("성명", "")),
        }
        for field in fields:
            item_id = field["item"].get("item_id", "")
            rec = get_record(sid, item_id)
            row[field["column"]] = rec.get(field["field"], "")
        rows.append(row)
    return pd.DataFrame(rows), items, fields


def save_record_matrix_df(edited_df, items, fields=None):
    if edited_df.empty:
        return 0
    fields = fields if fields is not None else build_input_fields(items)
    students = edited_df[["student_id", "학년", "반", "번호", "성명"]].copy()
    students = students[students["성명"].astype(str).str.strip() != ""].reset_index(drop=True)
    students["student_id"] = students["student_id"].apply(lambda x: clean_text(x) if clean_text(x) else make_id("mid_stu"))
    st.session_state.mid_students = sort_students_df(students)

    valid_student_ids = set(st.session_state.mid_students["student_id"].astype(str).tolist())
    valid_item_ids = {item.get("item_id", "") for item in items}
    new_records = {}
    saved_count = 0
    for _, row in edited_df.iterrows():
        sid = clean_text(row.get("student_id", ""))
        if sid not in valid_student_ids:
            continue
        by_item = {item_id: {"level": "", "comment": ""} for item_id in valid_item_ids}
        for field in fields:
            item_id = field["item"].get("item_id", "")
            if item_id not in valid_item_ids:
                continue
            value = clean_text(row.get(field["column"], ""))
            by_item.setdefault(item_id, {"level": "", "comment": ""})[field["field"]] = value
        for item_id, rec in by_item.items():
            if rec.get("level") or rec.get("comment"):
                new_records[record_key(sid, item_id)] = rec
            saved_count += 1
    st.session_state.mid_records = new_records
    sanitize_mid_state()
    return saved_count

def make_record_input_excel():
    students = st.session_state.mid_students.copy()
    items = all_used_items()
    fields = build_input_fields(items)
    wb = Workbook()
    ws = wb.active
    ws.title = "학생별성취수준"
    list_ws = wb.create_sheet("선택목록")
    list_ws.sheet_state = "hidden"

    base_headers = ["student_id", "학년", "반", "번호", "성명"]
    start_row = 4
    data_start_row = 6
    item_start_col = 6
    last_col = max(item_start_col + len(fields) - 1, 5)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1).value = "중학교 생기부 학생별 성취수준 입력표"
    ws.cell(1, 1).font = Font(bold=True, size=14, color="111827")
    ws.cell(1, 1).alignment = Alignment(vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1).value = "윗줄은 수행평가/관찰내용명, 아랫줄은 평가 요소명/입력구분입니다. 성취수준 칸은 웹앱에서 설정한 코드 기준으로 입력하고, 추가 코멘트 칸은 직접 작성하세요."
    ws.cell(2, 1).font = Font(size=10, color="6B7280")
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)

    for col_idx, header in enumerate(base_headers, start=1):
        ws.cell(start_row, col_idx).value = header
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=start_row + 1, end_column=col_idx)
    ws.column_dimensions["A"].hidden = True

    # 선택목록
    item_level_ranges = {}
    for idx, item in enumerate(items, start=1):
        col_idx = idx
        list_ws.cell(1, col_idx).value = item.get("item_id", "")
        levels = [clean_text(x) for x in item.get("levels", []) if clean_text(x)]
        for row_idx, level in enumerate(levels, start=2):
            list_ws.cell(row_idx, col_idx).value = level
        if levels:
            col_letter = get_column_letter(col_idx)
            item_level_ranges[item.get("item_id", "")] = f"'선택목록'!${col_letter}$2:${col_letter}${len(levels) + 1}"

    for offset, field in enumerate(fields):
        item = field["item"]
        col_idx = item_start_col + offset
        assessment_name = get_assessment_name(item.get("assessment_id", ""))
        field_label = "성취수준" if field["field"] == "level" else "추가 코멘트"
        ws.cell(start_row, col_idx).value = assessment_name
        ws.cell(start_row + 1, col_idx).value = f"{item.get('name', '')} / {field_label}"
        ws.cell(3, col_idx).value = field["column"]
    ws.row_dimensions[3].hidden = True

    # 같은 수행평가/관찰내용명은 가로 병합
    if fields:
        group_start = item_start_col
        previous = clean_text(ws.cell(start_row, item_start_col).value)
        for col_idx in range(item_start_col + 1, item_start_col + len(fields) + 1):
            current = clean_text(ws.cell(start_row, col_idx).value) if col_idx < item_start_col + len(fields) else "__END__"
            if current != previous:
                if previous and col_idx - group_start > 1:
                    ws.merge_cells(start_row=start_row, start_column=group_start, end_row=start_row, end_column=col_idx - 1)
                group_start = col_idx
                previous = current

    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    student_fill = PatternFill("solid", fgColor="E5E7EB")
    assessment_fill = PatternFill("solid", fgColor="DBEAFE")
    item_fill = PatternFill("solid", fgColor="FFEDD5")
    body_fill = PatternFill("solid", fgColor="F9FAFB")
    level_fill = PatternFill("solid", fgColor="FFF7ED")

    for row_idx in [start_row, start_row + 1]:
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color="111827")
            if col_idx < item_start_col:
                cell.fill = student_fill
            elif row_idx == start_row:
                cell.fill = assessment_fill
                cell.font = Font(bold=True, color="1E3A8A")
            else:
                cell.fill = item_fill
                cell.font = Font(bold=True, color="92400E")

    for row_offset, (_, student) in enumerate(students.iterrows()):
        row_idx = data_start_row + row_offset
        sid = student.get("student_id", "")
        values = [sid, student.get("학년", ""), student.get("반", ""), student.get("번호", ""), student.get("성명", "")]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.border = border
            cell.fill = body_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for offset, field in enumerate(fields):
            item = field["item"]
            col_idx = item_start_col + offset
            rec = get_record(sid, item.get("item_id", ""))
            cell = ws.cell(row_idx, col_idx)
            cell.value = rec.get(field["field"], "")
            cell.border = border
            cell.fill = level_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "F6"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    for col_idx in range(item_start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_final_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "최종생기부"
    ws.sheet_properties.tabColor = "FF2563EB"

    headers = ["학년", "반", "번호", "성명", "최종 생기부", "byte", "생성일시", "생성 원문", "API 입력자료"]
    ws.append(headers)
    students = st.session_state.mid_students.copy()
    for _, student in students.iterrows():
        sid = student.get("student_id", "")
        result = st.session_state.mid_results.get(sid, {})
        final_text = safe_result_text(result)
        ws.append([
            student.get("학년", ""),
            student.get("반", ""),
            student.get("번호", ""),
            student.get("성명", ""),
            final_text,
            byte_count(final_text),
            result.get("created_at", ""),
            result.get("generated", ""),
            result.get("material", ""),
        ])

    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10, color="111827")
    ws.freeze_panes = "E2"
    ws.sheet_view.showGridLines = False
    widths = {"A": 8, "B": 8, "C": 8, "D": 14, "E": 72, "F": 10, "G": 20, "H": 56, "I": 72}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 학생별 성취수준 시트: 수행평가/관찰내용명/평가 요소명 2단 헤더
    ws2 = wb.create_sheet("학생별성취수준")
    items = all_used_items()
    fields = build_input_fields(items)
    base_headers = ["student_id", "학년", "반", "번호", "성명"]
    for col_idx, header in enumerate(base_headers, start=1):
        ws2.cell(1, col_idx).value = header
        ws2.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
    start_col = 6
    for offset, field in enumerate(fields):
        item = field["item"]
        col_idx = start_col + offset
        field_label = "성취수준" if field["field"] == "level" else "추가 코멘트"
        ws2.cell(1, col_idx).value = get_assessment_name(item.get("assessment_id", ""))
        ws2.cell(2, col_idx).value = f"{item.get('name', '')} / {field_label}"
    if fields:
        group_start = start_col
        previous = clean_text(ws2.cell(1, start_col).value)
        for col_idx in range(start_col + 1, start_col + len(fields) + 1):
            current = clean_text(ws2.cell(1, col_idx).value) if col_idx < start_col + len(fields) else "__END__"
            if current != previous:
                if previous and col_idx - group_start > 1:
                    ws2.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=col_idx - 1)
                group_start = col_idx
                previous = current
    for row_offset, (_, student) in enumerate(students.iterrows(), start=3):
        sid = student.get("student_id", "")
        values = [sid, student.get("학년", ""), student.get("반", ""), student.get("번호", ""), student.get("성명", "")]
        for col_idx, value in enumerate(values, start=1):
            ws2.cell(row_offset, col_idx).value = value
        for offset, field in enumerate(fields):
            item = field["item"]
            rec = get_record(sid, item.get("item_id", ""))
            ws2.cell(row_offset, start_col + offset).value = rec.get(field["field"], "")
    ws2.column_dimensions["A"].hidden = True
    ws2.freeze_panes = "F3"
    ws2.sheet_view.showGridLines = False
    for row in ws2.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row <= 2:
                cell.font = Font(bold=True, color="111827")
                cell.fill = PatternFill("solid", fgColor="DBEAFE" if cell.column >= 6 else "E5E7EB")
    for col_idx in range(1, max(6 + len(fields), 6)):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output



def render_level_header_preview(items):
    """② 입력표 위에서 수행평가/관찰내용-평가 요소 구조를 색상으로 확인한다.
    실제 입력표 헤더는 평가 요소명만 표시하고, 이 구조표에서 상위 수행평가/관찰내용을 확인한다.
    """
    items = list(items or [])
    if not items:
        return

    assessment_groups = []
    current_name = None
    current_items = []

    for item in items:
        assessment_name = get_assessment_name(item.get("assessment_id", "")) or "수행평가/관찰내용명 미입력"
        if current_name is None:
            current_name = assessment_name
        if assessment_name != current_name:
            assessment_groups.append((current_name, current_items))
            current_name = assessment_name
            current_items = []
        current_items.append(item)
    if current_name is not None:
        assessment_groups.append((current_name, current_items))

    cards = []
    for assessment_index, (assessment_name, group_items) in enumerate(assessment_groups, start=1):
        item_chips = "".join([
            '<span class="mid-structure-item-chip">'
            f'🧾 {html.escape(clean_text(item.get("name", "이름 없는 평가 요소")))}'
            '</span>'
            for item in group_items
        ])
        cards.append(
            '<div class="mid-structure-card">'
            '<div class="mid-structure-assessment">'
            f'<span class="mid-structure-number">수행평가/관찰내용 {assessment_index}</span>'
            f'<span>📁 {html.escape(assessment_name)}</span>'
            '</div>'
            f'<div class="mid-structure-items">{item_chips}</div>'
            '</div>'
        )

    html_block = (
        '<div class="mid-color-guide">'
        '<span class="mid-color-chip"><span class="mid-blue-dot"></span>수행평가/관찰내용명</span>'
        '<span class="mid-color-chip"><span class="mid-yellow-dot"></span>평가 요소명</span>'
        '</div>'
        '<div class="mid-structure-wrap">'
        + "".join(cards) +
        '</div>'
    )
    st.markdown(html_block, unsafe_allow_html=True)

# =========================
# UI 스타일 / 이동
# =========================
st.markdown('<div id="mid-honey-top"></div>', unsafe_allow_html=True)
st.title(MID_APP_TITLE)
st.caption(MID_APP_SUBTITLE)

st.markdown(
    """
    <style>
    div[data-testid="stRadio"] > div[role="radiogroup"] {
        display: flex !important;
        flex-wrap: wrap !important;
        gap: 0 !important;
        border-bottom: 1px solid #D0D7DE !important;
        margin-bottom: 1.15rem !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] label {
        min-height: 42px !important;
        padding: 0.62rem 0.95rem !important;
        margin-right: 4px !important;
        margin-bottom: -1px !important;
        border: 1px solid #D0D7DE !important;
        border-bottom: none !important;
        border-radius: 10px 10px 0 0 !important;
        background: #F6F8FA !important;
        cursor: pointer !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] label:has(input:checked) {
        background: #FFFFFF !important;
        border-top: 3px solid #D92D20 !important;
        color: #111827 !important;
        font-weight: 800 !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] label p {
        font-weight: 700 !important;
        font-size: 0.95rem !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] input {
        display: none !important;
    }
    /* app.py와 같은 색상 규칙: 수행평가/관찰내용은 파란색, 평가 요소는 노란색/주황색 */
    /* 하위 박스: 평가 요소는 노란색/주황색 계열 */
    div[data-testid="stExpander"] details:has(.mid-item-card-content) {
        background: linear-gradient(180deg, #FFFBF5 0%, #FFF7ED 100%) !important;
        border: 2px solid #FED7AA !important;
        border-radius: 16px !important;
        box-shadow: 0 4px 12px rgba(234, 88, 12, 0.07) !important;
        padding: 0.12rem 0.3rem 0.32rem 0.3rem !important;
        margin: 0.65rem 0 0.95rem 0 !important;
    }
    div[data-testid="stExpander"] details:has(.mid-item-card-content) > summary {
        background: #FFEDD5 !important;
        border: 1px solid #FED7AA !important;
        border-radius: 12px !important;
        margin: 0.16rem 0 0.5rem 0 !important;
        padding: 0.12rem 0.4rem !important;
    }
    div[data-testid="stExpander"] details:has(.mid-item-card-content) > summary p {
        color: #111827 !important;
        font-weight: 800 !important;
    }

    /* 상위 박스: 수행평가/관찰내용은 파란색 계열. 아래 선언이 나중에 와야 중첩 expander에서도 파란색이 유지된다. */
    div[data-testid="stExpander"] details:has(.mid-assessment-card-content) {
        background: linear-gradient(180deg, #EFF6FF 0%, #DBEAFE 100%) !important;
        border: 2px solid #93C5FD !important;
        border-radius: 18px !important;
        box-shadow: 0 5px 15px rgba(37, 99, 235, 0.10) !important;
        padding: 0.18rem 0.38rem 0.42rem 0.38rem !important;
        margin: 0.9rem 0 1.25rem 0 !important;
    }
    div[data-testid="stExpander"] details:has(.mid-assessment-card-content) > summary {
        background: #DBEAFE !important;
        border: 1px solid #BFDBFE !important;
        border-radius: 14px !important;
        margin: 0.2rem 0 0.55rem 0 !important;
        padding: 0.15rem 0.45rem !important;
    }
    div[data-testid="stExpander"] details:has(.mid-assessment-card-content) > summary p {
        color: #0F172A !important;
        font-weight: 900 !important;
    }
    .mid-assessment-card-content, .mid-item-card-content { display: none !important; }

    .mid-header-preview-wrap {
        width: 100%;
        overflow-x: auto;
        border: 1px solid #D1D5DB;
        border-radius: 14px;
        margin: 0.6rem 0 0 0;
        background: #FFFFFF;
        border-bottom: none;
        border-radius: 14px 14px 0 0;
    }
    table.mid-header-preview {
        border-collapse: separate;
        border-spacing: 0;
        min-width: 760px;
        width: 100%;
        table-layout: fixed;
        font-size: 0.92rem;
    }
    table.mid-header-preview th {
        border-right: 1px solid #D1D5DB;
        border-bottom: 1px solid #D1D5DB;
        padding: 9px 10px;
        text-align: center;
        vertical-align: middle;
        line-height: 1.35;
        word-break: keep-all;
    }
    table.mid-header-preview th.mid-student-head {
        background: #E5E7EB;
        color: #111827;
        font-weight: 800;
        width: 72px;
    }
    table.mid-header-preview th.mid-assessment-head {
        background: #DBEAFE;
        color: #1E3A8A;
        font-weight: 900;
    }
    table.mid-header-preview th.mid-item-head {
        background: #FFEDD5;
        color: #92400E;
        font-weight: 900;
    }
    .mid-color-guide {
        display: flex;
        gap: 8px;
        flex-wrap: wrap;
        margin: 0.25rem 0 0.35rem 0;
        font-size: 0.9rem;
    }
    .mid-color-chip {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        border: 1px solid #D1D5DB;
        border-radius: 999px;
        padding: 5px 10px;
        background: #FFFFFF;
        font-weight: 800;
    }
    .mid-blue-dot, .mid-yellow-dot {
        width: 12px;
        height: 12px;
        border-radius: 999px;
        display: inline-block;
    }
    .mid-blue-dot { background: #DBEAFE; border: 1px solid #93C5FD; }
    .mid-yellow-dot { background: #FFEDD5; border: 1px solid #FED7AA; }
    .mid-structure-wrap {
        display: flex;
        flex-direction: column;
        gap: 10px;
        margin: 0.45rem 0 1.0rem 0;
    }
    .mid-structure-card {
        border: 1px solid #BFDBFE;
        border-radius: 14px;
        overflow: hidden;
        background: #FFFFFF;
        box-shadow: 0 1px 2px rgba(17, 24, 39, 0.04);
    }
    .mid-structure-assessment {
        display: flex;
        align-items: center;
        gap: 10px;
        background: #DBEAFE;
        color: #1E3A8A;
        border-bottom: 1px solid #BFDBFE;
        padding: 10px 12px;
        font-weight: 900;
    }
    .mid-structure-number {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        border: 1px solid #93C5FD;
        border-radius: 999px;
        background: #EFF6FF;
        color: #1E3A8A;
        padding: 3px 9px;
        font-size: 0.82rem;
        font-weight: 900;
        white-space: nowrap;
    }
    .mid-structure-items {
        display: flex;
        flex-wrap: wrap;
        gap: 8px;
        padding: 10px 12px;
        background: #FFFBF5;
    }
    .mid-structure-item-chip {
        display: inline-flex;
        align-items: center;
        border: 1px solid #FED7AA;
        border-radius: 999px;
        background: #FFEDD5;
        color: #92400E;
        padding: 6px 11px;
        font-weight: 900;
        line-height: 1.25;
    }
.mid-header-preview-wrap + div[data-testid="stDataFrame"],
    .mid-header-preview-wrap + div[data-testid="stDataEditor"] {
        margin-top: 0 !important;
    }
    div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] {
        margin-top: 0 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("작업 관리")

    uploaded_project = st.file_uploader(
        "프로젝트 JSON 불러오기",
        type=["json"],
        help="중학교 간편 생기부 프로젝트 파일을 다시 불러옵니다.",
    )
    if uploaded_project and st.button("프로젝트 불러오기"):
        load_mid_project_json(uploaded_project)
        st.success("프로젝트를 불러왔습니다.")
        st.rerun()

    st.download_button(
        "현재 프로젝트 JSON 저장",
        data=mid_project_to_json(),
        file_name="middle_개꿀생기부_project.json",
        mime="application/json",
    )

    st.divider()

    if st.button("샘플 데이터 불러오기", help="중학교 간편 생기부 샘플 프로젝트를 불러옵니다."):
        load_mid_sample_data()
        st.success("샘플 데이터를 불러왔습니다.")
        st.rerun()

    if st.button("전체 초기화", type="secondary"):
        for key in [
            "mid_settings",
            "mid_assessments",
            "mid_items",
            "mid_students",
            "mid_records",
            "mid_results",
            "mid_selected_result_student_id",
            "mid_generation_job",
            "mid_current_step",
        ]:
            if key in st.session_state:
                del st.session_state[key]
        init_mid_state()
        sanitize_mid_state()
        st.success("초기화했습니다.")
        st.rerun()


STEP_LABELS = [
    "① 기본 설정/수행평가·관찰내용 설계",
    "② 학생별 성취수준 입력",
    "③ 생기부 생성/다운로드",
]
NAV_WIDGET_KEY = "mid_step_nav_radio_v05"
PENDING_STEP_KEY = "mid_pending_step_index_v05"
SCROLL_TO_TOP_KEY = "mid_scroll_to_top_after_step_change_v05"

if "mid_current_step" not in st.session_state:
    st.session_state.mid_current_step = 0

programmatic_step_change = False
if PENDING_STEP_KEY in st.session_state:
    try:
        st.session_state.mid_current_step = int(st.session_state[PENDING_STEP_KEY])
    except Exception:
        st.session_state.mid_current_step = 0
    del st.session_state[PENDING_STEP_KEY]
    st.session_state[SCROLL_TO_TOP_KEY] = True
    programmatic_step_change = True

st.session_state.mid_current_step = max(0, min(int(st.session_state.mid_current_step), len(STEP_LABELS) - 1))
if NAV_WIDGET_KEY not in st.session_state or programmatic_step_change:
    st.session_state[NAV_WIDGET_KEY] = STEP_LABELS[st.session_state.mid_current_step]


def request_step_change(next_index: int):
    st.session_state[PENDING_STEP_KEY] = int(next_index)


def scroll_page_to_top_once():
    if not st.session_state.get(SCROLL_TO_TOP_KEY, False):
        return
    st.session_state[SCROLL_TO_TOP_KEY] = False
    components.html(
        """
        <script>
        function forceScrollTop() {
            try {
                const parentWindow = window.parent;
                const parentDoc = parentWindow.document;
                try { parentWindow.scrollTo(0, 0); } catch (e) {}
                try { parentDoc.documentElement.scrollTop = 0; } catch (e) {}
                try { parentDoc.body.scrollTop = 0; } catch (e) {}
                parentDoc.querySelectorAll('section, main, div').forEach(function(el) {
                    try {
                        if (el.scrollHeight > el.clientHeight + 80) { el.scrollTop = 0; }
                    } catch (e) {}
                });
            } catch (e) {}
        }
        forceScrollTop();
        setTimeout(forceScrollTop, 80);
        setTimeout(forceScrollTop, 250);
        setTimeout(forceScrollTop, 600);
        </script>
        """,
        height=0,
    )


st.markdown("### 작업 단계")
selected_step_label = st.radio(
    "이동할 단계를 선택하세요.",
    STEP_LABELS,
    horizontal=True,
    label_visibility="collapsed",
    key=NAV_WIDGET_KEY,
)
st.session_state.mid_current_step = STEP_LABELS.index(selected_step_label)
current_step = st.session_state.mid_current_step
scroll_page_to_top_once()


def render_next_step_button(current_index: int):
    if current_index >= len(STEP_LABELS) - 1:
        return
    next_index = current_index + 1
    next_label = STEP_LABELS[next_index]
    st.divider()
    st.button(
        f"다음 단계로 넘어가기 → {next_label}",
        type="primary",
        use_container_width=True,
        key=f"mid_next_step_button_{current_index}",
        on_click=request_step_change,
        args=(next_index,),
    )


# =========================
# ① 기본 설정/수행평가·관찰내용 설계
# =========================
if current_step == 0:
    st.subheader("① 기본 설정 / 수행평가·관찰내용 설계")

    settings = st.session_state.mid_settings

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        settings["school_year"] = st.text_input("학년도", value=clean_text(settings.get("school_year", "2026")))
    with col2:
        settings["semester"] = st.selectbox(
            "학기",
            ["1학기", "2학기"],
            index=1 if settings.get("semester") == "2학기" else 0,
        )
    with col3:
        settings["grade"] = st.text_input("학년", value=clean_text(settings.get("grade", "2")))
    with col4:
        settings["subject"] = st.text_input("과목명", value=clean_text(settings.get("subject", "과학")))

    col5, col6 = st.columns(2)
    with col5:
        settings["target_bytes_min"] = st.number_input(
            "목표 최소 byte",
            min_value=100,
            max_value=2000,
            value=int(settings.get("target_bytes_min", 250)),
            step=50,
        )
    with col6:
        settings["target_bytes_max"] = st.number_input(
            "목표 최대 byte",
            min_value=100,
            max_value=2000,
            value=int(settings.get("target_bytes_max", 450)),
            step=50,
        )

    st.markdown("#### 생기부 작성 규칙")
    settings["custom_rules"] = st.text_area(
        "공통 작성 규칙",
        value=settings.get("custom_rules", MID_DEFAULT_RULES),
        height=150,
    )

    st.divider()
    st.markdown("### 📚 AI가 참고할 수행평가/관찰내용별 평가 자료")
    st.markdown(
        """
        이 화면은 **수행평가/관찰내용**을 먼저 만들고, 각 수행평가/관찰내용 안에 학생별로 입력할 **평가 요소**를 넣는 구조입니다.  
        평가 요소 안에는 필요에 따라 **성취수준 코드, 평가 문구, 개별 코멘트**를 설정합니다.
        """
    )

    with st.expander("➕ 새 수행평가/관찰내용 추가", expanded=True):
        st.caption("먼저 상위 단위인 수행평가/관찰내용을 만들고, 그 안에 평가 요소를 추가합니다.")
        with st.form("mid_add_assessment_form_v18"):
            col1, col2 = st.columns([2, 1])
            with col1:
                new_assessment_name = st.text_input("수행평가/관찰내용명", placeholder="예: 혼합물 분리 탐구")
                new_area = st.text_input("영역/단원", placeholder="예: 물질의 특성")
            with col2:
                new_use = st.checkbox("사용", value=True)
                st.caption("순서는 아래의 드래그 정렬에서 바꿀 수 있습니다.")
            new_desc = st.text_area(
                "성취기준 / 관찰 기준 / 활동 내용",
                placeholder="예: 혼합물의 분리 방법을 선택하고 실험 과정과 결과를 보고서로 정리하는 활동",
                height=80,
            )
            submitted = st.form_submit_button("수행평가/관찰내용 추가")
            if submitted:
                if not clean_text(new_assessment_name):
                    st.warning("수행평가/관찰내용명을 입력하세요.")
                else:
                    st.session_state.mid_assessments.append({
                        "assessment_id": make_id("mid_assess"),
                        "name": clean_text(new_assessment_name),
                        "area": clean_text(new_area),
                        "unit": clean_text(new_area),
                        "description": clean_text(new_desc),
                        "order": len(st.session_state.mid_assessments) + 1,
                        "use": bool(new_use),
                    })
                    sanitize_mid_state()
                    st.success("수행평가/관찰내용을 추가했습니다.")
                    st.rerun()

    if not st.session_state.mid_assessments:
        st.info("아직 등록된 수행평가/관찰내용이 없습니다. 먼저 수행평가/관찰내용을 추가하세요.")

    normalize_assessment_orders()
    assessments = sort_assessments()

    if len(assessments) >= 2:
        with st.expander("수행평가/관찰내용 순서 드래그 정렬", expanded=True):
            st.caption("수행평가/관찰내용명을 마우스로 잡고 위아래로 옮긴 뒤 저장하세요. 항목을 추가하면 이 박스가 자동으로 새로 생성됩니다.")
            assessment_labels = [
                f"{idx}. {assessment.get('name', '이름 없는 수행평가/관찰내용')}"
                for idx, assessment in enumerate(assessments, start=1)
            ]
            st.caption("현재 수행평가/관찰내용 순서: " + " → ".join(assessment_labels))
            label_to_assessment_id = {
                label: assessment.get("assessment_id", "")
                for label, assessment in zip(assessment_labels, assessments)
            }
            sorted_labels = sort_labels_with_gray_box(
                assessment_labels,
                key="mid_assessment_drag_sort",
                header="수행평가/관찰내용 순서",
            )
            if st.button("수행평가/관찰내용 순서 저장", key="mid_save_assessment_drag_sort"):
                apply_assessment_drag_order(sorted_labels, label_to_assessment_id)
                normalize_assessment_orders()
                st.success("수행평가/관찰내용 순서를 저장했습니다.")
                st.rerun()

    rubric_updates = {}

    for assess_index, assessment in enumerate(assessments, start=1):
        aid = assessment.get("assessment_id", "")
        normalize_item_orders(aid)
        existing_items = sort_items_for_assessment(aid)
        status_badge = "사용" if assessment.get("use", True) else "미사용"
        area_text = clean_text(assessment.get("area", assessment.get("unit", ""))) or "영역/단원 미입력"
        item_count = len(existing_items)
        assessment_expander_title = (
            f"📁 수행평가/관찰내용 {assess_index}. {assessment.get('name', '이름 없는 수행평가/관찰내용')} "
            f"· 평가 요소 {item_count}개 · {status_badge}"
        )
        with st.expander(assessment_expander_title, expanded=True):
            st.markdown('<div class="mid-assessment-card-content"></div>', unsafe_allow_html=True)
            st.markdown(
                f"""
                ### 📁 수행평가/관찰내용 {assess_index}. {assessment.get('name', '이름 없는 수행평가/관찰내용')}
                **영역/단원:** {area_text} &nbsp;&nbsp;|&nbsp;&nbsp;
                **평가 요소:** {item_count}개 &nbsp;&nbsp;|&nbsp;&nbsp;
                **상태:** {status_badge}
                """
            )
            if assessment.get("description"):
                st.caption(f"설명/관찰 기준/활동 내용: {assessment.get('description', '')}")
            else:
                st.caption("설명/관찰 기준/활동 내용이 아직 입력되지 않았습니다.")

            with st.expander("⚙️ 수행평가/관찰내용 기본 정보 수정 / 삭제", expanded=False):
                col1, col2, col3 = st.columns([2, 2, 1])
                with col1:
                    assessment["name"] = st.text_input(
                        "수행평가/관찰내용명 수정",
                        value=assessment.get("name", ""),
                        key=f"mid_assess_name_{aid}",
                    )
                    new_area_value = st.text_input(
                        "영역/단원 수정",
                        value=clean_text(assessment.get("area", assessment.get("unit", ""))),
                        key=f"mid_assess_area_{aid}",
                    )
                    assessment["area"] = clean_text(new_area_value)
                    assessment["unit"] = clean_text(new_area_value)
                with col2:
                    assessment["description"] = st.text_area(
                        "설명/관찰 기준/활동 내용 수정",
                        value=assessment.get("description", ""),
                        key=f"mid_assess_desc_{aid}",
                        height=120,
                    )
                with col3:
                    st.caption("순서는 수행평가/관찰내용 목록 위의 드래그 정렬에서 변경합니다.")
                    assessment["use"] = st.checkbox(
                        "사용",
                        value=assessment.get("use", True),
                        key=f"mid_assess_use_{aid}",
                    )
                    if st.button("수행평가/관찰내용 삭제", key=f"mid_delete_assessment_{aid}"):
                        item_ids = [item.get("item_id", "") for item in get_items_for_assessment(aid)]
                        st.session_state.mid_assessments = [a for a in st.session_state.mid_assessments if a.get("assessment_id", "") != aid]
                        st.session_state.mid_items = [item for item in st.session_state.mid_items if item.get("assessment_id", "") != aid]
                        st.session_state.mid_records = {k: v for k, v in as_dict(st.session_state.mid_records).items() if k.split("::")[-1] not in item_ids}
                        st.success("수행평가/관찰내용을 삭제했습니다.")
                        st.rerun()

            st.markdown("#### 🧾 평가 요소")
            if existing_items:
                if len(existing_items) == 1:
                    st.caption("평가 요소가 1개라서 드래그 정렬 박스는 표시하지 않습니다. 평가 요소를 2개 이상 만들면 이곳에 드래그 정렬 박스가 나타납니다.")
                if len(existing_items) >= 2:
                    with st.expander("평가 요소 순서 드래그 정렬", expanded=True):
                        st.caption("평가 요소명을 마우스로 잡고 위아래로 옮긴 뒤 저장하세요. 항목을 추가하면 이 박스가 자동으로 새로 생성됩니다.")
                        item_labels = [
                            f"{idx}. {item.get('name', '이름 없는 평가 요소')}"
                            for idx, item in enumerate(existing_items, start=1)
                        ]
                        st.caption("현재 평가 요소 순서: " + " → ".join(item_labels))
                        label_to_item_id = {
                            label: item.get("item_id", "")
                            for label, item in zip(item_labels, existing_items)
                        }
                        sorted_item_labels = sort_labels_with_gray_box(
                            item_labels,
                            key=f"mid_item_drag_sort_{aid}",
                            header="평가 요소 순서",
                        )
                        if st.button("평가 요소 순서 저장", key=f"mid_save_item_drag_sort_{aid}"):
                            apply_item_drag_order(aid, sorted_item_labels, label_to_item_id)
                            normalize_item_orders(aid)
                            st.success("평가 요소 순서를 저장했습니다.")
                            st.rerun()

                for item_index, item in enumerate(existing_items, start=1):
                    item_id = item.get("item_id", "")
                    item_type_label = item_type_to_kor(item.get("type", "rubric"))
                    item_expander_title = (
                        f"🧾 평가 요소 {item_index}. {item.get('name', '이름 없는 평가 요소')} "
                        f"· {item_type_label}"
                    )
                    with st.expander(item_expander_title, expanded=True):
                        st.markdown('<div class="mid-item-card-content"></div>', unsafe_allow_html=True)
                        st.markdown(f"##### 🧾 평가 요소 {item_index}. {item.get('name', '이름 없는 평가 요소')}")
                        col1, col2, col3 = st.columns([2.2, 1.6, 1])
                        with col1:
                            item["name"] = st.text_input(
                                "항목명",
                                value=item.get("name", ""),
                                key=f"mid_item_name_{item_id}",
                            )
                        with col2:
                            type_options = ["성취도 선택형", "개별 코멘트형", "성취도 + 추가 코멘트형"]
                            current_type_label = item_type_to_kor(item.get("type", "rubric"))
                            new_type_label = st.selectbox(
                                "기록 방식",
                                type_options,
                                index=type_options.index(current_type_label) if current_type_label in type_options else 0,
                                key=f"mid_item_type_{item_id}",
                            )
                            new_type = item_type_from_kor(new_type_label)
                            if new_type != item.get("type"):
                                item["type"] = new_type
                                if new_type == "comment":
                                    item["levels"] = []
                                    item["rubrics"] = {}
                                elif not item.get("levels"):
                                    item["levels"] = ["A", "B", "C", "D", "E"]
                                    item["rubrics"] = {level: "" for level in item["levels"]}
                        with col3:
                            st.caption(f"현재 {item_index}번째")
                            if st.button("평가 요소 삭제", key=f"mid_delete_item_{item_id}"):
                                st.session_state.mid_items = [x for x in st.session_state.mid_items if x.get("item_id", "") != item_id]
                                st.session_state.mid_records = {k: v for k, v in as_dict(st.session_state.mid_records).items() if not k.endswith(f"::{item_id}")}
                                normalize_item_orders(aid)
                                st.success("평가 요소를 삭제했습니다.")
                                st.rerun()

                        if item.get("type", "rubric") in ["rubric", "rubric_plus"]:
                            levels, rubrics = render_rubric_input_block(
                                prefix=f"mid_edit_item_rubric_{item_id}",
                                current_levels=item.get("levels", []),
                                current_rubrics=item.get("rubrics", {}),
                            )
                            rubric_updates[item_id] = {"levels": levels, "rubrics": rubrics}
                            st.caption("성취수준/평가 문구는 화면 맨 아래의 전체 저장 버튼으로 한꺼번에 저장됩니다.")
                        else:
                            st.info("개별 코멘트형 항목입니다. 학생별 기록 입력 화면에서 학생별 서술형 코멘트를 입력합니다.")
            else:
                st.info("아직 이 수행평가/관찰내용에 등록된 평가 요소가 없습니다. 아래의 '평가 요소 추가'를 눌러 항목을 추가하세요.")

            st.divider()
            with st.expander("➕ 이 수행평가/관찰내용에 평가 요소 추가", expanded=(item_count == 0)):
                st.markdown(
                    """
                    수행평가/관찰내용 안에 존재하는 여러 관찰 및 평가 요소들을 추가해주세요.  
                    A, B, C와 같은 **성취도 선택형**으로 개별화시킬 수도 있고, 개인마다 다른 관찰 내용을 적어주는 **개별 코멘트형**으로 더욱 구체적인 개별화가 가능합니다.  
                    또한 이 두 가지를 융합한 **성취도 + 추가 코멘트형**도 가능합니다.
                    """
                )
                item_name = st.text_input("평가 요소명", placeholder="예: 실험 과정", key=f"mid_new_item_name_{aid}")
                item_type_label = st.selectbox(
                    "기록 방식",
                    ["성취도 선택형", "개별 코멘트형", "성취도 + 추가 코멘트형"],
                    key=f"mid_new_item_type_{aid}",
                    help="개별 코멘트형을 선택하면 성취수준 코드와 루브릭 입력칸이 사라집니다.",
                )
                item_type = item_type_from_kor(item_type_label)
                levels, rubrics = [], {}
                if item_type != "comment":
                    levels, rubrics = render_rubric_input_block(
                        prefix=f"mid_new_item_rubric_{aid}",
                        current_levels=["A", "B", "C", "D", "E"],
                        current_rubrics={
                            "A": "우수한 수준으로 수행함",
                            "B": "대체로 적절하게 수행함",
                            "C": "일부 보완이 필요함",
                            "D": "기본적인 참여가 이루어짐",
                            "E": "지속적인 보완이 필요함",
                        },
                    )
                else:
                    st.info("개별 코멘트형입니다. 성취수준 코드 없이 학생별 서술형 코멘트만 입력합니다.")

                if st.button("이 수행평가/관찰내용에 평가 요소 추가", key=f"mid_add_item_button_{aid}"):
                    if not clean_text(item_name):
                        st.warning("평가 요소명을 입력하세요.")
                    else:
                        st.session_state.mid_items.append({
                            "item_id": make_id("mid_item"),
                            "assessment_id": aid,
                            "name": clean_text(item_name),
                            "type": item_type,
                            "levels": levels if item_type != "comment" else [],
                            "rubrics": rubrics if item_type != "comment" else {},
                            "order": len(get_items_for_assessment(aid)) + 1,
                        })
                        normalize_item_orders(aid)
                        sanitize_mid_state()
                        st.success("평가 요소를 추가했습니다.")
                        st.rerun()

    if rubric_updates:
        st.divider()
        if st.button("전체 성취수준/평가 문구 한꺼번에 저장", type="primary", use_container_width=True):
            saved_count = 0
            for item in st.session_state.mid_items:
                item_id = item.get("item_id", "")
                if item_id in rubric_updates:
                    item["levels"] = rubric_updates[item_id]["levels"]
                    item["rubrics"] = rubric_updates[item_id]["rubrics"]
                    saved_count += 1
            sanitize_mid_state()
            st.success(f"성취수준/평가 문구를 {saved_count}개 평가 요소에 한꺼번에 저장했습니다.")
            st.rerun()

    render_next_step_button(0)


# =========================
# ② 학생별 성취수준 입력
# =========================
if current_step == 1:
    st.subheader("② 학생별 성취수준 입력")

    items = all_used_items()
    if not items:
        st.warning("먼저 ①에서 수행평가/관찰내용과 평가 요소를 추가하세요.")
    else:
        st.markdown(
            """
            학생은 행으로, 평가 요소는 열로 입력합니다.  
            위에는 ①에서 설정한 **수행평가/관찰내용-평가 요소 구조**를 색으로 보여주고, 실제 입력표에는 성취수준과 추가 코멘트를 기록 방식에 맞게 표시합니다. 이름은 관리용이며 AI 프롬프트에는 들어가지 않습니다.
            """
        )

        matrix_df, items, fields = build_record_matrix_df()
        visible_df = matrix_df.copy()

        st.markdown("#### 수행평가/관찰내용-평가 요소 구조")
        render_level_header_preview(items)

        column_config = {
            "student_id": None,
            "학년": st.column_config.TextColumn("학년", width="small"),
            "반": st.column_config.TextColumn("반", width="small"),
            "번호": st.column_config.TextColumn("번호", width="small"),
            "성명": st.column_config.TextColumn("성명", width="medium"),
        }
        for field in fields:
            item = field["item"]
            assessment_name = get_assessment_name(item.get("assessment_id", "")) or "수행평가/관찰내용명 미입력"
            item_name = clean_text(item.get("name", "")) or "평가 요소명 미입력"
            label = field.get("label", item_name)
            if field["field"] == "level":
                levels = [clean_text(x) for x in item.get("levels", []) if clean_text(x)]
                column_config[field["column"]] = st.column_config.SelectboxColumn(
                    label,
                    options=[""] + levels,
                    required=False,
                    width="medium",
                    help=f"{assessment_name} / {item_name} / 성취수준",
                )
            else:
                column_config[field["column"]] = st.column_config.TextColumn(
                    label,
                    width="large",
                    help=f"{assessment_name} / {item_name} / 추가 코멘트",
                )

        st.markdown("#### 학생별 성취수준/추가 코멘트 입력표")
        st.caption("성취도 선택형은 성취수준만, 개별 코멘트형은 추가 코멘트만, 성취도+추가 코멘트형은 두 칸이 표시됩니다.")

        edited_df = st.data_editor(
            visible_df,
            num_rows="dynamic",
            use_container_width=True,
            height=560,
            key="mid_record_matrix_editor_v18",
            column_config=column_config,
        )

        col_save, col_download = st.columns([1.4, 2.0])
        with col_save:
            if st.button("학생별 성취수준 저장", type="primary", use_container_width=True):
                saved_count = save_record_matrix_df(edited_df, items, fields)
                st.success(f"학생별 기록을 저장했습니다. 저장된 입력값: {saved_count}개")
                st.rerun()
        with col_download:
            st.download_button(
                "학생별 기록 입력표 엑셀 다운로드",
                data=make_record_input_excel(),
                file_name=f"middle_student_levels_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with st.expander("표 헤더 구조 설명", expanded=False):
            st.markdown(
                "웹 입력표의 성취수준 열 제목에는 평가 요소명만 표시합니다. "
                "대신 입력표 바로 위에 ①에서 설정한 수행평가/관찰내용명과 평가 요소명을 구조화해서 보여줍니다. "
                "엑셀 다운로드 파일에서는 윗줄 수행평가/관찰내용명, 아랫줄 평가 요소명의 2단 헤더를 적용했습니다."
            )

    render_next_step_button(1)


# =========================
# ③ 생기부 생성/다운로드
# =========================
if current_step == 2:
    st.subheader("③ 생기부 생성 / 수정 / 다운로드")

    if st.session_state.mid_students.empty:
        st.warning("먼저 ②에서 학생별 성취수준을 입력하고 저장하세요.")
    elif not all_used_items():
        st.warning("먼저 ①에서 수행평가/관찰내용과 평가 요소를 추가하세요.")
    else:
        st.markdown("#### AI 선택 및 API 설정")
        col_provider, col_model, col_key, col_variation = st.columns([1.2, 1.7, 2.4, 1.2], gap="small")
        with col_provider:
            ai_provider = st.selectbox("사용할 AI", AI_PROVIDER_OPTIONS, index=0)
        with col_model:
            model = st.text_input(
                "모델명",
                value=get_default_ai_model(ai_provider),
                key=f"mid_ai_model_name_{ai_provider}_v06",
            )
        with col_key:
            api_key = st.text_input(
                f"{ai_provider} API Key",
                value=get_default_ai_key(ai_provider),
                type="password",
                key=f"mid_api_key_{ai_provider}_v06",
                help=f"Streamlit Secrets에는 {AI_SECRET_KEY_NAMES.get(ai_provider, 'OPENAI_API_KEY')} 이름으로 저장해둘 수 있습니다.",
            )
        with col_variation:
            variation_level = st.selectbox("변주 강도", ["낮음", "보통", "높음"], index=1)

        students = st.session_state.mid_students.copy()
        student_labels = {
            f"{row.get('학년', '')}학년 {row.get('반', '')}반 {row.get('번호', '')}번 {row.get('성명', '')}": row
            for _, row in students.iterrows()
        }
        selected_label = st.selectbox("생성할 학생", list(student_labels.keys()))
        selected_student = student_labels[selected_label]

        col_a, col_b, col_spacer = st.columns([1.65, 1.9, 5.45], gap="small")
        with col_a:
            if st.button("선택 학생 생기부 생성", type="secondary", use_container_width=True):
                overlay_loading_started_at = datetime.now()
                sid = selected_student.get("student_id", "")
                variant_no = list(students["student_id"].astype(str)).index(str(sid)) + 1 if sid in students["student_id"].astype(str).tolist() else 1
                overlay_slot = None
                recent_preview_items = generation_preview_items_from_results(
                    students, st.session_state.mid_results, exclude_sid=sid
                )
                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "선택 학생 생기부 생성 중",
                    f"{selected_label}의 평가 요소별 성취수준을 정리하고 있습니다.",
                    0.20,
                    ["성취수준 입력값 확인", "평가 요소별 평가 문구 연결", "AI 입력 자료 구성"],
                    recent_items=recent_preview_items,
                    loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                )
                material = build_student_material(selected_student)
                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "선택 학생 생기부 생성 중",
                    f"{selected_label}의 문장을 생성하고 있습니다.",
                    0.50,
                    ["개인정보 제외", "중학교용 간결 문체 적용", f"변주 강도: {variation_level}"],
                    recent_items=recent_preview_items,
                    loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                )
                prompt = build_prompt(material, variation_level=variation_level, variant_no=variant_no)
                generated = None
                if api_key:
                    generated = generate_with_ai(prompt, ai_provider, api_key, model)
                if not generated:
                    overlay_slot = show_generation_overlay(
                        overlay_slot,
                        "선택 학생 생기부 생성 중",
                        "API 결과가 없어 내부 조합 방식으로 문장을 구성하고 있습니다.",
                        0.75,
                        ["교사의 평가 문구 추출", "중복 표현 정리", "명사형 종결 적용"],
                        recent_items=recent_preview_items,
                        loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                    )
                    generated = fallback_generate(material, variant_no=variant_no)
                generated = normalize_sentence(generated)
                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "선택 학생 생기부 생성 완료 처리 중",
                    "생성 문장을 결과표와 수정창에 저장하고 있습니다.",
                    0.95,
                    ["결과 저장", "byte 계산", "수정창 반영"],
                    recent_items=recent_preview_items,
                    loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                )
                st.session_state.mid_results[sid] = {
                    "material": material,
                    "generated": generated,
                    "edited": generated,
                    "bytes": byte_count(generated),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                st.success("생성했습니다.")
                st.rerun()
        with col_b:
            mid_job = st.session_state.mid_generation_job
            if not mid_job.get("active", False):
                if st.button("전체 학생 생기부 생성 시작", type="primary", use_container_width=True):
                    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    st.session_state.mid_generation_job = {
                        "active": True,
                        "stop_requested": False,
                        "student_ids": students["student_id"].astype(str).tolist(),
                        "index": 0,
                        "log": [],
                        "ai_provider": ai_provider,
                        "api_key": api_key,
                        "model": model,
                        "variation_level": variation_level,
                        "started_at": now_text,
                        "loading_started_at": now_text,
                        "finished_at": "",
                    }
                    st.success("전체 생성 작업을 시작합니다.")
                    st.rerun()
            else:
                if st.button("생기부 생성 중지", type="primary", use_container_width=True):
                    mid_job["active"] = False
                    mid_job["stop_requested"] = True
                    mid_job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    st.session_state.mid_generation_job = mid_job
                    st.warning("생성 중지를 요청했습니다. 이미 생성된 학생 결과는 저장되어 있습니다.")
                    st.rerun()

        mid_job = st.session_state.mid_generation_job
        if mid_job.get("active", False) or mid_job.get("log"):
            st.markdown("#### 전체 생성 진행 상황")
            total = len(mid_job.get("student_ids", []))
            done = int(mid_job.get("index", 0) or 0)
            if total > 0:
                st.progress(min(done / total, 1.0))
                st.caption(f"진행률: {done}/{total}명")

            if mid_job.get("active", False):
                st.info("전체 생성이 진행 중입니다. 위의 빨간색 '생기부 생성 중지' 버튼을 누르면 현재 처리 중인 학생 완료 후 멈춥니다.")
            elif mid_job.get("stop_requested", False):
                st.warning("전체 생성이 중지되었습니다. 중지 전까지 생성된 문구는 아래 결과표와 다운로드 엑셀에 반영됩니다.")
            elif mid_job.get("log"):
                st.success("전체 생성 작업이 완료되었습니다.")

            if mid_job.get("log"):
                latest_log = mid_job.get("log", [])[-1]
                latest_label = f"{latest_log.get('학년', '')}학년 {latest_log.get('반', '')}반 {latest_log.get('번호', '')}번 {latest_log.get('성명', '')}"
                st.caption(f"최근 생성 완료: {latest_label}")

        if mid_job.get("active", False) and not mid_job.get("stop_requested", False):
            student_ids = mid_job.get("student_ids", [])
            index = int(mid_job.get("index", 0) or 0)
            total = len(student_ids)

            if index >= total:
                mid_job["active"] = False
                mid_job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state.mid_generation_job = mid_job
                st.success("전체 학생 생기부 생성을 완료했습니다.")
                st.rerun()

            current_sid = clean_text(student_ids[index])
            matched = students[students["student_id"].astype(str) == current_sid]
            if matched.empty:
                mid_job["index"] = index + 1
                st.session_state.mid_generation_job = mid_job
                st.rerun()

            student = matched.iloc[0]
            label = f"{student.get('학년', '')}학년 {student.get('반', '')}반 {student.get('번호', '')}번 {student.get('성명', '')}"
            latest_preview = []
            if mid_job.get("log"):
                latest_log = mid_job.get("log", [])[-1]
                latest_label = f"{latest_log.get('학년', '')}학년 {latest_log.get('반', '')}반 {latest_log.get('번호', '')}번 {latest_log.get('성명', '')}"
                latest_preview = [{
                    "label": f"{latest_log.get('순서', '')}번 생성 · {latest_label}",
                    "text": latest_log.get("생성 문구", ""),
                }]

            overlay_slot = None
            loading_started_at = mid_job.get("loading_started_at") or mid_job.get("started_at")
            overlay_slot = show_generation_overlay(
                overlay_slot,
                "전체 학생 생기부 생성 중",
                f"{index + 1}/{total} 처리 중 · {label}",
                index / total if total else 0,
                ["평가 요소별 성취수준 확인", "AI 입력 자료 구성", "문장 생성", "결과 저장"],
                recent_items=latest_preview,
                loading_offset_seconds=loading_elapsed_seconds(loading_started_at),
            )

            material = build_student_material(student)
            prompt = build_prompt(
                material,
                variation_level=mid_job.get("variation_level", variation_level),
                variant_no=index + 1,
            )
            generated = None
            if mid_job.get("api_key"):
                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "전체 학생 생기부 생성 중",
                    f"{index + 1}/{total} · {label}의 문장을 AI가 생성하고 있습니다.",
                    (index + 0.55) / total if total else 0.55,
                    ["AI 응답 대기 중", "응답 후 결과표에 저장", "다음 학생으로 이동"],
                    recent_items=latest_preview,
                    loading_offset_seconds=loading_elapsed_seconds(loading_started_at),
                )
                generated = generate_with_ai(
                    prompt,
                    mid_job.get("ai_provider", ai_provider),
                    mid_job.get("api_key", ""),
                    mid_job.get("model", model),
                )
            if not generated:
                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "전체 학생 생기부 생성 중",
                    f"{index + 1}/{total} · {label}의 문장을 내부 조합 방식으로 구성하고 있습니다.",
                    (index + 0.75) / total if total else 0.75,
                    ["교사의 평가 문구 추출", "변주 표현 적용", "중학교 생기부 문체 정리"],
                    recent_items=latest_preview,
                    loading_offset_seconds=loading_elapsed_seconds(loading_started_at),
                )
                generated = fallback_generate(material, variant_no=index + 1)

            generated = normalize_sentence(generated)
            sid = student.get("student_id", "")
            st.session_state.mid_results[sid] = {
                "material": material,
                "generated": generated,
                "edited": generated,
                "bytes": byte_count(generated),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            mid_job.setdefault("log", []).append({
                "순서": index + 1,
                "학년": student.get("학년", ""),
                "반": student.get("반", ""),
                "번호": student.get("번호", ""),
                "성명": student.get("성명", ""),
                "생성 문구": generated,
                "byte": byte_count(generated),
                "상태": "저장 완료",
            })
            mid_job["index"] = index + 1

            if mid_job["index"] >= total:
                mid_job["active"] = False
                mid_job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state.mid_generation_job = mid_job
                st.success("전체 학생 생기부 생성을 완료했습니다.")
            else:
                st.session_state.mid_generation_job = mid_job
                st.rerun()

        st.divider()
        st.markdown("#### 전체 결과표 / 학생 선택")
        result_rows = []
        for _, student in students.iterrows():
            sid = student.get("student_id", "")
            result = st.session_state.mid_results.get(sid, {})
            text = safe_result_text(result)
            result_rows.append({
                "student_id": sid,
                "학년": student.get("학년", ""),
                "반": student.get("반", ""),
                "번호": student.get("번호", ""),
                "성명": student.get("성명", ""),
                "생성/수정 문구": text,
                "byte": byte_count(text),
                "생성일시": result.get("created_at", ""),
            })
        result_df = pd.DataFrame(result_rows)
        display_result_df = result_df.drop(columns=["student_id"], errors="ignore")

        if clean_text(st.session_state.get("mid_selected_result_student_id", "")) not in set(students["student_id"].astype(str)):
            st.session_state.mid_selected_result_student_id = selected_student.get("student_id", "")
        current_selected_sid = clean_text(st.session_state.mid_selected_result_student_id)

        selection_event = st.dataframe(
            display_result_df,
            use_container_width=True,
            height=340,
            hide_index=True,
            key="mid_generation_result_selector_v05",
            on_select="rerun",
            selection_mode="single-cell",
            column_config={
                "학년": st.column_config.TextColumn("학년", width="small"),
                "반": st.column_config.TextColumn("반", width="small"),
                "번호": st.column_config.TextColumn("번호", width="small"),
                "성명": st.column_config.TextColumn("성명", width="medium"),
                "생성/수정 문구": st.column_config.TextColumn("생성/수정 문구", width="large"),
                "byte": st.column_config.NumberColumn("byte", width="small"),
                "생성일시": st.column_config.TextColumn("생성일시", width="medium"),
            },
        )
        selected_row_index = None
        try:
            selected_cells = list(selection_event.selection.cells)
            if selected_cells:
                selected_row_index = int(selected_cells[-1][0])
        except Exception:
            selected_row_index = None
        if selected_row_index is not None and 0 <= selected_row_index < len(result_df):
            current_selected_sid = clean_text(result_df.iloc[selected_row_index].get("student_id", current_selected_sid))
            st.session_state.mid_selected_result_student_id = current_selected_sid

        selected_match = students[students["student_id"] == current_selected_sid]
        if selected_match.empty:
            selected_detail_student = selected_student
            current_selected_sid = selected_student.get("student_id", "")
        else:
            selected_detail_student = selected_match.iloc[0]
        selected_detail_label = f"{selected_detail_student.get('학년', '')}학년 {selected_detail_student.get('반', '')}반 {selected_detail_student.get('번호', '')}번 {selected_detail_student.get('성명', '')}"
        result = st.session_state.mid_results.get(current_selected_sid, {})
        initial_text = safe_result_text(result)

        st.markdown(f"#### 표에서 선택한 학생 수정: {selected_detail_label}")
        if not result:
            st.info("아직 이 학생의 생기부 문구가 생성되지 않았습니다. 직접 입력하거나 위에서 생성을 먼저 실행하세요.")
        editor_key = f"mid_generation_detail_text_v05_{current_selected_sid}"
        if editor_key not in st.session_state:
            st.session_state[editor_key] = initial_text
        edited = st.text_area("교사 수정 문구", key=editor_key, height=200)
        st.metric("현재 byte", byte_count(edited))
        if st.button("수정 문구 저장", type="primary"):
            result = st.session_state.mid_results.get(current_selected_sid, {})
            result["edited"] = edited
            result["bytes"] = byte_count(edited)
            if not result.get("generated"):
                result["generated"] = edited
            if not result.get("material"):
                result["material"] = build_student_material(selected_detail_student)
            if not result.get("created_at"):
                result["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.session_state.mid_results[current_selected_sid] = result
            st.success("수정 문구를 저장했습니다.")
            st.rerun()

        st.divider()
        st.markdown("#### 최종 결과 다운로드")
        st.caption("마지막 단계입니다. 수정 내용을 확인한 뒤 결과 엑셀을 내려받으세요.")
        st.markdown(
            """
            <style>
            section.main div[data-testid="stDownloadButton"] > button {
                min-height: 58px !important;
                font-size: 1.14rem !important;
                font-weight: 900 !important;
                border-radius: 14px !important;
                box-shadow: 0 4px 12px rgba(217, 45, 32, 0.18) !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        st.download_button(
            "📥 결과 엑셀 다운로드",
            data=export_final_excel(),
            file_name=f"middle_개꿀생기부_result_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
