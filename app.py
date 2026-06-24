import hashlib
import html
import json
import re
import uuid
from datetime import datetime
from io import BytesIO
from urllib import error as urllib_error
from urllib import parse as urllib_parse
from urllib import request as urllib_request

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from streamlit_sortables import sort_items
except Exception:
    sort_items = None


# =========================
# 앱 기본 설정
# =========================
st.set_page_config(
    page_title="개꿀 생기부",
    page_icon="🍯",
    layout="wide",
)

APP_TITLE = "🍯 개꿀 생기부 v62"
APP_SUBTITLE = "수행평가 기반 생기부 작성 도우미 · patched-20260624-v62"


DEFAULT_RULES = """- 명사형 종결을 사용한다. 예: 분석함, 정리함, 제시함, 탐색함.
- 학생 이름을 쓰지 않는다.
- 첫 문장을 '학생은'으로 시작하지 않는다.
- '깊은 이해', '창의융합', '혁신적' 같은 과장 표현을 피한다.
- 제공된 평가 자료에 근거한 내용만 작성한다.
- 한 문단으로 작성하고, 중복되는 내용은 자연스럽게 통합한다.
- 평가 자료에 없는 인성, 진로, 성격, 태도 내용을 임의로 추가하지 않는다."""


# 2026-06-21 공식 문서 확인 기준 기본 모델값.
# 사용자는 ⑥ 생성 화면에서 모델명을 직접 수정할 수 있다.
AI_PROVIDER_OPTIONS = ["ChatGPT", "Gemini", "Claude"]
AI_DEFAULT_MODELS = {
    "ChatGPT": "gpt-5.5",
    "Gemini": "gemini-3.1-pro-preview",
    "Claude": "claude-fable-5",
}
AI_SECRET_KEY_NAMES = {
    "ChatGPT": "OPENAI_API_KEY",
    "Gemini": "GEMINI_API_KEY",
    "Claude": "ANTHROPIC_API_KEY",
}


def get_default_ai_model(provider: str) -> str:
    provider = provider if provider in AI_DEFAULT_MODELS else "ChatGPT"
    return AI_DEFAULT_MODELS[provider]


def get_default_ai_key(provider: str) -> str:
    secret_name = AI_SECRET_KEY_NAMES.get(provider, "OPENAI_API_KEY")
    try:
        return st.secrets.get(secret_name, "")
    except Exception:
        return ""


MASTER_PROMPT = """너는 학교생활기록부 교과 세부능력 및 특기사항을 작성하는 교사 보조 도구이다.

[기본 역할]
- 제공된 평가 자료를 바탕으로 교과 세부능력 및 특기사항 문장을 작성한다.
- 학생의 학습 과정, 수행 수준, 탐구 활동의 특징이 드러나도록 작성한다.
- 교사가 관찰하고 기록하는 학교생활기록부 문체로 작성한다.

[근거 사용 원칙]
- 제공된 평가 자료만 근거로 사용한다.
- 평가 자료에 없는 내용은 추측하거나 임의로 추가하지 않는다.
- 성취수준 코드는 그대로 나열하지 않고, 성취수준에 연결된 교사의 평가 문구를 생활기록부 문장으로 자연스럽게 바꾸어 작성한다.
- 추가 코멘트가 있는 경우 교사의 평가 문구와 자연스럽게 연결하여 반영한다.
- 수행평가명, 평가 요소, 성취기준, 교사의 평가 문구가 서로 중복될 경우 같은 의미가 반복되지 않도록 통합한다.

[개인정보 및 표현 제한]
- 학생의 이름, 학년, 반, 번호, 학교명 등 개인을 식별할 수 있는 정보는 포함하지 않는다.
- 입력 자료에 개인정보가 있더라도 최종 문장에는 반영하지 않는다.
- 학생은, 이 학생은, 해당 학생은으로 문장을 시작하지 않는다.
- 근거 없는 인성 평가, 성격 판단, 진로 추정, 태도 평가는 작성하지 않는다.
- 과장된 표현이나 평가 자료보다 지나치게 확대된 표현은 사용하지 않는다.

[부정적 표현 회피]
- 부정적인 표현을 직접 사용하지 않고, 현재 수행한 내용과 앞으로 발전할 수 있는 방향을 중심으로 서술한다.
- 부족함, 미흡함, 노력이 필요함, 보완이 필요함, 성취 수준이 낮음, 소극적임과 같은 표현은 사용하지 않는다.
- 단순히 잘하지 못했다는 의미를 쓰기보다, 학생이 현재 수행한 내용과 앞으로 정교화할 수 있는 부분이 드러나도록 표현한다.
- 예시:
  - 개념 이해가 부족함 → 핵심 개념을 부분적으로 파악하였으며, 관련 사례와 연결해 이해를 넓혀 가고 있음.
  - 설명이 미흡함 → 주요 내용을 중심으로 설명하려는 모습을 보였으며, 근거를 구체화하면 설명의 완성도가 높아질 것으로 보임.
  - 탐구 과정이 부족함 → 탐구 과정의 기본 절차를 수행하였으며, 변인 통제와 결과 해석을 연결하는 경험을 쌓아 가고 있음.
  - 보완이 필요함 → 기본 개념을 바탕으로 활동에 참여하였고, 세부 근거를 추가하면 사고 과정이 더 분명하게 드러날 것으로 보임.
  - 성취 수준이 낮음 → 기초 내용을 바탕으로 과제를 수행하였고, 핵심 개념을 다양한 상황에 적용하는 과정에서 성장이 기대됨.

[작성 방식]
- 평가 자료에 제시된 활동 수행 내용과 교과 역량이 드러나도록 작성한다.
- 구체적이되 담백한 문장으로 작성한다.
- 문장 사이의 연결이 자연스럽도록 다듬는다.
- 한 문단으로 작성한다.
- 학교생활기록부에 바로 입력할 수 있는 완성된 문장으로 작성한다.
- 문장은 명사형 종결 어미를 사용하여 마무리한다.
- 명사형 종결 예시: 분석함, 정리함, 제시함, 설명함, 탐구함, 수행함, 비교함, 해석함, 적용함, 도출함, 확인함, 이해한 것으로 보임.

[출력 형식]
- 세부능력 및 특기사항 문장만 출력한다.
- 제목, 번호, 설명, 따옴표, 불필요한 안내 문구는 출력하지 않는다."""


# =========================
# 공통 유틸
# =========================
def make_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


def clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def byte_count(text: str) -> int:
    return len(clean_text(text).encode("utf-8"))


def loading_elapsed_seconds(started_at=None) -> float:
    """전체 생성 중 Streamlit이 rerun되어도 로딩 문구 순서가 처음으로 돌아가지 않게 경과 시간을 계산한다."""
    if started_at is None:
        return 0.0
    if isinstance(started_at, datetime):
        base = started_at
    else:
        text = clean_text(started_at)
        if not text:
            return 0.0
        base = None
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"]:
            try:
                base = datetime.strptime(text[:19], fmt)
                break
            except Exception:
                pass
        if base is None:
            return 0.0
    try:
        return max(0.0, (datetime.now() - base).total_seconds())
    except Exception:
        return 0.0


GENERATION_LOADING_MESSAGES = [
    '칭찬 쥐어짜는 중...',
    '학생의 작은 성취를 확대 해석하지 않게 조심하는 중...',
    '복붙 냄새를 환기시키는 중...',
    '생기부 문체로 점잖게 번역하는 중...',
    '과장 표현은 빼고 근거만 남기는 중...',
    '‘깊은 이해’를 몰래 검열하는 중...',
    '교사 양심과 업무 효율 사이에서 줄타기 중...',
    '비슷하지만 안 비슷한 문장 찾는 중...',
    '성취수준을 생활기록부 말투로 바꾸는 중...',
    '문장 끝을 ‘함’으로 착지시키는 중...',
    '학생의 장점을 현미경으로 찾는 중...',
    '루브릭을 문장으로 반죽하는 중...',
    '너무 멋있어 보이지 않게 톤을 낮추는 중...',
    '평가 문구를 세특 문장으로 숙성시키는 중...',
    '행정 문체에 감성을 한 방울 넣는 중...',
    '학생 이름이 들어가지 않았는지 몰래 확인하는 중...',
    '적당히 구체적이고 적당히 담백한 문장 찾는 중...',
    '똑같은 말인데 다르게 보이게 다듬는 중...',
    '평가 자료 밖으로 상상력이 탈출하지 못하게 막는 중...',
    '근거 없는 칭찬을 얌전히 돌려보내는 중...',
    '문장을 한 문단 안에 얌전히 앉히는 중...',
    '성취수준 A를 자랑스럽지만 과하지 않게 포장하는 중...',
    '성취수준 C도 품격 있게 표현하는 중...',
    '‘보완 필요’라는 말을 예쁘게 우회하는 중...',
    '생기부에 들어가도 안 민망한 표현 찾는 중...',
    '너무 AI 같지 않게 숨 고르는 중...',
    '문장 사이 중복을 살살 덜어내는 중...',
    '활동 내용과 평가 문구를 자연스럽게 이어 붙이는 중...',
    '칭찬은 하되 무리수는 두지 않는 중...',
    '관찰 기록을 세특 문장으로 승진시키는 중...',
    '‘우수함’을 세 번 말하지 않으려고 노력하는 중...',
    '학생별 차이를 티 안 나게 살리는 중...',
    '생기부 문장에 교사 말투를 입히는 중...',
    '과학 활동을 행정 문장으로 변환하는 중...',
    '평가 요소들을 한 문장 안에서 화해시키는 중...',
    '문장이 너무 신나지 않게 진정시키는 중...',
    '교사의 관찰처럼 보이도록 문장을 다듬는 중...',
    '문장 끝맺음을 공손하게 정렬하는 중...',
    '‘탐구함’과 ‘분석함’ 사이에서 고민하는 중...',
    '생활기록부에 어울리는 온도로 데우는 중...',
    '같은 활동, 다른 표현을 열심히 짜내는 중...',
    '학생의 수행 내용을 근거 중심으로 압축하는 중...',
    '표현은 다양하게, 의미는 안전하게 맞추는 중...',
    '너무 칭찬 같지도, 너무 평범하지도 않게 조절하는 중...',
    '마지막 문장까지 명사형 종결로 착지 준비 중...',
]

def generation_loading_ticker_html(loading_offset_seconds=0) -> str:
    """생성 대기 중에도 화면에서 7초마다 바뀌는 짧은 안내 문구를 만든다."""
    if not GENERATION_LOADING_MESSAGES:
        return ""

    display_messages = GENERATION_LOADING_MESSAGES + [GENERATION_LOADING_MESSAGES[0]]
    line_height = 30
    interval_seconds = 7
    total_seconds = len(GENERATION_LOADING_MESSAGES) * interval_seconds
    try:
        offset_seconds = float(loading_offset_seconds or 0) % total_seconds
    except Exception:
        offset_seconds = 0.0
    message_lines = "".join([
        f'<div class="generation-loading-line">{html.escape(message)}</div>'
        for message in display_messages
    ])
    return f"""
    <div class="generation-loading-box">
<div class="generation-loading-window" style="height:{line_height}px;">
            <div class="generation-loading-track" style="animation: generationLoadingTicker {total_seconds}s steps({len(GENERATION_LOADING_MESSAGES)}) infinite; animation-delay:-{offset_seconds:.2f}s;">
                {message_lines}
            </div>
        </div>
    </div>
    """

def show_generation_overlay(slot, title, detail, progress_ratio=None, step_lines=None, recent_items=None, loading_offset_seconds=0):
    """AI 생성 중 진행 상황을 화면 안쪽에 표시하고, 직전 생성 완료 문장 1개만 보여준다."""
    if slot is None:
        slot = st.empty()

    safe_title = html.escape(clean_text(title))
    safe_detail = html.escape(clean_text(detail))
    loading_html = generation_loading_ticker_html(loading_offset_seconds=loading_offset_seconds)

    recent_items = recent_items or []
    previous_html = ""
    previous_item = None
    for item in reversed(recent_items):
        if isinstance(item, dict) and clean_text(item.get("text", "")):
            previous_item = item
            break

    if previous_item:
        label = html.escape(clean_text(previous_item.get("label", "직전 생성 학생")))
        text = html.escape(clean_text(previous_item.get("text", "")))
        previous_html = f"""
        <div class="generation-inline-previous">
            <div class="generation-inline-previous-title">바로 이전 학생 생성 결과</div>
            <div class="generation-inline-previous-label">{label}</div>
            <div class="generation-inline-previous-text">{text}</div>
        </div>
        """
    else:
        previous_html = """
        <div class="generation-inline-empty">
            아직 바로 이전 생성 결과가 없습니다. 첫 번째 학생이 생성되면 다음 학생 생성 중 이곳에 표시됩니다.
        </div>
        """

    if progress_ratio is None:
        progress_html = ""
    else:
        try:
            progress_value = max(0, min(100, int(float(progress_ratio) * 100)))
        except Exception:
            progress_value = 0
        progress_html = f"""
        <div class="generation-inline-progress-wrap">
            <div class="generation-inline-progress-bar" style="width:{progress_value}%;"></div>
        </div>
        <div class="generation-inline-percent">{progress_value}%</div>
        """

    card_html = f"""
    <!doctype html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
    html, body {{ margin: 0; padding: 0; background: transparent; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; }}
    .generation-inline-card {{
        width: calc(100% - 2px);
        background: #FFFFFF;
        border: 1px solid #CBD5E1;
        border-left: 8px solid #D92D20;
        border-radius: 16px;
        box-shadow: 0 3px 10px rgba(15, 23, 42, 0.08);
        padding: 15px 17px 14px 17px;
        color: #111827;
        box-sizing: border-box;
        margin: 0;
    }}
    .generation-inline-title {{ font-size: 1.02rem; font-weight: 900; margin-bottom: 5px; }}
    .generation-inline-detail {{ font-size: 0.93rem; font-weight: 700; color: #334155; margin-bottom: 9px; }}
    .generation-inline-progress-wrap {{ width: 100%; height: 10px; border-radius: 999px; background: #E5E7EB; overflow: hidden; margin-top: 8px; }}
    .generation-inline-progress-bar {{ height: 10px; border-radius: 999px; background: linear-gradient(90deg, #EF4444 0%, #D92D20 100%); transition: width 0.25s ease; }}
    .generation-inline-percent {{ margin-top: 4px; text-align: right; color: #64748B; font-size: 0.78rem; font-weight: 800; }}
    .generation-loading-box {{ margin-top: 10px; background: #FFF7ED; border: 1px solid #FED7AA; border-radius: 12px; padding: 8px 11px; display: flex; align-items: center; overflow: hidden; }}
    .generation-loading-label {{ display: none; }}
    .generation-loading-window {{ flex: 1 1 auto; overflow: hidden; min-width: 0; }}
    .generation-loading-track {{ will-change: transform; }}
    .generation-loading-line {{ height: 30px; line-height: 30px; color: #7C2D12; font-size: 0.91rem; font-weight: 900; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    @keyframes generationLoadingTicker {{ from {{ transform: translateY(0); }} to {{ transform: translateY(-1350px); }} }}
    .generation-inline-previous {{ margin-top: 12px; background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 12px; padding: 11px 13px; }}
    .generation-inline-previous-title {{ font-weight: 900; color: #0F172A; font-size: 0.9rem; margin-bottom: 6px; }}
    .generation-inline-previous-label {{ font-weight: 900; color: #1E3A8A; font-size: 0.84rem; margin-bottom: 5px; }}
    .generation-inline-previous-text {{ color: #111827; font-size: 0.9rem; line-height: 1.6; white-space: pre-wrap; word-break: keep-all; overflow-wrap: anywhere; }}
    .generation-inline-empty {{ margin-top: 10px; color: #64748B; background: #F8FAFC; border: 1px dashed #CBD5E1; border-radius: 12px; padding: 10px 12px; font-size: 0.88rem; font-weight: 700; }}
    </style>
    </head>
    <body>
    <div class="generation-inline-card">
        <div class="generation-inline-title">{safe_title}</div>
        <div class="generation-inline-detail">{safe_detail}</div>
        {loading_html}
        {progress_html}
        {previous_html}
    </div>
    </body>
    </html>
    """
    slot.empty()
    with slot.container():
        components.html(card_html, height=360, scrolling=True)
    return slot

def generation_preview_items_from_log(log_entries):
    """전체 생성 중 바로 직전 생성 완료 문장 1개를 진행 상황 창에 표시하기 위한 자료로 변환한다."""
    preview_items = []
    for entry in (log_entries or [])[-1:]:
        if not isinstance(entry, dict):
            continue
        order_text = clean_text(entry.get("순서", ""))
        class_text = clean_text(entry.get("반", ""))
        number_text = clean_text(entry.get("번호", ""))
        name_text = clean_text(entry.get("성명", ""))
        label_parts = []
        if order_text:
            label_parts.append(f"{order_text}번 생성")
        student_label = f"{class_text}반 {number_text}번 {name_text}".strip()
        if student_label:
            label_parts.append(student_label)
        preview_items.append(
            {
                "label": " · ".join(label_parts) if label_parts else "이전 생성 문장",
                "text": clean_text(entry.get("생성 문구", "")),
            }
        )
    return preview_items



def looks_like_api_error_text(text: str) -> bool:
    """API 오류 문구가 결과표/엑셀에 저장되어 보이는 것을 막기 위한 감지 함수."""
    lower = clean_text(text).lower()
    if not lower:
        return False
    error_markers = [
        "api 생성 중 오류",
        "error code:",
        "insufficient_quota",
        "you exceeded your current quota",
        "please check your plan and billing details",
        "rate limit",
        "http 429",
        "429",
    ]
    return any(marker in lower for marker in error_markers)

def generation_preview_items_from_results(students_df, results, exclude_sid=""):
    """선택 생성 중 이미 생성되어 있는 최근 문장 1개를 표시한다."""
    if not isinstance(results, dict) or students_df is None or getattr(students_df, "empty", True):
        return []

    student_map = {clean_text(row.get("student_id", "")): row for _, row in students_df.iterrows()}
    entries = []
    for sid, result in results.items():
        sid = clean_text(sid)
        if not sid or sid == clean_text(exclude_sid) or not isinstance(result, dict):
            continue
        text = clean_text(result.get("edited", result.get("generated", "")))
        if looks_like_api_error_text(text):
            continue
        if not text:
            continue
        row = student_map.get(sid, {})
        label = f"{clean_text(row.get('반', ''))}반 {clean_text(row.get('번호', ''))}번 {clean_text(row.get('성명', ''))}".strip()
        entries.append(
            {
                "created_at": clean_text(result.get("created_at", "")),
                "label": label or "이전 생성 문장",
                "text": text,
            }
        )

    entries = sorted(entries, key=lambda x: x.get("created_at", ""))
    return [{"label": e["label"], "text": e["text"]} for e in entries[-1:]]

def to_int_or_big(value):
    text = re.sub(r"\D", "", clean_text(value))
    if text == "":
        return 999999
    return int(text)


def sort_students_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    result = df.copy()
    for col in ["student_id", "학년", "반", "번호", "성명"]:
        if col not in result.columns:
            result[col] = ""

    result["_학년정렬"] = result["학년"].map(to_int_or_big)
    result["_반정렬"] = result["반"].map(to_int_or_big)
    result["_번호정렬"] = result["번호"].map(to_int_or_big)

    result = result.sort_values(
        by=["_학년정렬", "_반정렬", "_번호정렬", "성명"],
        kind="stable",
    ).drop(columns=["_학년정렬", "_반정렬", "_번호정렬"])

    return result.reset_index(drop=True)



def needs_name_review(name) -> bool:
    """
    성명에 영어 알파벳이 포함된 경우에만 이름 확인 필요 학생으로 표시한다.
    블라인드 처리 후 생기는 * 문자는 확인 대상에 포함하지 않는다.
    """
    text = clean_text(name)
    if not text:
        return False

    return re.search(r"[A-Za-z]", text) is not None


def apply_name_review_edits(base_df: pd.DataFrame, review_df: pd.DataFrame) -> pd.DataFrame:
    """
    이름 확인 필요 학생만 따로 수정한 내용을 업로드 미리보기 명단에 반영한다.
    """
    if base_df.empty or review_df.empty or "student_id" not in base_df.columns or "student_id" not in review_df.columns:
        return base_df

    result = base_df.copy()
    review_map = review_df.set_index("student_id").to_dict(orient="index")

    for idx, row in result.iterrows():
        sid = row.get("student_id", "")
        if sid in review_map:
            for col in ["학년", "반", "번호", "성명"]:
                if col in review_map[sid]:
                    result.at[idx, col] = review_map[sid][col]

    return sort_students_df(result)



def mask_student_name(name, mask_char="*") -> str:
    """
    학생 이름을 개인정보 보호용으로 블라인드 처리한다.
    - 1글자: *
    - 2글자: 홍*
    - 3글자: 홍*동
    - 4글자 이상: 홍**동
    공백은 제거하지 않고 전체 문자열 기준으로 처리한다.
    """
    text = clean_text(name)
    if not text:
        return ""

    chars = list(text)
    length = len(chars)

    if length == 1:
        return mask_char
    if length == 2:
        return chars[0] + mask_char

    return chars[0] + (mask_char * (length - 2)) + chars[-1]


def mask_student_names_in_df(df: pd.DataFrame, mask_char="*") -> pd.DataFrame:
    result = df.copy()
    if "성명" not in result.columns:
        return result

    result["성명"] = result["성명"].apply(lambda x: mask_student_name(x, mask_char=mask_char))
    return result


def json_safe(obj):
    """
    프로젝트 JSON 저장 시 pandas/numpy 자료형 때문에 깨지지 않도록
    모든 값을 파이썬 기본 자료형 또는 문자열로 변환한다.
    """
    if isinstance(obj, dict):
        return {str(k): json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [json_safe(v) for v in obj]
    if isinstance(obj, tuple):
        return [json_safe(v) for v in obj]

    try:
        if pd.isna(obj):
            return ""
    except Exception:
        pass

    if hasattr(obj, "item"):
        try:
            return obj.item()
        except Exception:
            return str(obj)

    return obj


def init_state():
    if "settings" not in st.session_state:
        st.session_state.settings = {
            "school_year": "2026",
            "semester": "1학기",
            "school_level": "중학교",
            "grade": "1",
            "subject": "과학",
            "target_bytes_min": 700,
            "target_bytes_max": 800,
            "custom_rules": DEFAULT_RULES,
        }

    if "students" not in st.session_state:
        st.session_state.students = pd.DataFrame(
            columns=["student_id", "학년", "반", "번호", "성명"]
        )

    if "assessments" not in st.session_state:
        st.session_state.assessments = []

    # 주의: 세션 키 이름 items는 내부 메서드와 겹칠 수 있으므로 항상 대괄호 방식으로 접근한다.
    if "items" not in st.session_state:
        st.session_state["items"] = []

    if "records" not in st.session_state:
        st.session_state.records = {}

    if "results" not in st.session_state:
        st.session_state.results = {}

    if "generation_job" not in st.session_state:
        st.session_state.generation_job = {
            "active": False,
            "stop_requested": False,
            "student_ids": [],
            "index": 0,
            "log": [],
            "ai_provider": "ChatGPT",
            "api_key": "",
            "model": "",
            "started_at": "",
            "finished_at": "",
        }

    if "selected_generation_student_id" not in st.session_state:
        st.session_state.selected_generation_student_id = ""


def sanitize_state():
    """
    이전 버전에서 만들어진 불완전한 세션 데이터가 남아도 앱이 죽지 않도록 보정한다.
    """
    settings = st.session_state.get("settings", {})
    if "custom_rules" not in settings:
        # 예전 체크박스 규칙을 사용하던 프로젝트도 새 방식으로 변환
        settings["custom_rules"] = DEFAULT_RULES
    if "target_bytes_min" not in settings:
        settings["target_bytes_min"] = 700
    if "target_bytes_max" not in settings:
        settings["target_bytes_max"] = 800
    st.session_state.settings = settings

    if not isinstance(st.session_state.get("students"), pd.DataFrame):
        st.session_state.students = pd.DataFrame(st.session_state.get("students", []))

    for col in ["student_id", "학년", "반", "번호", "성명"]:
        if col not in st.session_state.students.columns:
            st.session_state.students[col] = ""

    # 학생 ID가 비어 있으면 보정
    if not st.session_state.students.empty:
        st.session_state.students["student_id"] = st.session_state.students["student_id"].apply(
            lambda x: clean_text(x) if clean_text(x) else make_id("stu")
        )
        st.session_state.students = sort_students_df(st.session_state.students)

    # 수행평가 보정
    clean_assessments = []
    for idx, assessment in enumerate(st.session_state.get("assessments", []), start=1):
        if not isinstance(assessment, dict):
            continue
        if not assessment.get("assessment_id"):
            assessment["assessment_id"] = make_id("assess")
        assessment["name"] = assessment.get("name") or "이름 없는 수행평가"
        assessment["area"] = assessment.get("area", "")
        assessment["description"] = assessment.get("description", "")
        assessment["order"] = int(assessment.get("order", idx) or idx)
        assessment["use"] = bool(assessment.get("use", True))
        clean_assessments.append(assessment)

    st.session_state.assessments = clean_assessments
    valid_assessment_ids = {a["assessment_id"] for a in clean_assessments}

    # 평가 요소 보정
    clean_items = []
    for idx, item in enumerate(st.session_state.get("items", []), start=1):
        if not isinstance(item, dict):
            continue
        if item.get("assessment_id") not in valid_assessment_ids:
            continue
        if not item.get("item_id"):
            item["item_id"] = make_id("item")
        item["name"] = item.get("name") or "이름 없는 평가 요소"
        if item.get("type") not in ["rubric", "comment", "rubric_plus"]:
            item["type"] = "rubric"
        if not isinstance(item.get("levels", []), list):
            item["levels"] = []
        if not isinstance(item.get("rubrics", {}), dict):
            item["rubrics"] = {}
        if item["type"] == "comment":
            item["levels"] = []
            item["rubrics"] = {}
        item["order"] = int(item.get("order", idx) or idx)
        clean_items.append(item)

    st.session_state["items"] = clean_items
    valid_item_ids = {item["item_id"] for item in clean_items}

    # records 보정
    clean_records = {}
    for key, value in st.session_state.get("records", {}).items():
        key = str(key)
        if "::" not in key:
            continue
        item_id = key.split("::")[-1]
        if item_id not in valid_item_ids:
            continue
        if not isinstance(value, dict):
            continue
        clean_records[key] = {
            "level": clean_text(value.get("level", "")),
            "comment": clean_text(value.get("comment", "")),
        }
    st.session_state.records = clean_records


init_state()
sanitize_state()


# =========================
# 나이스 엑셀 파싱
# =========================
def normalize_col_name(col) -> str:
    return clean_text(col).replace(" ", "").replace("\n", "")


def split_class_number(value):
    text = clean_text(value)
    nums = re.findall(r"\d+", text)

    if len(nums) >= 2:
        return str(int(nums[0])), str(int(nums[1]))
    if len(nums) == 1:
        return "", str(int(nums[0]))
    return "", ""


def find_header_row(raw_df: pd.DataFrame):
    for i in range(min(len(raw_df), 30)):
        row_values = [normalize_col_name(v) for v in raw_df.iloc[i].tolist()]

        has_name = any(v in ["성명", "이름", "학생명"] for v in row_values)
        has_grade = any(v in ["학년", "학년도"] for v in row_values)
        has_class_number = any(v in ["반/번호", "반번호", "반"] for v in row_values)

        if has_name and (has_grade or has_class_number):
            return i

    return None


def parse_neis_excel(uploaded_file):
    """
    나이스 세특 다운로드 파일에서 학생 명단을 추출한다.
    여러 시트에 명단이 있으면 모두 합친다.
    """
    file_bytes = uploaded_file.getvalue()
    xls = pd.ExcelFile(BytesIO(file_bytes))

    all_students = []
    found_sheets = []

    for sheet_name in xls.sheet_names:
        raw = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None)
        header_row = find_header_row(raw)

        if header_row is None:
            continue

        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=header_row)
        df.columns = [normalize_col_name(c) for c in df.columns]
        df = df.dropna(how="all").copy()

        grade_col = None
        name_col = None
        class_number_col = None
        class_col = None
        number_col = None
        school_year_col = None
        semester_col = None
        subject_col = None

        for col in df.columns:
            if col == "학년도":
                school_year_col = col
            elif col == "학기":
                semester_col = col
            elif col == "학년":
                grade_col = col
            elif col in ["성명", "이름", "학생명"]:
                name_col = col
            elif col in ["반/번호", "반번호"]:
                class_number_col = col
            elif col == "반":
                class_col = col
            elif col == "번호":
                number_col = col
            elif col == "과목명":
                subject_col = col

        if name_col is None:
            continue

        rows = []
        for _, row in df.iterrows():
            name = clean_text(row.get(name_col, ""))
            if not name or name in ["성명", "이름", "학생명"]:
                continue

            grade = clean_text(row.get(grade_col, "")) if grade_col else ""
            if not grade:
                grade = st.session_state.settings.get("grade", "")

            if class_number_col:
                class_no, number = split_class_number(row.get(class_number_col, ""))
            else:
                class_no = clean_text(row.get(class_col, "")) if class_col else ""
                number = clean_text(row.get(number_col, "")) if number_col else ""

            class_no = re.sub(r"\D", "", clean_text(class_no))
            number = re.sub(r"\D", "", clean_text(number))

            rows.append(
                {
                    "student_id": make_id("stu"),
                    "학년": grade,
                    "반": class_no,
                    "번호": number,
                    "성명": name,
                }
            )

        if rows:
            found_sheets.append(sheet_name)
            all_students.extend(rows)

            # 기본 설정 자동 반영은 최초 발견값만 사용
            if school_year_col and not df[school_year_col].dropna().empty:
                first_year = clean_text(df[school_year_col].dropna().iloc[0])
                if first_year:
                    st.session_state.settings["school_year"] = first_year
            if semester_col and not df[semester_col].dropna().empty:
                first_semester = clean_text(df[semester_col].dropna().iloc[0])
                if first_semester:
                    st.session_state.settings["semester"] = first_semester
            if subject_col and not df[subject_col].dropna().empty:
                first_subject = clean_text(df[subject_col].dropna().iloc[0])
                if first_subject:
                    st.session_state.settings["subject"] = first_subject

    if not all_students:
        return pd.DataFrame(columns=["student_id", "학년", "반", "번호", "성명"]), []

    result = pd.DataFrame(all_students)
    result = sort_students_df(result)
    return result, found_sheets


def combine_students(existing_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    existing_df = existing_df.copy()
    new_df = new_df.copy()

    for df in [existing_df, new_df]:
        for col in ["student_id", "학년", "반", "번호", "성명"]:
            if col not in df.columns:
                df[col] = ""

    combined = pd.concat([existing_df, new_df], ignore_index=True)

    # 학년/반/번호/성명이 같은 학생은 중복 제거
    combined = combined.drop_duplicates(
        subset=["학년", "반", "번호", "성명"],
        keep="first",
    )

    return sort_students_df(combined)


# =========================
# 프로젝트 저장/불러오기
# =========================
def project_to_json() -> str:
    data = {
        "settings": st.session_state.settings,
        "students": st.session_state.students.to_dict(orient="records"),
        "assessments": st.session_state.assessments,
        "items": st.session_state["items"],
        "records": st.session_state.records,
        "results": st.session_state.results,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "app": "개꿀 생기부",
        "version": "patched-20260624-v61",
    }
    return json.dumps(json_safe(data), ensure_ascii=False, indent=2, default=str)


def load_project_json(uploaded_file):
    data = json.load(uploaded_file)
    st.session_state.settings = data.get("settings", st.session_state.settings)
    st.session_state.students = pd.DataFrame(data.get("students", []))
    st.session_state.assessments = data.get("assessments", [])
    st.session_state["items"] = data.get("items", [])
    st.session_state.records = data.get("records", {})
    st.session_state.results = data.get("results", {})
    sanitize_state()


# =========================
# 수행평가/기록항목 관련 함수
# =========================
def get_items_for_assessment(assessment_id):
    return [
        item
        for item in st.session_state["items"]
        if isinstance(item, dict) and item.get("assessment_id", "") == assessment_id
    ]



def normalize_item_orders(assessment_id):
    """
    한 수행평가 안의 평가 요소 순서를 1, 2, 3...처럼 중복 없이 다시 정리한다.
    이전 버전에서 같은 순서가 저장되어 있어도 자동으로 보정된다.
    """
    items = get_items_for_assessment(assessment_id)
    items = sorted(
        items,
        key=lambda x: (
            int(x.get("order", 999) or 999),
            clean_text(x.get("name", "")),
            clean_text(x.get("item_id", "")),
        ),
    )

    for idx, item in enumerate(items, start=1):
        item["order"] = idx



def normalize_assessment_orders():
    """
    수행평가 순서를 1, 2, 3...처럼 중복 없이 다시 정리한다.
    """
    assessments = sorted(
        st.session_state.assessments,
        key=lambda x: (
            int(x.get("order", 999) or 999),
            clean_text(x.get("name", "")),
            clean_text(x.get("assessment_id", "")),
        ),
    )

    for idx, assessment in enumerate(assessments, start=1):
        assessment["order"] = idx


def apply_assessment_drag_order(sorted_labels, label_to_assessment_id):
    """
    드래그 앤 드롭 결과에 맞춰 수행평가 order 값을 재배치한다.
    """
    id_to_assessment = {
        assessment.get("assessment_id", ""): assessment
        for assessment in st.session_state.assessments
    }

    for idx, label in enumerate(sorted_labels, start=1):
        assessment_id = label_to_assessment_id.get(label)
        if assessment_id in id_to_assessment:
            id_to_assessment[assessment_id]["order"] = idx


def apply_item_drag_order(assessment_id, sorted_labels, label_to_item_id):
    """
    드래그 앤 드롭 결과에 맞춰 한 수행평가 안의 평가 요소 order 값을 재배치한다.
    """
    id_to_item = {
        item.get("item_id", ""): item
        for item in get_items_for_assessment(assessment_id)
    }

    for idx, label in enumerate(sorted_labels, start=1):
        item_id = label_to_item_id.get(label)
        if item_id in id_to_item:
            id_to_item[item_id]["order"] = idx


def sortable_style():
    """
    드래그 박스가 항목 수와 관계없이 세로 목록으로 안정적으로 보이도록 하는 최소 스타일.
    항목이 3개 이상이어도 한 줄로 압축되지 않고 아래로 이어지게 한다.
    """
    return """
    .sortable-component {
        width: 100% !important;
        border: 1px solid #D1D5DB !important;
        border-radius: 12px !important;
        padding: 10px !important;
        background-color: #F9FAFB !important;
        box-sizing: border-box !important;
        overflow: visible !important;
    }
    .sortable-container {
        width: 100% !important;
        background-color: #F9FAFB !important;
        border-radius: 10px !important;
        padding: 6px !important;
        box-sizing: border-box !important;
        overflow: visible !important;
    }
    .sortable-container-header {
        width: 100% !important;
        background-color: #F3F4F6 !important;
        color: #111827 !important;
        font-weight: 700 !important;
        padding: 8px 12px !important;
        border-radius: 8px !important;
        border: 1px solid #E5E7EB !important;
        box-sizing: border-box !important;
    }
    .sortable-container-body {
        width: 100% !important;
        display: flex !important;
        flex-direction: column !important;
        flex-wrap: nowrap !important;
        align-items: stretch !important;
        gap: 8px !important;
        background-color: #F9FAFB !important;
        padding-top: 8px !important;
        box-sizing: border-box !important;
        overflow: visible !important;
    }
    .sortable-item,
    .sortable-item:hover,
    .sortable-container-body > div,
    .sortable-container-body > div:hover {
        display: block !important;
        width: 100% !important;
        min-height: 44px !important;
        background-color: #FFFFFF !important;
        color: #111827 !important;
        font-weight: 700 !important;
        border: 1px solid #D1D5DB !important;
        border-radius: 10px !important;
        padding: 11px 14px !important;
        margin: 0 !important;
        box-shadow: 0 1px 2px rgba(17, 24, 39, 0.06) !important;
        box-sizing: border-box !important;
        white-space: normal !important;
        line-height: 1.35 !important;
        overflow-wrap: anywhere !important;
    }
    .sortable-item:hover,
    .sortable-container-body > div:hover {
        background-color: #F3F4F6 !important;
        border-color: #9CA3AF !important;
    }
    .sortable-item::before {
        content: "☰ " !important;
        color: #6B7280 !important;
        font-weight: 700 !important;
    }
    """


def sortable_key(base_key, labels):
    """
    streamlit-sortables 컴포넌트는 key가 그대로이면 이전 항목 개수를 기억하는 경우가 있다.
    그래서 항목이 추가/삭제/이름 변경될 때 key가 자동으로 달라지게 해 새 목록을 즉시 렌더링한다.
    """
    raw = "||".join([clean_text(label) for label in labels])
    digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:10]
    return f"{base_key}_{len(labels)}_{digest}"


def sort_labels_with_gray_box(labels, key, header="정렬"):
    """
    수행평가/평가 요소 공통 드래그 정렬 함수.
    항목 추가 직후에도 드래그 박스가 새 항목 수를 반영하도록 key를 동적으로 만든다.
    """
    if sort_items is None:
        st.warning("드래그 정렬 기능을 사용하려면 requirements.txt에 streamlit-sortables를 추가해야 합니다.")
        return labels

    if len(labels) < 2:
        return labels

    component_key = sortable_key(key, labels)

    try:
        return sort_items(
            labels,
            header=header,
            custom_style=sortable_style(),
            key=component_key,
        )
    except TypeError:
        st.warning("현재 설치된 streamlit-sortables가 회색 스타일을 지원하지 않아 기본 드래그 박스로 표시합니다.")
        try:
            return sort_items(labels, key=f"{component_key}_basic")
        except Exception as e:
            st.error(f"드래그 정렬 컴포넌트 오류: {e}")
            return labels
    except Exception as e:
        st.error(f"드래그 정렬 컴포넌트 오류: {e}")
        return labels

def shift_item_orders_for_insert(assessment_id, insert_order):
    """
    새 평가 요소를 특정 순서에 끼워 넣을 때,
    기존 평가 요소들의 순서를 뒤로 밀어 중복을 막는다.
    """
    normalize_item_orders(assessment_id)

    for item in get_items_for_assessment(assessment_id):
        current_order = int(item.get("order", 999) or 999)
        if current_order >= int(insert_order):
            item["order"] = current_order + 1


def move_item_to_order(assessment_id, item_id, new_order):
    """
    기존 평가 요소의 순서를 바꿀 때,
    같은 수행평가 안의 다른 평가 요소 순서도 함께 재배치한다.
    """
    items = sorted(
        get_items_for_assessment(assessment_id),
        key=lambda x: int(x.get("order", 999) or 999),
    )

    moving_item = None
    remaining_items = []

    for item in items:
        if item.get("item_id", "") == item_id:
            moving_item = item
        else:
            remaining_items.append(item)

    if moving_item is None:
        return

    new_order = max(1, min(int(new_order), len(items)))
    remaining_items.insert(new_order - 1, moving_item)

    for idx, item in enumerate(remaining_items, start=1):
        item["order"] = idx


def get_assessment_name(assessment_id):
    for assessment in st.session_state.assessments:
        if assessment.get("assessment_id", "") == assessment_id:
            return assessment.get("name", "")
    return ""


def record_key(student_id, item_id):
    return f"{student_id}::{item_id}"


def get_record(student_id, item_id):
    return st.session_state.records.get(
        record_key(student_id, item_id),
        {"level": "", "comment": ""},
    )


def set_record(student_id, item_id, level="", comment=""):
    st.session_state.records[record_key(student_id, item_id)] = {
        "level": clean_text(level),
        "comment": clean_text(comment),
    }


def parse_rubric_text(levels_text, rubrics_text):
    levels = [x.strip() for x in re.split(r"[,/|]", levels_text) if x.strip()]
    rubrics = {level: "" for level in levels}

    for line in rubrics_text.splitlines():
        line = line.strip()
        if not line:
            continue

        if "=" in line:
            key, value = line.split("=", 1)
        elif ":" in line:
            key, value = line.split(":", 1)
        else:
            continue

        key = key.strip()
        value = value.strip()

        if key:
            rubrics[key] = value
            if key not in levels:
                levels.append(key)

    return levels, rubrics


def item_type_to_kor(item_type):
    return {
        "rubric": "성취도 선택형",
        "comment": "개별 코멘트형",
        "rubric_plus": "성취도 + 추가 코멘트형",
    }.get(item_type, "성취도 선택형")


def item_type_from_kor(label):
    return {
        "성취도 선택형": "rubric",
        "개별 코멘트형": "comment",
        "성취도 + 추가 코멘트형": "rubric_plus",
    }.get(label, "rubric")



def default_level_code(index):
    default_codes = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    if 0 <= index < len(default_codes):
        return default_codes[index]
    return str(index + 1)


def render_rubric_input_block(prefix, current_levels=None, current_rubrics=None):
    """
    성취수준 개수를 1~10개 중 선택하고, 개수에 맞춰
    성취수준 코드와 교사의 평가 문구를 각각 입력받는다.
    """
    current_levels = current_levels if isinstance(current_levels, list) else []
    current_rubrics = current_rubrics if isinstance(current_rubrics, dict) else {}

    default_count = len(current_levels) if current_levels else 5
    default_count = max(1, min(10, int(default_count)))

    level_count = st.selectbox(
        "성취수준 개수",
        options=list(range(1, 11)),
        index=default_count - 1,
        key=f"{prefix}_level_count",
    )

    st.markdown("**성취수준 코드와 성취 수준별 교사의 평가 문구**")

    levels = []
    rubrics = {}

    for idx in range(level_count):
        existing_code = current_levels[idx] if idx < len(current_levels) else default_level_code(idx)
        existing_text = current_rubrics.get(existing_code, "")

        col_code, col_text = st.columns([1, 5])
        with col_code:
            code = st.text_input(
                f"{idx + 1}번 코드",
                value=existing_code,
                key=f"{prefix}_code_{idx}",
                label_visibility="collapsed",
                placeholder="A",
            )
        with col_text:
            comment = st.text_area(
                f"{idx + 1}번 평가 문구",
                value=existing_text,
                key=f"{prefix}_comment_{idx}",
                height=80,
                label_visibility="collapsed",
                placeholder="내용을 입력하세요. 교사가 평가한 내용을 입력해야 합니다.",
            )

        code = clean_text(code)
        if code:
            levels.append(code)
            rubrics[code] = clean_text(comment)

    return levels, rubrics


def render_add_item_expander(aid, item_count):
    """
    수행평가별 평가 요소 추가 영역.
    v20에서는 기존 평가 요소 목록을 모두 본 뒤 바로 아래에서 새 평가 요소를 추가할 수 있도록
    이 함수를 평가 요소 목록 하단에서 호출한다.
    """
    with st.expander("➕ 이 수행평가에 평가 요소 추가", expanded=(item_count == 0)):
        st.markdown(
            """
            수행평가 안에 존재하는 여러 평가 요소를 추가해주세요.  
            A, B, C와 같은 **성취도 선택형**으로 개별화시킬 수도 있고, 개인마다 다른 수행평가 코멘트를 적어주는 **개별 코멘트형**으로 더욱 구체적인 개별화가 가능합니다.  
            또한 이 두 가지를 융합한 **성취도 + 추가 코멘트형**도 가능합니다.
            """
        )

        item_name = st.text_input(
            "평가 요소명",
            placeholder="예: 생태지도 결과물 평가",
            key=f"new_item_name_{aid}",
        )

        item_type_label = st.selectbox(
            "기록 방식",
            ["성취도 선택형", "개별 코멘트형", "성취도 + 추가 코멘트형"],
            key=f"new_item_type_{aid}",
            help="개별 코멘트형을 선택하면 성취수준 코드와 루브릭 입력칸이 사라집니다.",
        )
        item_type = item_type_from_kor(item_type_label)

        levels = []
        rubrics = {}

        if item_type != "comment":
            levels, rubrics = render_rubric_input_block(
                prefix=f"new_item_rubric_{aid}",
                current_levels=["A", "B", "C", "D", "E"],
                current_rubrics={
                    "A": "우수한 수준으로 수행함",
                    "B": "대체로 적절하게 수행함",
                    "C": "일부 보완이 필요함",
                    "D": "기본적인 참여가 이루어짐",
                    "E": "지속적인 보완이 필요함",
                },
            )
        else:
            st.info("개별 코멘트형입니다. 성취수준 코드 없이 학생별 서술형 코멘트만 입력합니다.")

        item_order = len(get_items_for_assessment(aid)) + 1

        if st.button("이 수행평가에 평가 요소 추가", key=f"add_item_button_{aid}"):
            if not item_name.strip():
                st.warning("평가 요소명을 입력하세요.")
            else:
                if item_type == "comment":
                    levels, rubrics = [], {}

                st.session_state["items"].append(
                    {
                        "item_id": make_id("item"),
                        "assessment_id": aid,
                        "name": item_name.strip(),
                        "type": item_type,
                        "levels": levels,
                        "rubrics": rubrics,
                        "order": int(item_order),
                    }
                )
                normalize_item_orders(aid)
                sanitize_state()
                st.success("평가 요소를 추가했습니다.")
                st.rerun()


# =========================
# API 입력 자료 및 생성
# =========================
def format_selected_level_text(item, level):
    """
    API 입력 자료에서 선택 성취수준의 위치가 드러나도록 표기한다.
    예: 전체 성취수준 A, B, C, D, E 중 A
    """
    level = clean_text(level)
    levels = [clean_text(x) for x in item.get("levels", []) if clean_text(x)]

    if level and levels:
        return f"전체 성취수준 {', '.join(levels)} 중 {level}"
    if level:
        return f"성취수준 {level}"
    if levels:
        return f"전체 성취수준 {', '.join(levels)} 중 미입력"
    return "성취수준 미입력"


def build_student_material(student):
    student_id = student["student_id"]
    lines = []

    # 개인정보 보호를 위해 프롬프트 입력 자료에는 학년/반/번호/성명을 넣지 않는다.

    used_assessments = [a for a in st.session_state.assessments if a.get("use", True)]
    used_assessments = sorted(used_assessments, key=lambda x: int(x.get("order", 999) or 999))

    count = 1
    for assessment in used_assessments:
        assessment_id = assessment.get("assessment_id", "")
        chunks = []

        item_list = sorted(
            get_items_for_assessment(assessment_id),
            key=lambda x: int(x.get("order", 999) or 999),
        )

        for item in item_list:
            rec = get_record(student_id, item.get("item_id", ""))
            level = rec.get("level", "")
            comment = rec.get("comment", "")
            item_type = item.get("type", "rubric")
            teacher_comment = ""

            if item_type in ["rubric", "rubric_plus"]:
                teacher_comment = item.get("rubrics", {}).get(level, "")

            if item_type == "rubric":
                if level or teacher_comment:
                    level_text = format_selected_level_text(item, level)
                    chunks.append(
                        f"- {item.get('name', '')}: {level_text} / 교사의 평가: {teacher_comment}"
                    )

            elif item_type == "comment":
                if comment:
                    chunks.append(f"- {item.get('name', '')}: {comment}")

            elif item_type == "rubric_plus":
                if level or teacher_comment or comment:
                    level_text = format_selected_level_text(item, level)
                    text = f"- {item.get('name', '')}: {level_text} / 교사의 평가: {teacher_comment}"
                    if comment:
                        text += f" / 추가 코멘트: {comment}"
                    chunks.append(text)

        if chunks:
            lines.append(f"{count}. {assessment.get('name', '')}")
            if assessment.get("area"):
                lines.append(f"- 영역/단원: {assessment.get('area', '')}")
            if assessment.get("description"):
                lines.append(f"- 성취기준/수행평가 설명: {assessment.get('description', '')}")
            lines.extend(chunks)
            lines.append("")
            count += 1

    return "\n".join(lines).strip()


def build_prompt(material):
    settings = st.session_state.settings
    custom_rules = clean_text(settings.get("custom_rules", DEFAULT_RULES))

    prompt = f"""
{MASTER_PROMPT}

[선생님 추가 작성 규칙]
{custom_rules}

[작성 조건]
- 목표 분량: {settings.get('target_bytes_min', 700)}~{settings.get('target_bytes_max', 800)} byte
- 한 문단으로 작성한다.
- 문장 사이 연결을 자연스럽게 다듬는다.

[평가 자료]
{material}

[최종 출력]
세부능력 및 특기사항 문장만 출력하라.
""".strip()
    return prompt


def fallback_generate(material):
    sentences = []

    for line in material.splitlines():
        line = line.strip()
        if not line.startswith("- "):
            continue

        if "교사의 평가:" in line:
            part = line.split("교사의 평가:", 1)[1].strip()

            if " / 추가 코멘트:" in part:
                base, extra = part.split(" / 추가 코멘트:", 1)
                if base.strip():
                    sentences.append(base.strip().rstrip("."))
                if extra.strip():
                    sentences.append(extra.strip().rstrip("."))
            elif part:
                sentences.append(part.strip().rstrip("."))

        elif (
            ":" in line
            and "학년/반/번호" not in line
            and "영역/단원" not in line
            and "성취기준/수행평가 설명" not in line
        ):
            part = line.split(":", 1)[1].strip()
            if part:
                sentences.append(part.strip().rstrip("."))

    seen = set()
    unique = []
    for sentence in sentences:
        if sentence and sentence not in seen:
            unique.append(sentence)
            seen.add(sentence)

    if not unique:
        return "입력된 평가 자료가 부족하여 세부능력 및 특기사항 문장 생성이 어려움."

    text = ". ".join(unique).strip()
    if not text.endswith(("함", "임", "음", ".")):
        text += "함"
    return text



def format_api_error_notice(provider: str, error) -> str:
    """사용자에게 보여줄 API 오류 안내를 짧고 안전한 문구로 바꾼다."""
    raw = clean_text(error)
    lower = raw.lower()
    if "insufficient_quota" in lower or "exceeded your current quota" in lower or "billing" in lower:
        return f"{provider} API 할당량 또는 결제 한도 문제로 API 생성에 실패했습니다. 결과표에는 오류 문구를 넣지 않고 내부 조합 방식으로 대체했습니다. API 사용량/결제 상태를 확인하세요."
    if "429" in lower or "rate limit" in lower:
        return f"{provider} API 요청 한도에 걸려 API 생성에 실패했습니다. 결과표에는 오류 문구를 넣지 않고 내부 조합 방식으로 대체했습니다. 잠시 뒤 다시 시도하세요."
    if "invalid" in lower and "key" in lower:
        return f"{provider} API Key 문제로 API 생성에 실패했습니다. 결과표에는 오류 문구를 넣지 않고 내부 조합 방식으로 대체했습니다. API Key를 확인하세요."
    return f"{provider} API 호출에 실패했습니다. 결과표에는 오류 문구를 넣지 않고 내부 조합 방식으로 대체했습니다."


def remember_api_generation_error(provider: str, error) -> None:
    """API 오류를 화면에 원문 그대로 뿌리지 않고, 다음 렌더링 때 짧은 안내만 보이게 저장한다."""
    st.session_state["api_generation_notice"] = format_api_error_notice(provider, error)
    st.session_state["api_generation_error_provider"] = clean_text(provider)
    st.session_state["api_generation_error_raw"] = clean_text(error)[:500]


def safe_result_text(result: dict) -> str:
    """기존 세션/프로젝트에 API 오류 문구가 저장되어 있으면 표출하지 않고 가능한 경우 대체 문장을 만든다."""
    if not isinstance(result, dict):
        return ""
    text = clean_text(result.get("edited", result.get("generated", "")))
    if looks_like_api_error_text(text):
        material = clean_text(result.get("material", ""))
        return fallback_generate(material) if material else ""
    return text

def generate_with_openai(prompt, api_key, model):
    if not api_key:
        return None

    try:
        from openai import OpenAI

        client = OpenAI(api_key=api_key)
        response = client.responses.create(
            model=model,
            input=prompt,
        )
        return response.output_text.strip()

    except Exception as e:
        remember_api_generation_error("ChatGPT", e)
        return None


def _post_json(url, headers, payload, timeout=90):
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib_request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib_request.urlopen(req, timeout=timeout) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else {}
    except urllib_error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"HTTP {e.code}: {detail[:700]}") from e


def generate_with_gemini(prompt, api_key, model):
    if not api_key:
        return None

    try:
        model_path = urllib_parse.quote(clean_text(model), safe="")
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_path}:generateContent?key={urllib_parse.quote(api_key)}"
        payload = {
            "contents": [
                {
                    "role": "user",
                    "parts": [{"text": prompt}],
                }
            ],
            "generationConfig": {
                "temperature": 0.3,
            },
        }
        data = _post_json(
            url,
            headers={"Content-Type": "application/json"},
            payload=payload,
        )
        parts = data.get("candidates", [{}])[0].get("content", {}).get("parts", [])
        text = "".join([part.get("text", "") for part in parts if isinstance(part, dict)]).strip()
        return text or None
    except Exception as e:
        remember_api_generation_error("Gemini", e)
        return None


def generate_with_claude(prompt, api_key, model):
    if not api_key:
        return None

    try:
        payload = {
            "model": model,
            "max_tokens": 2000,
            "temperature": 0.3,
            "messages": [
                {
                    "role": "user",
                    "content": prompt,
                }
            ],
        }
        data = _post_json(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
            payload=payload,
        )
        blocks = data.get("content", [])
        text = "".join([block.get("text", "") for block in blocks if isinstance(block, dict) and block.get("type") == "text"]).strip()
        return text or None
    except Exception as e:
        remember_api_generation_error("Claude", e)
        return None


def generate_with_ai(prompt, ai_provider, api_key, model):
    provider = ai_provider if ai_provider in AI_PROVIDER_OPTIONS else "ChatGPT"
    model = clean_text(model) or get_default_ai_model(provider)

    if provider == "Gemini":
        return generate_with_gemini(prompt, api_key, model)
    if provider == "Claude":
        return generate_with_claude(prompt, api_key, model)
    return generate_with_openai(prompt, api_key, model)


def export_excel():
    """
    최종 다운로드 엑셀을 만든다.
    - 첫 번째 시트는 반드시 최종생기부로 배치한다.
    - 학교 업무용으로 바로 확인하기 쉽도록 헤더, 고정틀, 너비, 줄바꿈, 테두리, byte 조건부 서식을 적용한다.
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    students = st.session_state.students.copy()

    result_rows = []
    for _, student in students.iterrows():
        sid = student["student_id"]
        result = st.session_state.results.get(sid, {})
        final_text = safe_result_text(result)

        result_rows.append(
            {
                "학년": student.get("학년", ""),
                "반": student.get("반", ""),
                "번호": student.get("번호", ""),
                "성명": student.get("성명", ""),
                "최종 생기부": final_text,
                "byte": byte_count(final_text),
                "생성일시": result.get("created_at", ""),
                "생성 원문": result.get("generated", ""),
                "API 입력자료": result.get("material", ""),
            }
        )

    result_df = pd.DataFrame(result_rows)

    item_rows = []
    for item in st.session_state["items"]:
        item_rows.append(
            {
                "수행평가": get_assessment_name(item.get("assessment_id", "")),
                "기록항목": item.get("name", ""),
                "기록방식": item_type_to_kor(item.get("type", "")),
                "성취수준": ", ".join(item.get("levels", [])),
                "루브릭": "\n".join([f"{k}: {v}" for k, v in item.get("rubrics", {}).items()]),
                "순서": item.get("order", ""),
            }
        )

    item_df = pd.DataFrame(item_rows)

    record_rows = []
    for _, student in students.iterrows():
        for item in st.session_state["items"]:
            rec = get_record(student["student_id"], item.get("item_id", ""))

            record_rows.append(
                {
                    "학년": student.get("학년", ""),
                    "반": student.get("반", ""),
                    "번호": student.get("번호", ""),
                    "성명": student.get("성명", ""),
                    "수행평가": get_assessment_name(item.get("assessment_id", "")),
                    "기록항목": item.get("name", ""),
                    "기록방식": item_type_to_kor(item.get("type", "")),
                    "성취수준": rec.get("level", ""),
                    "개별코멘트": rec.get("comment", ""),
                }
            )

    record_df = pd.DataFrame(record_rows)

    def excel_rgb(value, alpha=False):
        """openpyxl 색상 오류를 막기 위해 #RRGGBB/RRGGBB 값을 정리한다.
        탭 색상은 ARGB 8자리, 셀 채우기는 RGB 6자리를 사용한다.
        """
        text = clean_text(value).replace("#", "").upper()
        if len(text) == 8:
            return text if alpha else text[-6:]
        if len(text) == 6:
            return f"FF{text}" if alpha else text
        return "FF64748B" if alpha else "64748B"

    def style_worksheet(ws, header_fill_color="#1F2937", tab_color="#64748B", freeze_cell="A2"):
        """엑셀 시트를 읽기 좋게 공통 서식화한다."""
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = freeze_cell
        ws.sheet_properties.tabColor = excel_rgb(tab_color, alpha=True)

        header_fill = PatternFill("solid", fgColor=excel_rgb(header_fill_color, alpha=False))
        header_font = Font(bold=True, color="FFFFFF", size=11)
        body_font = Font(size=10, color="111827")
        thin_side = Side(style="thin", color="D1D5DB")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        body_alignment = Alignment(vertical="top", wrap_text=True)

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            ws.row_dimensions[1].height = 28

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = border
            ws.row_dimensions[row[0].row].height = 42

        # 자동 필터는 표가 비어 있어도 헤더 확인에 도움이 된다.
        if ws.max_column >= 1 and ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions

        # 기본 폭 자동 산정. 긴 문장 열은 너무 넓어지지 않게 제한한다.
        text_heavy_headers = {"최종 생기부", "생성 원문", "API 입력자료", "루브릭", "개별코멘트", "성취기준/수행평가 설명", "description"}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_value = clean_text(ws.cell(row=1, column=col_idx).value)
            max_len = max(8, len(header_value) + 2)

            for row_idx in range(2, min(ws.max_row, 80) + 1):
                value = clean_text(ws.cell(row=row_idx, column=col_idx).value)
                if "\n" in value:
                    value = max(value.splitlines(), key=len, default=value)
                max_len = max(max_len, len(value) + 2)

            if header_value in text_heavy_headers:
                width = min(max(max_len, 32), 72)
            elif header_value in ["학년", "반", "번호", "byte", "순서"]:
                width = 9
            elif header_value in ["성명", "기록방식", "성취수준"]:
                width = min(max(max_len, 12), 18)
            else:
                width = min(max(max_len, 12), 34)

            ws.column_dimensions[col_letter].width = width

    def highlight_byte_column(ws):
        """최종생기부 시트의 byte 열을 목표 범위 기준으로 표시한다."""
        if ws.max_row < 2:
            return

        headers = [clean_text(cell.value) for cell in ws[1]]
        if "byte" not in headers:
            return

        byte_col_idx = headers.index("byte") + 1
        byte_col_letter = get_column_letter(byte_col_idx)
        byte_range = f"{byte_col_letter}2:{byte_col_letter}{ws.max_row}"
        target_min = int(st.session_state.settings.get("target_bytes_min", 700))
        target_max = int(st.session_state.settings.get("target_bytes_max", 800))

        low_fill = PatternFill("solid", fgColor="FEF3C7")
        high_fill = PatternFill("solid", fgColor="FEE2E2")
        ok_fill = PatternFill("solid", fgColor="DCFCE7")

        ws.conditional_formatting.add(
            byte_range,
            CellIsRule(operator="lessThan", formula=[str(target_min)], fill=low_fill),
        )
        ws.conditional_formatting.add(
            byte_range,
            CellIsRule(operator="greaterThan", formula=[str(target_max)], fill=high_fill),
        )
        ws.conditional_formatting.add(
            byte_range,
            CellIsRule(operator="between", formula=[str(target_min), str(target_max)], fill=ok_fill),
        )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 최종생기부를 첫 번째 시트로 먼저 작성한다.
        result_df.to_excel(writer, sheet_name="최종생기부", index=False)
        record_df.to_excel(writer, sheet_name="학생별기록", index=False)
        item_df.to_excel(writer, sheet_name="평가요소_루브릭", index=False)
        pd.DataFrame(st.session_state.assessments).to_excel(
            writer, sheet_name="수행평가", index=False
        )
        students.drop(columns=["student_id"], errors="ignore").to_excel(
            writer, sheet_name="학생명단", index=False
        )

        workbook = writer.book

        # 첫 시트는 특히 왼쪽 학생 정보가 고정되도록 E2에서 틀 고정한다.
        final_ws = workbook["최종생기부"]
        style_worksheet(final_ws, header_fill_color="#1E3A8A", tab_color="#2563EB", freeze_cell="E2")
        highlight_byte_column(final_ws)

        # 최종 생기부 열은 실제 검토가 쉽도록 넓게 잡는다.
        final_widths = {
            "A": 8,
            "B": 8,
            "C": 8,
            "D": 14,
            "E": 72,
            "F": 10,
            "G": 20,
            "H": 56,
            "I": 72,
        }
        for col_letter, width in final_widths.items():
            final_ws.column_dimensions[col_letter].width = width

        if "학생별기록" in workbook.sheetnames:
            style_worksheet(workbook["학생별기록"], header_fill_color="#0F766E", tab_color="#14B8A6", freeze_cell="E2")
        if "평가요소_루브릭" in workbook.sheetnames:
            style_worksheet(workbook["평가요소_루브릭"], header_fill_color="#92400E", tab_color="#F59E0B", freeze_cell="A2")
        if "수행평가" in workbook.sheetnames:
            style_worksheet(workbook["수행평가"], header_fill_color="#5B21B6", tab_color="#8B5CF6", freeze_cell="A2")
        if "학생명단" in workbook.sheetnames:
            style_worksheet(workbook["학생명단"], header_fill_color="#374151", tab_color="#9CA3AF", freeze_cell="A2")

        # 첫 번째 시트가 실제로 최종생기부가 되도록 명시적으로 활성화한다.
        workbook.active = workbook.sheetnames.index("최종생기부")

    output.seek(0)
    return output

# =========================
# 학생별 기록 입력 엑셀 양식
# =========================
def student_display_label(student) -> str:
    """학생을 화면/엑셀에서 표시할 때 쓰는 이름."""
    return f"{clean_text(student.get('반', ''))}반 {clean_text(student.get('번호', ''))}번 {clean_text(student.get('성명', ''))}"


def selected_record_rows(selected_items):
    """
    선택된 수행평가/평가요소를 학생별 기록 입력용 입력 항목 구조로 변환한다.
    - 성취도 선택형: 성취수준 1개 열
    - 개별 코멘트형: 개별코멘트 1개 열
    - 성취도+추가코멘트형: 성취수준 1개 열 + 추가코멘트 1개 열
    """
    rows = []
    for item in selected_items:
        item_type = item.get("type", "rubric")
        base = {
            "assessment_id": item.get("assessment_id", ""),
            "assessment_name": get_assessment_name(item.get("assessment_id", "")),
            "item_id": item.get("item_id", ""),
            "item_name": item.get("name", ""),
            "item_type": item_type,
            "levels": item.get("levels", []),
        }

        if item_type == "rubric":
            rows.append({**base, "field": "level", "field_label": "성취수준"})
        elif item_type == "comment":
            rows.append({**base, "field": "comment", "field_label": "개별코멘트"})
        elif item_type == "rubric_plus":
            rows.append({**base, "field": "level", "field_label": "성취수준"})
            rows.append({**base, "field": "comment", "field_label": "추가코멘트"})

    return rows


def record_column_label(record_row, used_labels=None):
    """
    웹 입력표는 엑셀처럼 다중 헤더를 직접 편집하기 어렵기 때문에
    수행평가 > 평가요소 > 입력구분을 한 열 제목에 압축해서 표시한다.
    """
    base = f"{record_row['assessment_name']} | {record_row['item_name']} | {record_row['field_label']}"
    if used_labels is None:
        return base

    label = base
    count = 2
    while label in used_labels:
        label = f"{base} ({count})"
        count += 1
    used_labels.add(label)
    return label


def build_record_matrix_df(students: pd.DataFrame, selected_items):
    """
    웹 화면에서 학생을 행으로, 평가요소 입력칸을 열로 배치한 전치형 입력표를 만든다.
    """
    record_rows = selected_record_rows(selected_items)
    used_labels = set()
    record_columns = []

    for record_row in record_rows:
        label = record_column_label(record_row, used_labels)
        record_columns.append({**record_row, "column_label": label})

    matrix_rows = []
    for _, student in students.iterrows():
        sid = student.get("student_id", "")
        row = {
            "_student_id": sid,
            "학년": clean_text(student.get("학년", "")),
            "반": clean_text(student.get("반", "")),
            "번호": clean_text(student.get("번호", "")),
            "성명": clean_text(student.get("성명", "")),
        }

        for record_col in record_columns:
            rec = get_record(sid, record_col["item_id"])
            if record_col["field"] == "level":
                row[record_col["column_label"]] = rec.get("level", "")
            else:
                row[record_col["column_label"]] = rec.get("comment", "")

        matrix_rows.append(row)

    return pd.DataFrame(matrix_rows), record_columns


def save_record_matrix_df(edited_df: pd.DataFrame, base_df: pd.DataFrame, record_columns):
    """전치형 웹 입력표에서 수정한 내용을 세션 기록에 저장한다."""
    if edited_df.empty or base_df.empty:
        return 0, []

    item_map = {
        item.get("item_id", ""): item
        for item in st.session_state["items"]
        if isinstance(item, dict)
    }
    warnings = []
    saved_count = 0

    for idx, row in edited_df.iterrows():
        if idx not in base_df.index:
            continue

        sid = clean_text(base_df.loc[idx, "_student_id"])
        student_label = student_display_label(base_df.loc[idx])
        if not sid:
            continue

        for record_col in record_columns:
            item_id = record_col["item_id"]
            field = record_col["field"]
            col_label = record_col["column_label"]
            item = item_map.get(item_id)
            if item is None or col_label not in edited_df.columns:
                continue

            value = clean_text(row.get(col_label, ""))
            rec = get_record(sid, item_id)

            if field == "level":
                allowed_levels = [clean_text(x) for x in item.get("levels", [])]
                if value and allowed_levels and value not in allowed_levels:
                    warnings.append(
                        f"{student_label} / {item.get('name', '')}: '{value}'는 등록된 성취수준 {allowed_levels}에 없습니다."
                    )
                set_record(sid, item_id, level=value, comment=rec.get("comment", ""))
                saved_count += 1
            elif field == "comment":
                set_record(sid, item_id, level=rec.get("level", ""), comment=value)
                saved_count += 1

    return saved_count, warnings


def merge_same_values(ws, row_idx, start_col, end_col):
    """같은 값이 이어지는 헤더 셀을 가로 병합한다."""
    if start_col > end_col:
        return

    group_start = start_col
    previous_value = clean_text(ws.cell(row_idx, start_col).value)

    for col_idx in range(start_col + 1, end_col + 2):
        current_value = clean_text(ws.cell(row_idx, col_idx).value) if col_idx <= end_col else "__END__"
        if current_value != previous_value:
            if previous_value and col_idx - group_start > 1:
                ws.merge_cells(start_row=row_idx, start_column=group_start, end_row=row_idx, end_column=col_idx - 1)
            group_start = col_idx
            previous_value = current_value


def make_student_record_excel(students: pd.DataFrame, selected_items, selected_assess_name: str = "전체"):
    """학생별 기록 입력용 전치형 엑셀 양식을 생성한다. 성취수준 입력칸에는 드롭다운을 넣는다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "학생별기록"
    list_ws = wb.create_sheet("선택목록")
    list_ws.sheet_state = "hidden"

    record_rows = selected_record_rows(selected_items)
    students = students.copy()

    title = f"BigHoneySangkibu 학생별 기록 입력 양식 - {selected_assess_name}"
    record_start_col = 6  # A 숨김 student_id, B 학년, C 반, D 번호, E 성명, F부터 입력 항목
    student_start_row = 8
    record_count = len(record_rows)
    record_end_col = record_start_col + record_count - 1
    last_visible_col = max(record_end_col, 5)

    # 제목/안내
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_visible_col)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=14, color="1F2937")
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_visible_col)
    ws.cell(2, 1).value = "학생은 행으로, 수행평가·평가요소는 열로 배치했습니다. 노란색 성취수준 칸은 드롭다운으로 선택하고, 코멘트 칸은 직접 입력한 뒤 이 파일을 다시 웹앱에 업로드하세요. 숨겨진 행/열은 삭제하지 마세요."
    ws.cell(2, 1).font = Font(size=10, color="6B7280")
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")

    # 숨김 메타 행: 각 입력 열의 item_id / field
    ws.row_dimensions[3].hidden = True
    ws.row_dimensions[4].hidden = True
    ws.cell(3, 1).value = "__item_id_row__"
    ws.cell(4, 1).value = "__field_row__"

    # 학생 기본 정보 헤더
    base_headers = {
        1: "__student_id__",
        2: "학년",
        3: "반",
        4: "번호",
        5: "성명",
    }
    for col_idx, header in base_headers.items():
        ws.cell(7, col_idx).value = header
        if col_idx >= 2:
            ws.merge_cells(start_row=5, start_column=col_idx, end_row=7, end_column=col_idx)
            ws.cell(5, col_idx).value = header

    ws.column_dimensions["A"].hidden = True

    # 선택목록 시트에 성취수준 목록 저장
    item_level_ranges = {}
    list_col = 1
    for item in selected_items:
        levels = [clean_text(x) for x in item.get("levels", []) if clean_text(x)]
        if not levels:
            continue

        item_id = item.get("item_id", "")
        list_ws.cell(1, list_col).value = item_id
        for row_offset, level in enumerate(levels, start=2):
            list_ws.cell(row_offset, list_col).value = level

        col_letter = get_column_letter(list_col)
        item_level_ranges[item_id] = f"'선택목록'!${col_letter}$2:${col_letter}${len(levels) + 1}"
        list_col += 1

    # 입력 항목 헤더 작성
    for offset, record_row in enumerate(record_rows):
        col_idx = record_start_col + offset
        item_id = record_row["item_id"]
        field = record_row["field"]

        ws.cell(3, col_idx).value = item_id
        ws.cell(4, col_idx).value = field
        ws.cell(5, col_idx).value = record_row["assessment_name"]
        ws.cell(6, col_idx).value = record_row["item_name"]
        ws.cell(7, col_idx).value = record_row["field_label"]

    if record_count > 0:
        merge_same_values(ws, 5, record_start_col, record_end_col)
        merge_same_values(ws, 6, record_start_col, record_end_col)

    # 스타일
    level_fill = PatternFill("solid", fgColor="FFF7ED")
    comment_fill = PatternFill("solid", fgColor="F9FAFB")
    assessment_fill = PatternFill("solid", fgColor="EFF6FF")
    item_fill = PatternFill("solid", fgColor="FFFBEB")
    field_fill = PatternFill("solid", fgColor="F3F4F6")
    student_fill = PatternFill("solid", fgColor="E5E7EB")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for row_idx in range(5, 8):
        for col_idx in range(1, last_visible_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color="111827")
            if col_idx < record_start_col:
                cell.fill = student_fill
            elif row_idx == 5:
                cell.fill = assessment_fill
                cell.font = Font(bold=True, color="1E3A8A")
            elif row_idx == 6:
                cell.fill = item_fill
                cell.font = Font(bold=True, color="92400E")
            else:
                cell.fill = field_fill

    item_map = {item.get("item_id", ""): item for item in selected_items}

    # 학생별 본문 작성
    for row_offset, (_, student) in enumerate(students.iterrows()):
        row_idx = student_start_row + row_offset
        sid = student.get("student_id", "")
        ws.cell(row_idx, 1).value = sid
        ws.cell(row_idx, 2).value = clean_text(student.get("학년", ""))
        ws.cell(row_idx, 3).value = clean_text(student.get("반", ""))
        ws.cell(row_idx, 4).value = clean_text(student.get("번호", ""))
        ws.cell(row_idx, 5).value = clean_text(student.get("성명", ""))

        for col_idx in range(1, 6):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="F9FAFB")

        for offset, record_row in enumerate(record_rows):
            col_idx = record_start_col + offset
            item_id = record_row["item_id"]
            field = record_row["field"]
            rec = get_record(sid, item_id)
            cell = ws.cell(row_idx, col_idx)
            cell.value = rec.get("level", "") if field == "level" else rec.get("comment", "")
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.fill = level_fill if field == "level" else comment_fill

    # 성취수준 드롭다운 적용
    student_end_row = student_start_row + len(students) - 1
    if len(students) > 0:
        for offset, record_row in enumerate(record_rows):
            item_id = record_row["item_id"]
            field = record_row["field"]
            col_idx = record_start_col + offset
            item = item_map.get(item_id, {})
            if field == "level" and item_id in item_level_ranges:
                dv = DataValidation(
                    type="list",
                    formula1=item_level_ranges[item_id],
                    allow_blank=True,
                    showDropDown=False,
                )
                dv.error = "등록된 성취수준 코드 중에서 선택하세요."
                dv.errorTitle = "성취수준 선택 오류"
                dv.prompt = "드롭다운에서 성취수준을 선택하세요."
                dv.promptTitle = "성취수준 선택"
                ws.add_data_validation(dv)
                col_letter = get_column_letter(col_idx)
                dv.add(f"{col_letter}{student_start_row}:{col_letter}{student_end_row}")

    # 열 너비와 고정
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    for col_idx in range(record_start_col, last_visible_col + 1):
        field = clean_text(ws.cell(4, col_idx).value)
        ws.column_dimensions[get_column_letter(col_idx)].width = 18 if field == "level" else 28

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 42
    ws.row_dimensions[7].height = 28
    for row_idx in range(student_start_row, student_start_row + len(students)):
        ws.row_dimensions[row_idx].height = 34

    ws.freeze_panes = "F8"
    ws.auto_filter.ref = f"B7:{get_column_letter(last_visible_col)}{max(student_start_row + len(students) - 1, 7)}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_student_record_excel(uploaded_file):
    """학생별 기록 입력 엑셀 파일을 읽어 현재 세션 기록에 반영한다."""
    wb = load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True)
    ws = wb["학생별기록"] if "학생별기록" in wb.sheetnames else wb.active

    # v30 전치형 양식: A열 hidden student_id, 3행 item_id, 4행 field, 8행부터 학생
    if clean_text(ws.cell(3, 1).value) == "__item_id_row__" and clean_text(ws.cell(4, 1).value) == "__field_row__":
        student_id_col = 1
        record_start_col = 6
        student_start_row = 8
        current_student_ids = set(st.session_state.students["student_id"].astype(str).tolist())

        record_cols = []
        for col_idx in range(record_start_col, ws.max_column + 1):
            item_id = clean_text(ws.cell(3, col_idx).value)
            field = clean_text(ws.cell(4, col_idx).value)
            if item_id and field:
                record_cols.append((col_idx, item_id, field))

        if not record_cols:
            return 0, ["엑셀 양식에서 평가요소 입력 열을 찾지 못했습니다. 웹앱에서 새 양식을 다운로드한 뒤 다시 입력해 주세요."]

        item_map = {
            item.get("item_id", ""): item
            for item in st.session_state["items"]
            if isinstance(item, dict)
        }

        saved_count = 0
        warnings = []
        matched_students = 0

        for row_idx in range(student_start_row, ws.max_row + 1):
            sid = clean_text(ws.cell(row_idx, student_id_col).value)
            if not sid or sid not in current_student_ids:
                continue
            matched_students += 1
            student_label = f"{clean_text(ws.cell(row_idx, 3).value)}반 {clean_text(ws.cell(row_idx, 4).value)}번 {clean_text(ws.cell(row_idx, 5).value)}"

            for col_idx, item_id, field in record_cols:
                if item_id not in item_map:
                    continue

                item = item_map[item_id]
                value = clean_text(ws.cell(row_idx, col_idx).value)
                rec = get_record(sid, item_id)

                if field == "level":
                    allowed_levels = [clean_text(x) for x in item.get("levels", [])]
                    if value and allowed_levels and value not in allowed_levels:
                        warnings.append(
                            f"{student_label} / {item.get('name', '')}: '{value}'는 등록된 성취수준 {allowed_levels}에 없습니다."
                        )
                    set_record(sid, item_id, level=value, comment=rec.get("comment", ""))
                    saved_count += 1
                elif field == "comment":
                    set_record(sid, item_id, level=rec.get("level", ""), comment=value)
                    saved_count += 1

        if matched_students == 0:
            return 0, ["현재 웹앱 학생 명단과 엑셀 양식의 학생 정보가 맞지 않습니다. 현재 학생 명단 기준으로 양식을 다시 다운로드해 주세요."]

        return saved_count, warnings

    # v29 이전 행 중심 양식도 업로드할 수 있도록 기존 방식 유지
    header_row = None
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [clean_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, min(ws.max_column, 30) + 1)]
        if "수행평가" in values and "평가요소" in values and "입력구분" in values:
            header_row = row_idx
            break

    if header_row is None:
        return 0, ["학생별 기록 입력 양식의 헤더를 찾지 못했습니다. 웹앱에서 새 양식을 다운로드한 뒤 다시 입력해 주세요."]

    headers = {
        clean_text(ws.cell(header_row, col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
        if clean_text(ws.cell(header_row, col_idx).value)
    }

    item_id_col = headers.get("__item_id__")
    field_col = headers.get("__field__")
    if not item_id_col or not field_col:
        return 0, ["숨겨진 메타데이터 열을 찾지 못했습니다. 엑셀 양식의 숨겨진 열을 삭제하지 않았는지 확인해 주세요."]

    student_start_col = 4
    student_end_col = item_id_col - 2
    student_id_row = header_row - 1

    current_student_ids = set(st.session_state.students["student_id"].astype(str).tolist())
    student_cols = []
    for col_idx in range(student_start_col, student_end_col + 1):
        sid = clean_text(ws.cell(student_id_row, col_idx).value)
        if sid and sid in current_student_ids:
            student_cols.append((col_idx, sid, clean_text(ws.cell(header_row, col_idx).value)))

    if not student_cols:
        return 0, ["현재 웹앱 학생 명단과 엑셀 양식의 학생 정보가 맞지 않습니다. 현재 학생 명단 기준으로 양식을 다시 다운로드해 주세요."]

    item_map = {
        item.get("item_id", ""): item
        for item in st.session_state["items"]
        if isinstance(item, dict)
    }

    saved_count = 0
    warnings = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        item_id = clean_text(ws.cell(row_idx, item_id_col).value)
        field = clean_text(ws.cell(row_idx, field_col).value)
        if not item_id or item_id not in item_map:
            continue

        item = item_map[item_id]
        allowed_levels = [clean_text(x) for x in item.get("levels", [])]

        for col_idx, sid, student_label in student_cols:
            value = clean_text(ws.cell(row_idx, col_idx).value)
            rec = get_record(sid, item_id)

            if field == "level":
                if value and allowed_levels and value not in allowed_levels:
                    warnings.append(
                        f"{student_label} / {item.get('name', '')}: '{value}'는 등록된 성취수준 {allowed_levels}에 없습니다."
                    )
                set_record(sid, item_id, level=value, comment=rec.get("comment", ""))
                saved_count += 1
            elif field == "comment":
                set_record(sid, item_id, level=rec.get("level", ""), comment=value)
                saved_count += 1

    return saved_count, warnings

def build_sample_project_data():
    """
    사이드바의 샘플 데이터 불러오기 버튼에서 사용하는 내장 샘플 프로젝트.
    실제 학생 정보가 아닌 가상 학생 3개 반, 반별 10명과 임의 학생별 기록을 포함한다.
    """
    data = json.loads(r'''{
  "settings": {
    "school_year": "2026",
    "semester": "1학기",
    "school_level": "고등학교",
    "grade": "2",
    "subject": "통합과학",
    "target_bytes_min": 700,
    "target_bytes_max": 800,
    "custom_rules": "- 꾸며내지 말 것\n- 교사의 관찰이 드러나도록 할 것"
  },
  "students": [
    {
      "student_id": "sample_stu_1_01",
      "학년": "2",
      "반": "1",
      "번호": "1",
      "성명": "김민준"
    },
    {
      "student_id": "sample_stu_1_02",
      "학년": "2",
      "반": "1",
      "번호": "2",
      "성명": "박서연"
    },
    {
      "student_id": "sample_stu_1_03",
      "학년": "2",
      "반": "1",
      "번호": "3",
      "성명": "최도윤"
    },
    {
      "student_id": "sample_stu_1_04",
      "학년": "2",
      "반": "1",
      "번호": "4",
      "성명": "이하은"
    },
    {
      "student_id": "sample_stu_1_05",
      "학년": "2",
      "반": "1",
      "번호": "5",
      "성명": "정우진"
    },
    {
      "student_id": "sample_stu_1_06",
      "학년": "2",
      "반": "1",
      "번호": "6",
      "성명": "한지우"
    },
    {
      "student_id": "sample_stu_1_07",
      "학년": "2",
      "반": "1",
      "번호": "7",
      "성명": "오서준"
    },
    {
      "student_id": "sample_stu_1_08",
      "학년": "2",
      "반": "1",
      "번호": "8",
      "성명": "윤채아"
    },
    {
      "student_id": "sample_stu_1_09",
      "학년": "2",
      "반": "1",
      "번호": "9",
      "성명": "Kim Alex"
    },
    {
      "student_id": "sample_stu_1_10",
      "학년": "2",
      "반": "1",
      "번호": "10",
      "성명": "임하준"
    },
    {
      "student_id": "sample_stu_2_01",
      "학년": "2",
      "반": "2",
      "번호": "1",
      "성명": "강민서"
    },
    {
      "student_id": "sample_stu_2_02",
      "학년": "2",
      "반": "2",
      "번호": "2",
      "성명": "송예린"
    },
    {
      "student_id": "sample_stu_2_03",
      "학년": "2",
      "반": "2",
      "번호": "3",
      "성명": "장도현"
    },
    {
      "student_id": "sample_stu_2_04",
      "학년": "2",
      "반": "2",
      "번호": "4",
      "성명": "유나경"
    },
    {
      "student_id": "sample_stu_2_05",
      "학년": "2",
      "반": "2",
      "번호": "5",
      "성명": "조현우"
    },
    {
      "student_id": "sample_stu_2_06",
      "학년": "2",
      "반": "2",
      "번호": "6",
      "성명": "신아린"
    },
    {
      "student_id": "sample_stu_2_07",
      "학년": "2",
      "반": "2",
      "번호": "7",
      "성명": "배준서"
    },
    {
      "student_id": "sample_stu_2_08",
      "학년": "2",
      "반": "2",
      "번호": "8",
      "성명": "Park Lina"
    },
    {
      "student_id": "sample_stu_2_09",
      "학년": "2",
      "반": "2",
      "번호": "9",
      "성명": "문태오"
    },
    {
      "student_id": "sample_stu_2_10",
      "학년": "2",
      "반": "2",
      "번호": "10",
      "성명": "서지민"
    },
    {
      "student_id": "sample_stu_3_01",
      "학년": "2",
      "반": "3",
      "번호": "1",
      "성명": "홍시우"
    },
    {
      "student_id": "sample_stu_3_02",
      "학년": "2",
      "반": "3",
      "번호": "2",
      "성명": "권예준"
    },
    {
      "student_id": "sample_stu_3_03",
      "학년": "2",
      "반": "3",
      "번호": "3",
      "성명": "남서아"
    },
    {
      "student_id": "sample_stu_3_04",
      "학년": "2",
      "반": "3",
      "번호": "4",
      "성명": "구민재"
    },
    {
      "student_id": "sample_stu_3_05",
      "학년": "2",
      "반": "3",
      "번호": "5",
      "성명": "하은별"
    },
    {
      "student_id": "sample_stu_3_06",
      "학년": "2",
      "반": "3",
      "번호": "6",
      "성명": "Choi Ryan"
    },
    {
      "student_id": "sample_stu_3_07",
      "학년": "2",
      "반": "3",
      "번호": "7",
      "성명": "백다온"
    },
    {
      "student_id": "sample_stu_3_08",
      "학년": "2",
      "반": "3",
      "번호": "8",
      "성명": "민서율"
    },
    {
      "student_id": "sample_stu_3_09",
      "학년": "2",
      "반": "3",
      "번호": "9",
      "성명": "노하윤"
    },
    {
      "student_id": "sample_stu_3_10",
      "학년": "2",
      "반": "3",
      "번호": "10",
      "성명": "차지호"
    }
  ],
  "assessments": [
    {
      "assessment_id": "assess_feaff4e0",
      "name": "우리 주변의 다양한 물질 탐구하기",
      "area": "Ⅰ. 물질과 규칙성",
      "description": "[10통과01-03]세상을 이루는 물질은 원소들로 이루어져 있으며, 원소들의 성질이 주기성을 나타내는 현상을 통해 자연의 규칙성을찾아낼 수 있다.\n[10통과01-04]지구와 생명체를 구성하는 주요 원소들이 결합을 형성하는 이유와 원소들의 성질에 따라 형성되는 결합의 종류를 추론할 수 있다.\n\n수행 평가 요약\n우리 주변의 흥미로운 물질 3가지를 조사하여 특징 및 구성 원소의 배열을 이해하고 주어진 조건에 맞게 원자 결합을 비유적으로 표현하기",
      "order": 1,
      "use": true
    },
    {
      "assessment_id": "assess_7289a43c",
      "name": "중력을 받는 물체의 운동 탐구",
      "area": "Ⅱ. 시스템과 상호 작용",
      "description": "[10통과03-01]자유 낙하와 수평으로 던진 물체의 운동을 이용하여 중력의 작용에 의한 역학적 시스템을 설명할 수 있다.\n\n수행 평가 요약\n중력을 받는 물체가 자유 낙하할 때와 수평 방향으로 던졌을 때 운동을 분석하고 중력가속도를 측정하기",
      "order": 2,
      "use": true
    }
  ],
  "items": [
    {
      "item_id": "item_03972701",
      "assessment_id": "assess_feaff4e0",
      "name": "물질 선정과 특징",
      "type": "rubric",
      "levels": [
        "A",
        "B",
        "C",
        "D"
      ],
      "rubrics": {
        "A": "화합물의 종류와 선정 이유를 분명하게 밝히고 그 특징에 대해 명확하게 서술함",
        "B": "화합물의 종류와 선정 이유 및 특징을 대부분 명확하게 서술함",
        "C": "화합물의 종류와 선정 이유 및 특징을 일부만 명확하게 서술함",
        "D": "화합물의 선정 이유와 특징 모두 미흡하게 설명함"
      },
      "order": 1
    },
    {
      "item_id": "item_95d74cad",
      "assessment_id": "assess_feaff4e0",
      "name": "결합 종류와 원자 배치 모형",
      "type": "rubric",
      "levels": [
        "A",
        "B",
        "C",
        "D",
        "E"
      ],
      "rubrics": {
        "A": "결합 종류를 분명하게 밝히고 원자 배치 모형 명확하게 제시함",
        "B": "결합 종류와 원자 배치 모형을 대부분 명확하게 제시함",
        "C": "결합 종류와 원자 배치 모형을 일부만 명확하게 제시함",
        "D": "결합 종류와 원자 배치 모형 모두 미흡하게 제시함",
        "E": "제출하지 않음. (이 영역 작성 X)"
      },
      "order": 2
    },
    {
      "item_id": "item_f41fa21a",
      "assessment_id": "assess_feaff4e0",
      "name": "화학 결합 비유적 표현",
      "type": "rubric_plus",
      "levels": [
        "A",
        "B",
        "C",
        "D"
      ],
      "rubrics": {
        "A": "화학 결합을 과학적 원리가 드러나도록 그림으로 표현하고 담긴 의도를 적절하게 설명함",
        "B": "화학 결합을 그림으로 표현하고 그림에 담긴 의도를 적절하게 설명하지 못했으나 과학적인 원리가 잘 표현됨",
        "C": "화학 결합을 그림으로 표현하고 그림에 담긴 의도를 적절하게 설명하였으나 과학적인 원리가 잘 표현하지 못함",
        "D": "화학 결합을 그린 그림 및 그림에 담긴 의도 모두 미흡하게 표현함"
      },
      "order": 3
    },
    {
      "item_id": "item_299588fe",
      "assessment_id": "assess_feaff4e0",
      "name": "전체 총평",
      "type": "comment",
      "levels": [],
      "rubrics": {},
      "order": 4
    },
    {
      "item_id": "item_e058012d",
      "assessment_id": "assess_7289a43c",
      "name": "실험 결과를 정리해 보고서 상의 표에 정리하기",
      "type": "rubric",
      "levels": [
        "A",
        "B1",
        "B2",
        "B3",
        "C"
      ],
      "rubrics": {
        "A": "실험 보고서에 빈 칸 없이 정확하게 기록함",
        "B1": "실험 보고서에 대체로 정확하게 기록하였으나, 단위가 잘못됨",
        "B2": "실험 보고서에 대체로 정확하게 기록하였으나, 계산이 잘못됨",
        "B3": "실험 보고서에 대체로 정확하게 기록하였으나, 사소한 실수가 있음",
        "C": "실험 보고서 작성이 미흡함"
      },
      "order": 1
    },
    {
      "item_id": "item_aa648d0f",
      "assessment_id": "assess_7289a43c",
      "name": "실험 결과를 토대로 연직 아래 방향으로 던진 물체의 운동을 분석하기",
      "type": "rubric",
      "levels": [
        "A",
        "B",
        "C",
        "D"
      ],
      "rubrics": {
        "A": "자유 낙하하는 물체의 속도 변화를 물체에 작용하는 힘과 관련지어 구체적으로 분석함.",
        "B": "자유 낙하하는 물체의 속도 변화를 물체에 작용하는 힘과 관련지어 대체로 잘 분석함.",
        "C": "자유 낙하하는 물체의 속도 변화를 물체에 작용하는 힘과 관련짓지는 못했지만, 속도 변화 또는 물체에 작용하는 힘을 분석함.",
        "D": "자유 낙하하는 물체의 속도 변화와 물체에 작용하는 힘을 분석하지 못함."
      },
      "order": 2
    },
    {
      "item_id": "item_2494ee70",
      "assessment_id": "assess_7289a43c",
      "name": "총평",
      "type": "comment",
      "levels": [],
      "rubrics": {},
      "order": 3
    }
  ],
  "records": {
    "sample_stu_1_01::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_01::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_01::item_f41fa21a": {
      "level": "A",
      "comment": "카페인과 향료 성분을 예로 들며 같은 물질도 쓰임에 따라 관심 지점이 달라질 수 있음을 질문함"
    },
    "sample_stu_1_01::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_01::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_01::item_299588fe": {
      "level": "",
      "comment": "조사한 물질을 단순 나열하지 않고 구성 원소와 쓰임을 함께 비교하여 물질 선택의 이유를 분명히 하려는 모습이 드러남"
    },
    "sample_stu_1_01::item_2494ee70": {
      "level": "",
      "comment": "낙하 시간과 이동 거리의 관계를 표로 정리한 뒤 값의 변화가 일정하지 않다는 점을 스스로 확인함"
    },
    "sample_stu_1_02::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_02::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_02::item_f41fa21a": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_02::item_e058012d": {
      "level": "B1",
      "comment": ""
    },
    "sample_stu_1_02::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_02::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_02::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_03::item_03972701": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_03::item_95d74cad": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_03::item_f41fa21a": {
      "level": "C",
      "comment": "샴푸 성분표에서 계면활성제를 찾아보고 분자 구조와 생활 속 기능을 연결하려는 시도가 드러남"
    },
    "sample_stu_1_03::item_e058012d": {
      "level": "B2",
      "comment": ""
    },
    "sample_stu_1_03::item_aa648d0f": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_03::item_299588fe": {
      "level": "",
      "comment": "생활 속 성분표와 교과 개념을 연결해 물질의 특징을 설명하려고 하였고, 낯선 화합물 이름을 스스로 찾아 정리함"
    },
    "sample_stu_1_03::item_2494ee70": {
      "level": "",
      "comment": "동영상 분석에서 기준점을 잡는 위치가 결과에 영향을 줄 수 있음을 언급하며 측정 오차를 고려함"
    },
    "sample_stu_1_04::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_04::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_04::item_f41fa21a": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_04::item_e058012d": {
      "level": "B3",
      "comment": ""
    },
    "sample_stu_1_04::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_04::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_04::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_05::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_05::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_05::item_f41fa21a": {
      "level": "A",
      "comment": "소금과 설탕을 비교하며 겉보기 성질만으로 물질의 결합 방식을 판단하기 어렵다는 점을 언급함"
    },
    "sample_stu_1_05::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_05::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_05::item_299588fe": {
      "level": "",
      "comment": "원소의 배열과 결합 종류를 모형으로 표현하는 과정에서 실제 물질의 성질을 근거로 삼으려는 시도가 나타남"
    },
    "sample_stu_1_05::item_2494ee70": {
      "level": "",
      "comment": "수평으로 던진 물체의 운동을 가로 방향과 세로 방향으로 나누어 보려는 관점이 나타남"
    },
    "sample_stu_1_06::item_03972701": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_1_06::item_95d74cad": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_1_06::item_f41fa21a": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_1_06::item_e058012d": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_06::item_aa648d0f": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_1_06::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_06::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_07::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_07::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_07::item_f41fa21a": {
      "level": "B",
      "comment": "탄산음료 속 이산화 탄소를 사례로 들어 분자 모형이 실제 물질의 성질 설명에 어떻게 쓰이는지 궁금해함"
    },
    "sample_stu_1_07::item_e058012d": {
      "level": "B1",
      "comment": ""
    },
    "sample_stu_1_07::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_07::item_299588fe": {
      "level": "",
      "comment": "선정한 물질의 사용 장면을 먼저 떠올린 뒤 그 성질이 어떤 구성 원소와 관련되는지 탐색함"
    },
    "sample_stu_1_07::item_2494ee70": {
      "level": "",
      "comment": "중력가속도 계산 과정에서 단위 변환을 다시 확인하며 실험값과 이론값의 차이를 비교함"
    },
    "sample_stu_1_08::item_03972701": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_08::item_95d74cad": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_08::item_f41fa21a": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_08::item_e058012d": {
      "level": "B2",
      "comment": ""
    },
    "sample_stu_1_08::item_aa648d0f": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_1_08::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_08::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_09::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_09::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_09::item_f41fa21a": {
      "level": "A",
      "comment": "유리와 플라스틱의 구성 원소를 비교하며 같은 생활용품도 원소 조합에 따라 성질이 달라진다는 점을 정리함"
    },
    "sample_stu_1_09::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_09::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_09::item_299588fe": {
      "level": "",
      "comment": "물질의 겉보기 특징과 내부 구조를 구분하려고 하였으며, 조사 자료를 자신의 표현으로 다시 정리함"
    },
    "sample_stu_1_09::item_2494ee70": {
      "level": "",
      "comment": "자유 낙하 운동을 속도 변화와 힘의 방향으로 연결하여 설명하려고 하였고, 그래프의 기울기 의미를 질문함"
    },
    "sample_stu_1_10::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_10::item_95d74cad": {
      "level": "E",
      "comment": ""
    },
    "sample_stu_1_10::item_f41fa21a": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_1_10::item_e058012d": {
      "level": "B3",
      "comment": ""
    },
    "sample_stu_1_10::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_1_10::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_1_10::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_01::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_01::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_01::item_f41fa21a": {
      "level": "A",
      "comment": "비누가 기름때를 제거하는 과정을 조사하며 물질의 구조와 기능을 연결하려는 관점이 나타남"
    },
    "sample_stu_2_01::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_01::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_01::item_299588fe": {
      "level": "",
      "comment": "생활 속 사례를 과학 개념으로 설명하려는 태도가 나타났고, 화학 결합을 비유로 표현하는 데 자신만의 관점을 반영함"
    },
    "sample_stu_2_01::item_2494ee70": {
      "level": "",
      "comment": "실험 장치의 높이와 출발 시점 설정이 결과에 영향을 줄 수 있음을 찾아 보고서에 반영함"
    },
    "sample_stu_2_02::item_03972701": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_02::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_02::item_f41fa21a": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_02::item_e058012d": {
      "level": "B1",
      "comment": ""
    },
    "sample_stu_2_02::item_aa648d0f": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_02::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_02::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_03::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_03::item_95d74cad": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_03::item_f41fa21a": {
      "level": "B",
      "comment": "배터리 소재에 쓰이는 리튬을 찾아보고 원소의 성질이 기술 활용과 연결될 수 있음을 제시함"
    },
    "sample_stu_2_03::item_e058012d": {
      "level": "B2",
      "comment": ""
    },
    "sample_stu_2_03::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_03::item_299588fe": {
      "level": "",
      "comment": "주기율표와 조사 자료를 함께 확인하며 원소의 성질이 물질의 특징과 연결되는 과정을 정리함"
    },
    "sample_stu_2_03::item_2494ee70": {
      "level": "",
      "comment": "연속 사진 자료에서 물체 사이 간격이 점점 커지는 모습을 근거로 속도 변화를 설명함"
    },
    "sample_stu_2_04::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_04::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_04::item_f41fa21a": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_04::item_e058012d": {
      "level": "B3",
      "comment": ""
    },
    "sample_stu_2_04::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_04::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_04::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_05::item_03972701": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_2_05::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_05::item_f41fa21a": {
      "level": "D",
      "comment": "치약 속 플루오린 화합물을 사례로 들어 원소 이름과 실제 화합물의 성질을 구분하려는 모습을 보임"
    },
    "sample_stu_2_05::item_e058012d": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_05::item_aa648d0f": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_2_05::item_299588fe": {
      "level": "",
      "comment": "익숙한 물질을 과학적으로 다시 바라보며 구성 원소, 결합 방식, 쓰임 사이의 관계를 설명하려고 함"
    },
    "sample_stu_2_05::item_2494ee70": {
      "level": "",
      "comment": "계산값이 예상과 다르게 나온 이유를 측정 간격, 영상 프레임, 반응 시간 측면에서 나누어 검토함"
    },
    "sample_stu_2_06::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_06::item_95d74cad": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_2_06::item_f41fa21a": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_06::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_06::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_06::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_06::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_07::item_03972701": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_07::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_07::item_f41fa21a": {
      "level": "C",
      "comment": "스마트폰 화면 소재를 조사하며 투명성과 전기적 성질이 모두 요구되는 물질의 특징을 정리함"
    },
    "sample_stu_2_07::item_e058012d": {
      "level": "B1",
      "comment": ""
    },
    "sample_stu_2_07::item_aa648d0f": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_07::item_299588fe": {
      "level": "",
      "comment": "자료 조사 과정에서 용어를 그대로 옮기기보다 핵심 개념을 골라 모형과 설명에 반영하려는 모습이 보임"
    },
    "sample_stu_2_07::item_2494ee70": {
      "level": "",
      "comment": "수평 방향 운동과 연직 방향 운동을 독립적으로 분석하는 과정에서 힘이 작용하는 방향을 구분함"
    },
    "sample_stu_2_08::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_08::item_95d74cad": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_08::item_f41fa21a": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_08::item_e058012d": {
      "level": "B2",
      "comment": ""
    },
    "sample_stu_2_08::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_08::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_08::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_09::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_09::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_09::item_f41fa21a": {
      "level": "B",
      "comment": "건조제 속 실리카겔을 사례로 들어 물질의 구조가 물 흡착 성질과 관련될 수 있음을 탐색함"
    },
    "sample_stu_2_09::item_e058012d": {
      "level": "B3",
      "comment": ""
    },
    "sample_stu_2_09::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_09::item_299588fe": {
      "level": "",
      "comment": "물질을 선정한 이유가 분명하고, 원자 배치 모형과 비유적 표현을 연결해 설명하려는 흐름이 나타남"
    },
    "sample_stu_2_09::item_2494ee70": {
      "level": "",
      "comment": "그래프를 먼저 그린 뒤 운동의 특징을 설명하려고 하였고, 표와 그래프의 정보가 서로 연결됨을 확인함"
    },
    "sample_stu_2_10::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_10::item_95d74cad": {
      "level": "E",
      "comment": ""
    },
    "sample_stu_2_10::item_f41fa21a": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_2_10::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_2_10::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_2_10::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_2_10::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_01::item_03972701": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_01::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_01::item_f41fa21a": {
      "level": "A",
      "comment": "철의 부식과 스테인리스강을 비교하며 원소 조합이 물질의 안정성에 영향을 준다는 점을 정리함"
    },
    "sample_stu_3_01::item_e058012d": {
      "level": "B1",
      "comment": ""
    },
    "sample_stu_3_01::item_aa648d0f": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_01::item_299588fe": {
      "level": "",
      "comment": "복합적인 물질의 성질을 하나의 원인으로 단정하지 않고 여러 구성 요소를 함께 고려하려는 관점이 드러남"
    },
    "sample_stu_3_01::item_2494ee70": {
      "level": "",
      "comment": "낙하 물체의 운동을 일상적 낙하 사례와 비교하며 공기 저항이 결과에 미치는 가능성을 언급함"
    },
    "sample_stu_3_02::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_02::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_02::item_f41fa21a": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_02::item_e058012d": {
      "level": "B2",
      "comment": ""
    },
    "sample_stu_3_02::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_02::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_02::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_03::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_03::item_95d74cad": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_03::item_f41fa21a": {
      "level": "C",
      "comment": "화장품 성분을 조사하며 물질 이름이 복잡해도 구성 원소와 결합을 기준으로 살펴볼 수 있음을 말함"
    },
    "sample_stu_3_03::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_03::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_03::item_299588fe": {
      "level": "",
      "comment": "화합물의 이름, 구성 원소, 결합 특징을 구분해 정리하면서 물질을 체계적으로 이해하려는 모습을 보임"
    },
    "sample_stu_3_03::item_2494ee70": {
      "level": "",
      "comment": "측정값을 평균 내는 이유를 실험 오차 감소와 연결해 설명하고, 반복 측정의 필요성을 제시함"
    },
    "sample_stu_3_04::item_03972701": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_3_04::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_04::item_f41fa21a": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_3_04::item_e058012d": {
      "level": "B3",
      "comment": ""
    },
    "sample_stu_3_04::item_aa648d0f": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_3_04::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_04::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_05::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_05::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_05::item_f41fa21a": {
      "level": "A",
      "comment": "물과 과산화수소를 비교하며 비슷한 원소로 이루어진 물질도 배열에 따라 성질이 달라짐을 설명함"
    },
    "sample_stu_3_05::item_e058012d": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_05::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_05::item_299588fe": {
      "level": "",
      "comment": "일상적 소재를 교과 개념과 연결해 설명하려는 시도가 꾸준했고, 모형 표현에서 과학적 의미를 담으려 함"
    },
    "sample_stu_3_05::item_2494ee70": {
      "level": "",
      "comment": "보고서 정리 과정에서 계산 과정과 해석 문장을 분리해 적어 결과를 확인하기 쉽게 구성함"
    },
    "sample_stu_3_06::item_03972701": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_06::item_95d74cad": {
      "level": "D",
      "comment": ""
    },
    "sample_stu_3_06::item_f41fa21a": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_06::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_06::item_aa648d0f": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_06::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_06::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_07::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_07::item_95d74cad": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_07::item_f41fa21a": {
      "level": "C",
      "comment": "제설제의 원리를 찾아보고 이온 결합 물질이 물에 녹을 때 나타나는 특징과 연결함"
    },
    "sample_stu_3_07::item_e058012d": {
      "level": "B1",
      "comment": ""
    },
    "sample_stu_3_07::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_07::item_299588fe": {
      "level": "",
      "comment": "조사한 물질의 쓰임과 성질을 연결하면서 원소의 주기성과 결합 개념을 실제 사례에 적용하려는 모습이 나타남"
    },
    "sample_stu_3_07::item_2494ee70": {
      "level": "",
      "comment": "수평으로 던진 물체가 아래로 휘어지는 이유를 중력의 지속적인 작용과 연결해 설명함"
    },
    "sample_stu_3_08::item_03972701": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_08::item_95d74cad": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_08::item_f41fa21a": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_08::item_e058012d": {
      "level": "B2",
      "comment": ""
    },
    "sample_stu_3_08::item_aa648d0f": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_08::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_08::item_2494ee70": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_09::item_03972701": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_09::item_95d74cad": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_09::item_f41fa21a": {
      "level": "B",
      "comment": "단열재 소재를 조사하며 가벼움과 열전도성 같은 성질이 물질 선택 기준이 될 수 있음을 제시함"
    },
    "sample_stu_3_09::item_e058012d": {
      "level": "B3",
      "comment": ""
    },
    "sample_stu_3_09::item_aa648d0f": {
      "level": "B",
      "comment": ""
    },
    "sample_stu_3_09::item_299588fe": {
      "level": "",
      "comment": "화학 결합을 비유적으로 표현하는 과정에서 핵심 원리를 놓치지 않으려는 시도가 보이며 설명의 맥락이 분명함"
    },
    "sample_stu_3_09::item_2494ee70": {
      "level": "",
      "comment": "속도 변화, 가속도, 힘의 방향을 함께 고려하며 중력을 받는 물체의 운동을 분석하려는 모습이 드러남"
    },
    "sample_stu_3_10::item_03972701": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_10::item_95d74cad": {
      "level": "E",
      "comment": ""
    },
    "sample_stu_3_10::item_f41fa21a": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_10::item_e058012d": {
      "level": "A",
      "comment": ""
    },
    "sample_stu_3_10::item_aa648d0f": {
      "level": "C",
      "comment": ""
    },
    "sample_stu_3_10::item_299588fe": {
      "level": "",
      "comment": ""
    },
    "sample_stu_3_10::item_2494ee70": {
      "level": "",
      "comment": ""
    }
  },
  "results": {},
  "saved_at": "2026-06-23T07:35:00",
  "app": "개꿀 생기부",
  "version": "sample-project-v61"
}''')
    data["saved_at"] = datetime.now().isoformat(timespec="seconds")
    data["version"] = "sample-project-v61"
    return data


def apply_project_data(data):
    """
    JSON 파일이나 내장 샘플 프로젝트 dict를 현재 세션에 반영한다.
    """
    st.session_state.settings = data.get("settings", st.session_state.settings)
    st.session_state.students = pd.DataFrame(data.get("students", []))
    st.session_state.assessments = data.get("assessments", [])
    st.session_state["items"] = data.get("items", [])
    st.session_state.records = data.get("records", {})
    st.session_state.results = data.get("results", {})
    if "selected_generation_student_id" in st.session_state:
        del st.session_state["selected_generation_student_id"]
    sanitize_state()


def load_sample_data():
    sample_data = build_sample_project_data()
    apply_project_data(sample_data)


# =========================
# 앱 UI
# =========================
st.markdown('<div id="big-honey-top"></div>', unsafe_allow_html=True)
st.title(APP_TITLE)
st.caption(APP_SUBTITLE)

with st.sidebar:
    st.header("작업 관리")

    uploaded_project = st.file_uploader(
        "프로젝트 JSON 불러오기",
        type=["json"],
        help="이전에 저장한 BigHoneySangkibu 프로젝트 파일을 다시 불러옵니다.",
    )

    if uploaded_project and st.button("프로젝트 불러오기"):
        load_project_json(uploaded_project)
        st.success("프로젝트를 불러왔습니다.")
        st.rerun()

    st.download_button(
        "현재 프로젝트 JSON 저장",
        data=project_to_json(),
        file_name="BigHoneySangkibu_project.json",
        mime="application/json",
    )

    st.divider()

    if st.button("샘플 데이터 불러오기", help="고등학교 통합과학 예시 학생 3개 반, 반별 10명과 학생별 기록이 들어간 샘플 프로젝트를 불러옵니다."):
        load_sample_data()
        st.success("고등학교 통합과학 예시 학생 3개 반과 학생별 기록이 들어간 샘플 데이터를 불러왔습니다.")
        st.rerun()

    if st.button("전체 초기화", type="secondary"):
        for key in ["settings", "students", "assessments", "items", "records", "results", "generation_job", "selected_generation_student_id"]:
            if key in st.session_state:
                del st.session_state[key]
        init_state()
        sanitize_state()
        st.success("초기화했습니다.")
        st.rerun()


STEP_LABELS = [
    "① 기본 설정",
    "② 학생 명단 업로드",
    "③ 수행평가 설계",
    "④ 학생별 기록 입력",
    "⑤ API 자료 확인",
    "⑥ 생기부 생성/다운로드",
]

NAV_WIDGET_KEY = "step_nav_radio_v28"
PENDING_STEP_KEY = "pending_step_index_v28"
SCROLL_TO_TOP_KEY = "scroll_to_top_after_step_change_v28"


if "current_step" not in st.session_state:
    st.session_state["current_step"] = 0

try:
    st.session_state["current_step"] = int(st.session_state.get("current_step", 0))
except Exception:
    st.session_state["current_step"] = 0

# 다음 단계 버튼은 URL을 바꾸지 않고 세션 상태만 변경한다.
# URL 이동은 새로고침처럼 동작해 입력 중인 정보가 날아갈 수 있으므로 사용하지 않는다.
programmatic_step_change = False
if PENDING_STEP_KEY in st.session_state:
    try:
        st.session_state["current_step"] = int(st.session_state[PENDING_STEP_KEY])
    except Exception:
        st.session_state["current_step"] = 0
    del st.session_state[PENDING_STEP_KEY]
    st.session_state[SCROLL_TO_TOP_KEY] = True
    programmatic_step_change = True

st.session_state["current_step"] = max(0, min(st.session_state["current_step"], len(STEP_LABELS) - 1))

# 중요: 단계 탭 라디오는 사용자가 직접 클릭해서도 이동할 수 있어야 한다.
# 따라서 매 실행마다 NAV_WIDGET_KEY를 강제로 덮어쓰면 클릭값이 무시된다.
# 처음 실행하거나, 하단 '다음 단계' 버튼으로 이동한 경우에만 위젯 값을 동기화한다.
if NAV_WIDGET_KEY not in st.session_state or programmatic_step_change:
    st.session_state[NAV_WIDGET_KEY] = STEP_LABELS[st.session_state["current_step"]]


def request_step_change(next_index: int):
    """다음 단계 이동 요청을 저장한다. 실제 반영은 다음 rerun의 위쪽에서 처리한다."""
    st.session_state[PENDING_STEP_KEY] = int(next_index)


def scroll_page_to_top_once():
    """다음 단계 이동 직후 새로고침 없이 화면을 앱 맨 위로 올린다."""
    if not st.session_state.get(SCROLL_TO_TOP_KEY, False):
        return

    st.session_state[SCROLL_TO_TOP_KEY] = False
    components.html(
        """
        <script>
        function forceScrollTop() {
            try {
                const parentWindow = window.parent;
                const parentDoc = parentWindow.document;

                try { parentWindow.scrollTo(0, 0); } catch (e) {}
                try { parentDoc.documentElement.scrollTop = 0; } catch (e) {}
                try { parentDoc.body.scrollTop = 0; } catch (e) {}

                const selectors = [
                    'section.main',
                    'main',
                    '.main',
                    '.block-container',
                    '[data-testid="stAppViewContainer"]',
                    '[data-testid="stMain"]',
                    '[data-testid="stMainBlockContainer"]'
                ];

                selectors.forEach(function(selector) {
                    parentDoc.querySelectorAll(selector).forEach(function(el) {
                        try { el.scrollTop = 0; } catch (e) {}
                        try { el.scrollTo(0, 0); } catch (e) {}
                    });
                });

                // Streamlit 버전에 따라 실제 스크롤 컨테이너가 달라질 수 있어
                // 스크롤 가능한 주요 요소를 한 번 더 찾아서 맨 위로 올린다.
                parentDoc.querySelectorAll('section, main, div').forEach(function(el) {
                    try {
                        if (el.scrollHeight > el.clientHeight + 80) {
                            el.scrollTop = 0;
                        }
                    } catch (e) {}
                });
            } catch (e) {}
        }

        forceScrollTop();
        setTimeout(forceScrollTop, 50);
        setTimeout(forceScrollTop, 150);
        setTimeout(forceScrollTop, 350);
        setTimeout(forceScrollTop, 700);
        setTimeout(forceScrollTop, 1100);
        </script>
        """,
        height=0,
    )

st.markdown(
    """
    <style>
    /* 기존 st.tabs처럼 보이도록 단계 선택 라디오를 탭 형태로 정리한다. */
    div[data-testid="stRadio"] > div[role="radiogroup"] {
        display: flex !important;
        flex-wrap: wrap !important;
        gap: 0 !important;
        border-bottom: 1px solid #D0D7DE !important;
        margin-bottom: 1.15rem !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] label {
        min-height: 42px !important;
        padding: 0.62rem 0.95rem !important;
        margin-right: 4px !important;
        margin-bottom: -1px !important;
        border: 1px solid #D0D7DE !important;
        border-bottom: none !important;
        border-radius: 10px 10px 0 0 !important;
        background: #F6F8FA !important;
        cursor: pointer !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] label:has(input:checked) {
        background: #FFFFFF !important;
        border-top: 3px solid #D92D20 !important;
        color: #111827 !important;
        font-weight: 800 !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] label p {
        font-weight: 700 !important;
        font-size: 0.95rem !important;
    }
    div[data-testid="stRadio"] > div[role="radiogroup"] input {
        display: none !important;
    }


    /* 수행평가 설계 화면 계층 구분
       주의: 수행평가 expander 안에 평가요소 expander가 들어가므로, :has(.item-card-content)는 바깥 수행평가에도 같이 걸릴 수 있다.
       그래서 평가요소(하위) 스타일을 먼저 선언하고, 수행평가(상위) 파란 스타일을 뒤에서 다시 덮어쓴다. */

    /* 하위 박스: 평가요소는 노란색/주황색 계열 */
    div[data-testid="stExpander"] details:has(.item-card-content) {
        background: linear-gradient(180deg, #FFFBF5 0%, #FFF7ED 100%) !important;
        border: 2px solid #FED7AA !important;
        border-radius: 16px !important;
        box-shadow: 0 4px 12px rgba(234, 88, 12, 0.07) !important;
        padding: 0.12rem 0.3rem 0.32rem 0.3rem !important;
        margin: 0.65rem 0 0.95rem 0 !important;
    }
    div[data-testid="stExpander"] details:has(.item-card-content) > summary {
        background: #FFEDD5 !important;
        border: 1px solid #FED7AA !important;
        border-radius: 12px !important;
        margin: 0.16rem 0 0.5rem 0 !important;
        padding: 0.12rem 0.4rem !important;
    }
    div[data-testid="stExpander"] details:has(.item-card-content) > summary p {
        color: #111827 !important;
        font-weight: 800 !important;
    }
    div[data-testid="stExpander"] details:has(.item-card-content) .item-card-content {
        display: none !important;
    }

    /* 상위 박스: 수행평가는 연한 파란색 계열. 아래 선언이 나중에 오므로 바깥 수행평가 박스는 확실히 파란색으로 보인다. */
    div[data-testid="stExpander"] details:has(.assessment-card-content) {
        background: linear-gradient(180deg, #EFF6FF 0%, #DBEAFE 100%) !important;
        border: 2px solid #93C5FD !important;
        border-radius: 18px !important;
        box-shadow: 0 5px 15px rgba(37, 99, 235, 0.10) !important;
        padding: 0.18rem 0.38rem 0.42rem 0.38rem !important;
        margin: 0.9rem 0 1.25rem 0 !important;
    }
    div[data-testid="stExpander"] details:has(.assessment-card-content) > summary {
        background: #DBEAFE !important;
        border: 1px solid #BFDBFE !important;
        border-radius: 14px !important;
        margin: 0.2rem 0 0.55rem 0 !important;
        padding: 0.15rem 0.45rem !important;
    }
    div[data-testid="stExpander"] details:has(.assessment-card-content) > summary p {
        color: #0F172A !important;
        font-weight: 900 !important;
    }
    div[data-testid="stExpander"] details:has(.assessment-card-content) .assessment-card-content {
        display: none !important;
    }


    a.next-step-link {
        display: flex !important;
        width: 100% !important;
        min-height: 46px !important;
        align-items: center !important;
        justify-content: center !important;
        background: #D92D20 !important;
        color: #FFFFFF !important;
        text-decoration: none !important;
        border-radius: 10px !important;
        font-weight: 800 !important;
        border: 1px solid #B42318 !important;
        box-shadow: 0 1px 2px rgba(16, 24, 40, 0.08) !important;
        margin-top: 0.65rem !important;
        margin-bottom: 0.5rem !important;
    }
    a.next-step-link:hover {
        background: #B42318 !important;
        color: #FFFFFF !important;
        text-decoration: none !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown("### 작업 단계")
selected_step_label = st.radio(
    "이동할 단계를 선택하세요.",
    STEP_LABELS,
    horizontal=True,
    label_visibility="collapsed",
    key=NAV_WIDGET_KEY,
)
st.session_state["current_step"] = STEP_LABELS.index(selected_step_label)
current_step = st.session_state["current_step"]
scroll_page_to_top_once()


def render_next_step_button(current_index: int):
    """각 단계 하단에서 다음 단계로 이동한다. 마지막 단계에서는 첫 단계로 순환한다."""
    next_index = (current_index + 1) % len(STEP_LABELS)
    next_label = STEP_LABELS[next_index]

    st.divider()
    st.button(
        f"다음 단계로 넘어가기 → {next_label}",
        type="primary",
        use_container_width=True,
        key=f"next_step_button_{current_index}",
        on_click=request_step_change,
        args=(next_index,),
    )


# =========================
# ① 기본 설정
# =========================
if current_step == 0:
    st.subheader("① 기본 설정")

    settings = st.session_state.settings

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        settings["school_year"] = st.text_input("학년도", value=clean_text(settings.get("school_year", "2026")))
    with col2:
        settings["semester"] = st.selectbox(
            "학기",
            ["1학기", "2학기"],
            index=1 if settings.get("semester") == "2학기" else 0,
        )
    with col3:
        school_options = ["중학교", "고등학교", "초등학교"]
        current_school = settings.get("school_level", "중학교")
        settings["school_level"] = st.selectbox(
            "학교급",
            school_options,
            index=school_options.index(current_school) if current_school in school_options else 0,
        )
    with col4:
        settings["grade"] = st.text_input("학년", value=clean_text(settings.get("grade", "1")))

    col5, col6, col7 = st.columns(3)
    with col5:
        settings["subject"] = st.text_input("과목명", value=clean_text(settings.get("subject", "과학")))
    with col6:
        settings["target_bytes_min"] = st.number_input(
            "목표 최소 byte",
            min_value=100,
            max_value=3000,
            value=int(settings.get("target_bytes_min", 700)),
            step=50,
        )
    with col7:
        settings["target_bytes_max"] = st.number_input(
            "목표 최대 byte",
            min_value=100,
            max_value=3000,
            value=int(settings.get("target_bytes_max", 800)),
            step=50,
        )

    st.markdown("#### 생기부 작성 규칙")
    st.caption("여기에 적은 규칙은 API 프롬프트의 맨 앞부분에 모든 학생 공통 규칙으로 들어갑니다.")

    settings["custom_rules"] = st.text_area(
        "공통 작성 규칙",
        value=settings.get("custom_rules", DEFAULT_RULES),
        height=220,
        help="예: 명사형 종결, 금지 표현, 학생 이름 제외, 분량, 문체, 과목 특성 등을 자유롭게 적으세요.",
    )

    with st.expander("API 프롬프트에 들어가는 규칙 예시 보기"):
        st.code(settings["custom_rules"], language="text")

    st.info("작업을 이어서 하려면 왼쪽의 '현재 프로젝트 JSON 저장'을 눌러 파일로 저장하세요.")


    render_next_step_button(0)


# =========================
# ② 학생 명단 업로드
# =========================
if current_step == 1:
    st.subheader("② 학생 명단 업로드")

    st.markdown(
        """
        나이스 세특 파일에서 `학년`, `반/번호`, `성명`을 자동 추출합니다.  
        여러 반 파일을 한 번에 올리거나, 파일을 나중에 하나씩 추가해도 기존 명단에 누적할 수 있습니다.
        """
    )

    st.info("아래의 '업로드 파일 미리보기'는 아직 최종 명단이 아닙니다. 최종 명단은 하단의 '현재 학생 명단'이며, 그 표에서 직접 수정할 수 있습니다.")

    uploaded_excels = st.file_uploader(
        "나이스 엑셀 파일 업로드",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )

    st.info("나이스 파일 다운로드 경로: [성적처리] → [과목별세부능력특기사항] → 강의실 선택 후 [조회] → [엑셀 다운로드]")

    parsed_students = pd.DataFrame(columns=["student_id", "학년", "반", "번호", "성명"])
    parsed_info = []

    if uploaded_excels:
        parsed_frames = []

        for file in uploaded_excels:
            try:
                students_df, sheet_names = parse_neis_excel(file)
                if not students_df.empty:
                    students_df["원본파일"] = file.name
                    parsed_frames.append(students_df)
                    parsed_info.append(f"{file.name}: {len(students_df)}명 / 시트 {', '.join(sheet_names)}")
                else:
                    parsed_info.append(f"{file.name}: 학생 명단을 찾지 못함")
            except Exception as e:
                parsed_info.append(f"{file.name}: 오류 - {e}")

        if parsed_frames:
            parsed_students = pd.concat(parsed_frames, ignore_index=True)
            parsed_students = parsed_students.drop(columns=["원본파일"], errors="ignore")
            parsed_students = sort_students_df(parsed_students)

        for info in parsed_info:
            st.caption(info)

        if not parsed_students.empty:
            st.markdown("#### 업로드 파일 미리보기")
            st.caption("이 표는 업로드한 파일에서 읽어온 임시 명단입니다. 아래 버튼을 눌러야 현재 학생 명단에 반영됩니다.")

            st.dataframe(
                parsed_students.drop(columns=["student_id"], errors="ignore"),
                use_container_width=True,
                height=260,
            )

            review_students = parsed_students[parsed_students["성명"].map(needs_name_review)].copy()
            adjusted_parsed_students = parsed_students.copy()

            if not review_students.empty:
                st.markdown("#### 이름 확인 필요 학생")
                st.warning(
                    "성명에 한글 외 문자가 포함된 학생만 따로 모았습니다. "
                    "필요하면 여기서 이름을 수정한 뒤 아래의 추가/교체 버튼을 누르세요. "
                    "수정이 필요 없으면 그냥 넘어가도 됩니다."
                )

                edited_review = st.data_editor(
                    review_students,
                    use_container_width=True,
                    height=220,
                    num_rows="fixed",
                    disabled=["student_id"],
                    column_config={
                        "student_id": None,
                    },
                    key="name_review_editor",
                )

                adjusted_parsed_students = apply_name_review_edits(parsed_students, edited_review)
            else:
                st.success("한글 외 문자가 포함된 성명은 발견되지 않았습니다.")

            col_a, col_b, col_blank = st.columns([3.4, 3.4, 3.2])
            with col_a:
                if st.button("업로드 명단을 현재 명단에 추가", type="primary", use_container_width=True):
                    st.session_state.students = combine_students(st.session_state.students, adjusted_parsed_students)
                    st.success("업로드 명단을 현재 학생 명단에 추가했습니다.")
                    st.rerun()

            with col_b:
                if st.button("현재 명단을 업로드 명단으로 교체", type="primary", use_container_width=True):
                    st.session_state.students = sort_students_df(adjusted_parsed_students)
                    st.success("현재 학생 명단을 업로드 명단으로 교체했습니다.")
                    st.rerun()

    st.divider()
    st.markdown("#### 현재 학생 명단")
    st.caption("이 표가 최종 명단입니다. 여기서 직접 수정한 뒤 '현재 학생 명단 저장'을 누르면 이후 모든 입력 화면에 반영됩니다. API 입력 전 개인정보 보호가 필요하면 저장 후 '이름 블라인드 처리'를 누르세요.")

    editable_students = st.session_state.students.copy()

    if editable_students.empty:
        editable_students = pd.DataFrame(
            [
                {
                    "student_id": make_id("stu"),
                    "학년": st.session_state.settings.get("grade", "1"),
                    "반": "1",
                    "번호": "1",
                    "성명": "",
                }
            ]
        )

    for col in ["student_id", "학년", "반", "번호", "성명"]:
        if col not in editable_students.columns:
            editable_students[col] = ""

    edited_students = st.data_editor(
        editable_students[["student_id", "학년", "반", "번호", "성명"]],
        num_rows="dynamic",
        use_container_width=True,
        height=360,
        key="final_students_editor",
        column_config={
            "student_id": None,
            "학년": st.column_config.TextColumn("학년", width="small"),
            "반": st.column_config.TextColumn("반", width="small"),
            "번호": st.column_config.TextColumn("번호", width="small"),
            "성명": st.column_config.TextColumn("성명", width="medium"),
        },
    )

    # 버튼은 왼쪽부터 '반/번호 순대로 다시 정렬 → 이름 블라인드 처리 → 저장' 순서로 붙여 배치한다.
    col_sort, col_mask, col_save, col_blank = st.columns([2.25, 2.25, 1.45, 4.05])

    with col_sort:
        if st.button("반/번호 순대로 다시 정렬"):
            st.session_state.students = sort_students_df(st.session_state.students)
            st.success("현재 학생 명단을 반/번호 순대로 정렬했습니다.")
            st.rerun()

    with col_mask:
        if st.button("🔒 이름 블라인드 처리", type="primary", help="현재 학생 명단의 성명을 홍*동, 홍*, 홍**동 형식으로 바꿉니다."):
            if st.session_state.students.empty:
                st.warning("블라인드 처리할 학생 명단이 없습니다.")
            else:
                st.session_state.students = mask_student_names_in_df(st.session_state.students, mask_char="*")
                st.warning("학생 이름을 블라인드 처리했습니다. 원래 이름으로 되돌리려면 암호화 전 저장한 JSON 또는 나이스 파일을 다시 불러와야 합니다.")
                st.rerun()

    with col_save:
        if st.button("현재 명단 저장", type="primary"):
            new_df = edited_students.copy()

            for col in ["student_id", "학년", "반", "번호", "성명"]:
                if col not in new_df.columns:
                    new_df[col] = ""

            new_df = new_df[new_df["성명"].astype(str).str.strip() != ""].reset_index(drop=True)

            # 기존 행은 student_id를 유지하고, 새로 추가된 행만 새 ID를 부여한다.
            new_df["student_id"] = new_df["student_id"].apply(
                lambda x: clean_text(x) if clean_text(x) else make_id("stu")
            )

            st.session_state.students = sort_students_df(new_df[["student_id", "학년", "반", "번호", "성명"]])
            st.success("현재 학생 명단을 저장했습니다.")
            st.rerun()

    if not st.session_state.students.empty:
        review_current = st.session_state.students[st.session_state.students["성명"].map(needs_name_review)]
        if not review_current.empty:
            with st.expander(f"현재 명단의 이름 확인 필요 학생 {len(review_current)}명 보기"):
                st.dataframe(
                    review_current.drop(columns=["student_id"], errors="ignore"),
                    use_container_width=True,
                    height=180,
                )


    render_next_step_button(1)


# =========================
# ③ 수행평가 설계
# =========================
if current_step == 2:
    st.subheader("③ 수행평가 설계")

    st.markdown(
        """
        이 화면은 **수행평가**를 먼저 만들고, 각 수행평가 안에 학생별로 입력할 **평가 요소**를 넣는 구조입니다.  
        평가 요소 안에는 필요에 따라 **성취수준 코드, 평가 문구, 개별 코멘트**를 설정합니다.
        """
    )

    with st.expander("➕ 새 수행평가 추가", expanded=True):
        st.caption("먼저 상위 단위인 수행평가를 만들고, 그 안에 평가 요소를 추가합니다.")

        with st.form("add_assessment_form"):
            col1, col2 = st.columns([2, 1])
            with col1:
                new_assessment_name = st.text_input("수행평가명", placeholder="예: 생태지도 만들기")
                new_area = st.text_input("영역/단원", placeholder="예: 생물과 환경")
            with col2:
                st.caption("새로 추가한 수행평가는 기본적으로 사용 상태로 저장됩니다.")
                st.caption("순서는 아래의 드래그 정렬에서 바꿀 수 있습니다.")

            new_desc = st.text_area(
                "성취기준 / 수행평가 설명",
                placeholder="예: 생태지도 제작 수행평가의 성취기준과 평가 내용을 입력하세요",
                height=80,
            )

            submitted = st.form_submit_button("수행평가 추가")
            if submitted:
                if not new_assessment_name.strip():
                    st.warning("수행평가명을 입력하세요.")
                else:
                    st.session_state.assessments.append(
                        {
                            "assessment_id": make_id("assess"),
                            "name": new_assessment_name.strip(),
                            "area": new_area.strip(),
                            "description": new_desc.strip(),
                            "order": len(st.session_state.assessments) + 1,
                            "use": True,
                        }
                    )
                    sanitize_state()
                    st.success("수행평가를 추가했습니다.")
                    st.rerun()

    st.divider()
    st.markdown("### 📚 AI가 참고할 수행평가 자료")

    if not st.session_state.assessments:
        st.info("아직 등록된 수행평가가 없습니다. 먼저 수행평가를 추가하세요.")

    normalize_assessment_orders()
    sorted_assessments = sorted(
        st.session_state.assessments,
        key=lambda x: int(x.get("order", 999) or 999),
    )

    if len(sorted_assessments) >= 2:
        with st.expander("수행평가 순서 드래그 정렬", expanded=True):
            if sort_items is None:
                st.warning("드래그 정렬 기능을 사용하려면 requirements.txt에 streamlit-sortables를 추가해야 합니다.")
            else:
                st.caption("수행평가명을 마우스로 잡고 위아래로 옮긴 뒤 저장하세요. 항목을 추가하면 이 박스가 자동으로 새로 생성됩니다.")
                assessment_labels = [
                    f"{idx}. {assessment.get('name', '이름 없는 수행평가')}"
                    for idx, assessment in enumerate(sorted_assessments, start=1)
                ]
                st.caption("현재 수행평가 순서: " + " → ".join(assessment_labels))
                label_to_assessment_id = {
                    label: assessment.get("assessment_id", "")
                    for label, assessment in zip(assessment_labels, sorted_assessments)
                }

                sorted_labels = sort_labels_with_gray_box(
                    assessment_labels,
                    key="assessment_drag_sort",
                    header="수행평가 순서",
                )

                if st.button("수행평가 순서 저장"):
                    apply_assessment_drag_order(sorted_labels, label_to_assessment_id)
                    normalize_assessment_orders()
                    st.success("수행평가 순서를 저장했습니다.")
                    st.rerun()

    rubric_updates = {}

    for assess_index, assessment in enumerate(sorted_assessments, start=1):
        aid = assessment.get("assessment_id", "")
        normalize_item_orders(aid)
        existing_items = sorted(
            get_items_for_assessment(aid),
            key=lambda x: int(x.get("order", 999) or 999),
        )

        status_badge = "사용" if assessment.get("use", True) else "미사용"
        area_text = assessment.get("area", "") or "영역/단원 미입력"
        item_count = len(existing_items)

        assessment_expander_title = (
            f"📁 수행평가 {assess_index}. {assessment.get('name', '이름 없는 수행평가')} "
            f"· 평가 요소 {item_count}개 · {status_badge}"
        )
        with st.expander(assessment_expander_title, expanded=True):
            st.markdown('<div class="assessment-card-content"></div>', unsafe_allow_html=True)
            st.markdown(
                f"""
                ### 📁 수행평가 {assess_index}. {assessment.get('name', '이름 없는 수행평가')}
                **영역/단원:** {area_text} &nbsp;&nbsp;|&nbsp;&nbsp;
                **평가 요소:** {item_count}개 &nbsp;&nbsp;|&nbsp;&nbsp;
                **상태:** {status_badge}
                """
            )

            if assessment.get("description"):
                st.caption(f"수행평가 설명: {assessment.get('description', '')}")
            else:
                st.caption("수행평가 설명이 아직 입력되지 않았습니다.")

            with st.expander("⚙️ 수행평가 기본 정보 수정 / 삭제", expanded=False):
                col1, col2, col3 = st.columns([2, 2, 1])

                with col1:
                    assessment["name"] = st.text_input(
                        "수행평가명 수정",
                        value=assessment.get("name", ""),
                        key=f"assess_name_{aid}",
                    )
                    assessment["area"] = st.text_input(
                        "영역/단원 수정",
                        value=assessment.get("area", ""),
                        key=f"assess_area_{aid}",
                    )

                with col2:
                    assessment["description"] = st.text_area(
                        "수행평가 설명 수정",
                        value=assessment.get("description", ""),
                        key=f"assess_desc_{aid}",
                        height=120,
                    )

                with col3:
                    st.caption("순서는 수행평가 목록 위의 드래그 정렬에서 변경합니다.")
                    assessment["use"] = st.checkbox(
                        "사용",
                        value=assessment.get("use", True),
                        key=f"assess_use_{aid}",
                    )
                    st.caption("사용 해제 시 이 수행평가는 학생별 기록 입력, API 입력 자료, 생기부 생성에서 제외됩니다.")
                    if st.button("수행평가 삭제", key=f"delete_assessment_{aid}"):
                        item_ids = [it.get("item_id", "") for it in get_items_for_assessment(aid)]
                        st.session_state.assessments = [
                            x for x in st.session_state.assessments
                            if x.get("assessment_id", "") != aid
                        ]
                        st.session_state["items"] = [
                            x for x in st.session_state["items"]
                            if x.get("assessment_id", "") != aid
                        ]
                        st.session_state.records = {
                            k: v for k, v in st.session_state.records.items()
                            if k.split("::")[-1] not in item_ids
                        }
                        st.success("수행평가를 삭제했습니다.")
                        st.rerun()

            st.markdown("#### 🧾 평가 요소")

            if existing_items:
                if len(existing_items) == 1:
                    st.caption("평가 요소가 1개라서 드래그 정렬 박스는 표시하지 않습니다. 평가 요소를 2개 이상 만들면 이곳에 드래그 정렬 박스가 나타납니다.")

                if len(existing_items) >= 2:
                    with st.expander("평가 요소 순서 드래그 정렬", expanded=True):
                        if sort_items is None:
                            st.warning("드래그 정렬 기능을 사용하려면 requirements.txt에 streamlit-sortables를 추가해야 합니다.")
                        else:
                            st.caption("평가 요소명을 마우스로 잡고 위아래로 옮긴 뒤 저장하세요. 항목을 추가하면 이 박스가 자동으로 새로 생성됩니다.")
                            item_labels = [
                                f"{idx}. {item.get('name', '이름 없는 평가 요소')}"
                                for idx, item in enumerate(existing_items, start=1)
                            ]
                            st.caption("현재 평가 요소 순서: " + " → ".join(item_labels))
                            label_to_item_id = {
                                label: item.get("item_id", "")
                                for label, item in zip(item_labels, existing_items)
                            }

                            sorted_item_labels = sort_labels_with_gray_box(
                                item_labels,
                                key=f"item_drag_sort_{aid}",
                                header="평가 요소 순서",
                            )

                            if st.button("평가 요소 순서 저장", key=f"save_item_drag_sort_{aid}"):
                                apply_item_drag_order(aid, sorted_item_labels, label_to_item_id)
                                normalize_item_orders(aid)
                                st.success("평가 요소 순서를 저장했습니다.")
                                st.rerun()

                for item_index, item in enumerate(existing_items, start=1):
                    item_id = item.get("item_id", "")
                    item_type_label = item_type_to_kor(item.get("type", "rubric"))
                    level_count = len(item.get("levels", [])) if item.get("type") != "comment" else 0

                    item_expander_title = (
                        f"🧾 평가 요소 {item_index}. {item.get('name', '이름 없는 평가 요소')} "
                        f"· {item_type_label}"
                    )
                    with st.expander(item_expander_title, expanded=True):
                        st.markdown('<div class="item-card-content"></div>', unsafe_allow_html=True)
                        st.markdown(
                            f"##### 🧾 평가 요소 {item_index}. {item.get('name', '이름 없는 평가 요소')}"
                        )

                        col1, col2, col3 = st.columns([2.2, 1.6, 1])

                        with col1:
                            item["name"] = st.text_input(
                                "항목명",
                                value=item.get("name", ""),
                                key=f"item_name_{item_id}",
                            )

                        with col2:
                            current_type_label = item_type_to_kor(item.get("type", "rubric"))
                            type_options = ["성취도 선택형", "개별 코멘트형", "성취도 + 추가 코멘트형"]
                            new_type_label = st.selectbox(
                                "기록 방식",
                                type_options,
                                index=type_options.index(current_type_label) if current_type_label in type_options else 0,
                                key=f"item_type_{item_id}",
                            )
                            new_type = item_type_from_kor(new_type_label)

                            if new_type != item.get("type"):
                                item["type"] = new_type
                                if new_type == "comment":
                                    item["levels"] = []
                                    item["rubrics"] = {}
                                elif not item.get("levels"):
                                    item["levels"] = ["A", "B", "C", "D", "E"]
                                    item["rubrics"] = {level: "" for level in item["levels"]}

                        with col3:
                            st.caption(f"현재 {item_index}번째")
                            if st.button("평가 요소 삭제", key=f"delete_item_{item_id}"):
                                st.session_state["items"] = [
                                    x for x in st.session_state["items"]
                                    if x.get("item_id", "") != item_id
                                ]
                                st.session_state.records = {
                                    k: v for k, v in st.session_state.records.items()
                                    if not k.endswith(f"::{item_id}")
                                }
                                normalize_item_orders(aid)
                                st.success("평가 요소를 삭제했습니다.")
                                st.rerun()

                        if item.get("type") in ["rubric", "rubric_plus"]:
                            levels, rubrics = render_rubric_input_block(
                                prefix=f"edit_item_rubric_{item_id}",
                                current_levels=item.get("levels", []),
                                current_rubrics=item.get("rubrics", {}),
                            )
                            rubric_updates[item_id] = {
                                "levels": levels,
                                "rubrics": rubrics,
                            }
                            st.caption("성취수준/평가 문구는 화면 맨 아래의 전체 저장 버튼으로 한꺼번에 저장됩니다.")
                        else:
                            st.info("개별 코멘트형 항목입니다. 학생별 기록 입력 화면에서 학생별 서술형 코멘트를 입력합니다.")
            else:
                st.info("아직 이 수행평가에 등록된 평가 요소가 없습니다. 아래의 '평가 요소 추가'를 눌러 항목을 추가하세요.")

            st.divider()
            render_add_item_expander(aid, item_count)

    if rubric_updates:
        st.divider()
        if st.button("전체 성취수준/평가 문구 한꺼번에 저장", type="primary", use_container_width=True):
            saved_count = 0
            for item in st.session_state["items"]:
                item_id = item.get("item_id", "")
                if item_id in rubric_updates:
                    item["levels"] = rubric_updates[item_id]["levels"]
                    item["rubrics"] = rubric_updates[item_id]["rubrics"]
                    saved_count += 1
            st.success(f"성취수준/평가 문구를 {saved_count}개 평가 요소에 한꺼번에 저장했습니다.")
            st.rerun()


    render_next_step_button(2)


# =========================
# ④ 학생별 기록 입력
# =========================
if current_step == 3:
    st.subheader("④ 학생별 기록 입력")

    if st.session_state.students.empty:
        st.warning("먼저 학생 명단을 업로드하거나 입력하세요.")

    elif not st.session_state["items"]:
        st.warning("먼저 수행평가와 평가 요소를 추가하세요.")

    else:
        students = st.session_state.students.copy()
        assessments = [a for a in st.session_state.assessments if a.get("use", True)]
        assessments = sorted(assessments, key=lambda x: int(x.get("order", 999) or 999))
        assess_options = ["전체"] + [a.get("name", "") for a in assessments]
        selected_assess_name = st.selectbox("수행평가 필터", assess_options)

        if selected_assess_name == "전체":
            selected_assess_ids = [a.get("assessment_id", "") for a in assessments]
        else:
            selected_assess_ids = [
                a.get("assessment_id", "")
                for a in assessments
                if a.get("name", "") == selected_assess_name
            ]

        selected_items = [
            item for item in st.session_state["items"]
            if isinstance(item, dict) and item.get("assessment_id", "") in selected_assess_ids
        ]
        selected_items = sorted(
            selected_items,
            key=lambda x: (get_assessment_name(x.get("assessment_id", "")), int(x.get("order", 999) or 999)),
        )

        if not selected_items:
            st.info("선택한 수행평가에 평가 요소가 없습니다.")
        else:
            st.markdown("#### 엑셀로 입력하기")
            st.caption(
                "현재 필터 기준으로 학생별 기록 입력 양식을 다운로드할 수 있습니다. "
                "엑셀에서 성취수준은 드롭다운으로 선택하고, 코멘트를 입력한 뒤 다시 업로드하면 웹앱 기록에 반영됩니다."
            )

            excel_file = make_student_record_excel(students, selected_items, selected_assess_name)
            col_download, col_upload = st.columns([1, 1])

            with col_download:
                st.download_button(
                    "학생별 기록 입력 양식 다운로드",
                    data=excel_file,
                    file_name=f"BigHoneySangkibu_student_records_{selected_assess_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            with col_upload:
                uploaded_record_excel = st.file_uploader(
                    "입력 완료 엑셀 업로드",
                    type=["xlsx"],
                    key=f"record_excel_upload_{selected_assess_name}",
                    help="이 화면에서 다운로드한 학생별 기록 입력 양식을 다시 업로드하세요.",
                )

            if uploaded_record_excel is not None:
                if st.button("업로드한 엑셀 기록 반영", type="primary", use_container_width=True):
                    try:
                        saved_count, warnings = import_student_record_excel(uploaded_record_excel)
                        if saved_count > 0:
                            st.success(f"엑셀 기록을 반영했습니다. 반영된 입력값: {saved_count}개")
                        else:
                            st.warning("반영된 입력값이 없습니다.")

                        if warnings:
                            with st.expander(f"확인 필요 항목 {len(warnings)}개 보기", expanded=True):
                                for warning in warnings[:80]:
                                    st.warning(warning)
                                if len(warnings) > 80:
                                    st.caption(f"나머지 {len(warnings) - 80}개 항목은 생략했습니다.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"엑셀 기록을 반영하는 중 오류가 발생했습니다: {e}")

            st.divider()
            st.markdown("#### 웹에서 직접 입력하기")
            st.caption(
                "엑셀 양식과 같은 전치형 구조입니다. 학생은 행으로, 수행평가·평가요소·입력구분은 열로 배치됩니다. "
                "성취수준 열은 ③ 수행평가 설계에서 정한 코드나 기호를 선택하거나 입력하세요."
            )

            matrix_df, record_columns = build_record_matrix_df(students, selected_items)
            visible_df = matrix_df.drop(columns=["_student_id"], errors="ignore")

            column_config = {
                "학년": st.column_config.TextColumn("학년", width="small"),
                "반": st.column_config.TextColumn("반", width="small"),
                "번호": st.column_config.TextColumn("번호", width="small"),
                "성명": st.column_config.TextColumn("성명", width="medium"),
            }
            for record_col in record_columns:
                label = record_col["column_label"]
                if record_col["field"] == "level":
                    levels = [clean_text(x) for x in record_col.get("levels", []) if clean_text(x)]
                    column_config[label] = st.column_config.SelectboxColumn(
                        label,
                        options=[""] + levels,
                        required=False,
                        width="medium",
                    )
                else:
                    column_config[label] = st.column_config.TextColumn(label, width="large")

            edited_df = st.data_editor(
                visible_df,
                use_container_width=True,
                num_rows="fixed",
                disabled=["학년", "반", "번호", "성명"],
                column_config=column_config,
                key=f"record_matrix_editor_{selected_assess_name}",
                height=560,
            )

            if st.button("학생별 기록 저장", type="primary", use_container_width=True):
                saved_count, warnings = save_record_matrix_df(edited_df, matrix_df, record_columns)
                st.success(f"학생별 기록을 저장했습니다. 저장된 입력값: {saved_count}개")
                if warnings:
                    with st.expander(f"성취수준 코드 확인 필요 {len(warnings)}개", expanded=True):
                        for warning in warnings[:80]:
                            st.warning(warning)
                        if len(warnings) > 80:
                            st.caption(f"나머지 {len(warnings) - 80}개 항목은 생략했습니다.")

    render_next_step_button(3)


# =========================
# ⑤ API 자료 확인
# =========================
if current_step == 4:
    st.subheader("⑤ API 입력 자료 확인")

    if st.session_state.students.empty:
        st.warning("먼저 학생 명단을 입력하세요.")
    else:
        students = sort_students_df(st.session_state.students.copy())

        class_values = sorted(
            students["반"].map(clean_text).unique().tolist(),
            key=to_int_or_big,
        )
        class_label_map = {
            (f"{class_value}반" if class_value else "반 미입력"): class_value
            for class_value in class_values
        }

        col_class, col_student = st.columns([1, 2])
        with col_class:
            selected_class_label = st.selectbox(
                "반 선택",
                list(class_label_map.keys()),
                key="api_preview_class_select",
            )

        selected_class = class_label_map[selected_class_label]
        class_students = students[students["반"].map(clean_text) == selected_class].copy()
        class_students = sort_students_df(class_students)

        student_label_map = {}
        label_counts = {}
        for _, row in class_students.iterrows():
            number_text = clean_text(row.get("번호", ""))
            name_text = clean_text(row.get("성명", ""))
            base_label = f"{number_text}번 {name_text}" if number_text else f"번호 미입력 {name_text}"
            label_counts[base_label] = label_counts.get(base_label, 0) + 1
            label = base_label if label_counts[base_label] == 1 else f"{base_label} ({label_counts[base_label]})"
            student_label_map[label] = row

        with col_student:
            selected_label = st.selectbox(
                "번호/이름 선택",
                list(student_label_map.keys()),
                key="api_preview_student_select",
            )

        selected_student = student_label_map[selected_label]
        st.caption(f"선택한 학생: {selected_class_label} {selected_label}")
        material = build_student_material(selected_student)

        st.markdown("#### API 입력 자료 미리보기")
        st.text_area("이 내용이 API 프롬프트의 수행평가 자료로 들어갑니다.", value=material, height=420)

        with st.expander("실제 프롬프트 보기"):
            st.text_area("프롬프트", value=build_prompt(material), height=420)


    render_next_step_button(4)


# =========================
# ⑥ 생기부 생성/다운로드
# =========================
if current_step == 5:
    st.subheader("⑥ 생기부 생성 / 수정 / 다운로드")

    if st.session_state.students.empty:
        st.warning("먼저 학생 명단을 입력하세요.")
    else:
        st.markdown("#### AI 선택 및 API 설정")
        col_provider, col_model, col_key = st.columns([1.25, 1.7, 2.6], gap="small")

        with col_provider:
            ai_provider = st.selectbox(
                "사용할 AI",
                AI_PROVIDER_OPTIONS,
                index=0,
                help="생기부 문장을 생성할 AI 제공자를 선택합니다.",
            )

        with col_model:
            model = st.text_input(
                "모델명",
                value=get_default_ai_model(ai_provider),
                key=f"ai_model_name_{ai_provider}",
                help="접속 시점의 최신 기본 모델명을 넣어두었습니다. 필요하면 직접 수정할 수 있습니다.",
            )

        with col_key:
            api_key = st.text_input(
                f"{ai_provider} API Key",
                value=get_default_ai_key(ai_provider),
                type="password",
                key=f"api_key_{ai_provider}",
                help=f"Streamlit Secrets에는 {AI_SECRET_KEY_NAMES.get(ai_provider, 'OPENAI_API_KEY')} 이름으로 저장해둘 수 있습니다. 입력하지 않으면 API 없이 간단 조합 방식으로 생성됩니다.",
            )

        st.caption(
            f"현재 기본 모델: {get_default_ai_model(ai_provider)} · "
            "모델명 칸은 직접 수정할 수 있습니다. API 키는 화면에 표시되지 않는 암호 입력칸입니다."
        )

        api_notice = clean_text(st.session_state.pop("api_generation_notice", ""))
        if api_notice:
            st.warning(api_notice)

        students = st.session_state.students.copy()
        student_labels = {
            f"{row['반']}반 {row['번호']}번 {row['성명']}": row
            for _, row in students.iterrows()
        }

        selected_label = st.selectbox("생성할 학생", list(student_labels.keys()))
        selected_student = student_labels[selected_label]

        col_a, col_b, col_spacer = st.columns([1.65, 1.9, 5.45], gap="small")

        with col_a:
            if st.button("선택 학생 생기부 생성", type="secondary", use_container_width=True):
                overlay_loading_started_at = datetime.now()
                overlay_slot = None
                recent_preview_items = generation_preview_items_from_results(
                    students, st.session_state.results, selected_student.get("student_id", "")
                )
                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "선택 학생 생기부 생성 중",
                    f"{selected_label}의 수행평가 자료를 정리하고 있습니다.",
                    0.20,
                    ["학생별 수행평가 입력 자료 수집", "AI 입력 프롬프트 구성", "생기부 문장 생성 준비"],
                    recent_items=recent_preview_items,
                    loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                )
                material = build_student_material(selected_student)

                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "선택 학생 생기부 생성 중",
                    f"{selected_label}의 프롬프트를 구성했습니다. AI 응답을 기다리는 중입니다.",
                    0.45,
                    ["개인정보 제외 확인", "성취수준 문구 연결", "선택한 AI 모델 호출"],
                    recent_items=recent_preview_items,
                    loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                )
                prompt = build_prompt(material)

                generated = None
                if api_key:
                    with st.spinner("API로 생성 중..."):
                        generated = generate_with_ai(prompt, ai_provider, api_key, model)

                if not generated:
                    overlay_slot = show_generation_overlay(
                        overlay_slot,
                        "선택 학생 생기부 생성 중",
                        "API 결과가 없어 내부 조합 방식으로 문장을 구성하고 있습니다.",
                        0.75,
                        ["교사의 평가 문구 추출", "중복 문구 정리", "명사형 종결 형태로 정리"],
                        recent_items=recent_preview_items,
                        loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                    )
                    generated = fallback_generate(material)

                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "선택 학생 생기부 생성 완료 처리 중",
                    "생성 문장을 결과표와 수정창에 저장하고 있습니다.",
                    0.95,
                    ["생성 원문 저장", "교사 수정 문구 초기화", "byte 계산"],
                    recent_items=recent_preview_items,
                    loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                )
                sid = selected_student["student_id"]
                st.session_state.results[sid] = {
                    "material": material,
                    "generated": generated,
                    "edited": generated,
                    "bytes": byte_count(generated),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                st.success("생성했습니다.")
                st.rerun()

        with col_b:
            job = st.session_state.generation_job

            if not job.get("active", False):
                if st.button("전체 학생 생기부 생성 시작", type="primary", use_container_width=True):
                    st.session_state.generation_job = {
                        "active": True,
                        "stop_requested": False,
                        "student_ids": students["student_id"].tolist(),
                        "index": 0,
                        "log": [],
                        "ai_provider": ai_provider,
                        "api_key": api_key,
                        "model": model,
                        "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "finished_at": "",
                    }
                    st.success("전체 생성 작업을 시작합니다.")
                    st.rerun()
            else:
                if st.button("생기부 생성 중지", type="primary", use_container_width=True):
                    job["active"] = False
                    job["stop_requested"] = True
                    job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    st.session_state.generation_job = job
                    st.warning("생성 중지를 요청했습니다. 이미 생성된 학생 결과는 저장되어 있습니다.")
                    st.rerun()

        # 전체 생성 작업 상태 표시 및 한 명씩 자동 처리
        job = st.session_state.generation_job
        if job.get("active", False) or job.get("log"):
            st.markdown("#### 전체 생성 진행 상황")
            total = len(job.get("student_ids", []))
            done = int(job.get("index", 0))

            if total > 0:
                st.progress(min(done / total, 1.0))
                st.caption(f"진행률: {done}/{total}명")

            if job.get("active", False):
                st.info("전체 생성이 진행 중입니다. 중지하려면 오른쪽의 '생기부 생성 중지' 버튼을 누르세요. 현재 처리 중인 학생은 완료된 뒤 멈춥니다.")
            elif job.get("stop_requested", False):
                st.warning("전체 생성이 중지되었습니다. 중지 전까지 생성된 문구는 아래 결과표와 다운로드 엑셀에 자동 반영됩니다.")
            elif job.get("log"):
                st.success("전체 생성 작업이 완료되었습니다.")

            if job.get("log"):
                latest_log = job.get("log", [])[-1]
                latest_label = f"{latest_log.get('반', '')}반 {latest_log.get('번호', '')}번 {latest_log.get('성명', '')}"
                st.caption(f"최근 생성 완료: {latest_label} · 생성 결과는 아래 수정 가능 표와 다운로드 엑셀에 반영됩니다.")

        if job.get("active", False) and not job.get("stop_requested", False):
            student_ids = job.get("student_ids", [])
            index = int(job.get("index", 0))
            total = len(student_ids)

            if index >= total:
                job["active"] = False
                job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state.generation_job = job
                st.success("전체 학생 생기부 생성을 완료했습니다.")
                st.rerun()

            current_sid = student_ids[index]
            matched = students[students["student_id"] == current_sid]

            if matched.empty:
                job["index"] = index + 1
                st.session_state.generation_job = job
                st.rerun()

            student = matched.iloc[0]
            label = f"{student.get('반', '')}반 {student.get('번호', '')}번 {student.get('성명', '')}"

            overlay_slot = None
            overlay_loading_started_at = job.get("started_at", "")
            progress_base = index / total if total else 0
            recent_preview_items = generation_preview_items_from_log(job.get("log", []))
            overlay_slot = show_generation_overlay(
                overlay_slot,
                "전체 학생 생기부 생성 중",
                f"{index + 1}/{total} 처리 중 · {label}",
                progress_base,
                ["현재 학생 수행평가 자료 정리", "프롬프트 구성", "AI 생성 또는 내부 조합", "결과 저장"],
                recent_items=recent_preview_items,
                loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
            )
            with st.spinner(f"{index + 1}/{total} 생성 중: {label}"):
                material = build_student_material(student)
                overlay_slot = show_generation_overlay(
                    overlay_slot,
                    "전체 학생 생기부 생성 중",
                    f"{index + 1}/{total} · {label}의 API 입력 자료를 구성했습니다.",
                    (index + 0.25) / total if total else 0,
                    ["개인정보 제외", "수행평가 입력 자료 정리", "성취수준 평가 문구 연결"],
                    recent_items=recent_preview_items,
                    loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                )
                prompt = build_prompt(material)

                generated = None
                if job.get("api_key"):
                    overlay_slot = show_generation_overlay(
                        overlay_slot,
                        "전체 학생 생기부 생성 중",
                        f"{index + 1}/{total} · {label}의 문장을 AI가 생성하고 있습니다.",
                        (index + 0.55) / total if total else 0,
                        ["AI 응답 대기 중", "응답이 끝나면 결과표에 자동 저장", "다음 학생으로 자동 이동"],
                        recent_items=recent_preview_items,
                        loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                    )
                    generated = generate_with_ai(
                        prompt,
                        job.get("ai_provider", ai_provider),
                        job.get("api_key", ""),
                        job.get("model", model),
                    )

                if not generated:
                    overlay_slot = show_generation_overlay(
                        overlay_slot,
                        "전체 학생 생기부 생성 중",
                        f"{index + 1}/{total} · {label}의 문장을 내부 조합 방식으로 구성하고 있습니다.",
                        (index + 0.75) / total if total else 0,
                        ["교사의 평가 문구 추출", "중복 표현 정리", "생기부 문체로 변환"],
                        recent_items=recent_preview_items,
                        loading_offset_seconds=loading_elapsed_seconds(overlay_loading_started_at),
                    )
                    generated = fallback_generate(material)

            st.session_state.results[current_sid] = {
                "material": material,
                "generated": generated,
                "edited": generated,
                "bytes": byte_count(generated),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

            job.setdefault("log", []).append(
                {
                    "순서": index + 1,
                    "반": student.get("반", ""),
                    "번호": student.get("번호", ""),
                    "성명": student.get("성명", ""),
                    "생성 문구": generated,
                    "byte": byte_count(generated),
                    "상태": "저장 완료",
                }
            )
            job["index"] = index + 1

            if job["index"] >= total:
                job["active"] = False
                job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state.generation_job = job
                st.success("전체 학생 생기부 생성을 완료했습니다.")
            else:
                st.session_state.generation_job = job
                st.rerun()

        st.divider()
        st.markdown("#### 전체 결과표 / 학생 선택")
        st.caption(
            "표에서 학생의 성명, byte, 생성 문구 등 아무 셀이나 클릭하면 해당 학생 한 명만 아래 큰 수정창의 대상으로 선택됩니다. "
            "문구 수정은 아래 큰 수정창에서 하거나, 필요하면 접힌 `표 형태로 여러 문구 한꺼번에 수정` 영역을 열어 처리하세요."
        )

        # 표 선택값이 없으면, 위의 '생성할 학생' 선택값을 기본 수정 대상으로 사용한다.
        valid_student_ids = set(students["student_id"].astype(str).tolist())
        if clean_text(st.session_state.get("selected_generation_student_id", "")) not in valid_student_ids:
            st.session_state.selected_generation_student_id = selected_student["student_id"]

        current_selected_sid = clean_text(st.session_state.get("selected_generation_student_id", selected_student["student_id"]))

        result_rows = []
        for _, student in students.iterrows():
            sid = student["student_id"]
            result = st.session_state.results.get(sid, {})
            text = safe_result_text(result)
            result_rows.append(
                {
                    "student_id": sid,
                    "학년": student.get("학년", ""),
                    "반": student.get("반", ""),
                    "번호": student.get("번호", ""),
                    "성명": student.get("성명", ""),
                    "생성/수정 문구": text,
                    "byte": byte_count(text),
                    "생성일시": result.get("created_at", ""),
                }
            )

        result_df = pd.DataFrame(result_rows)
        display_result_df = result_df.drop(columns=["student_id"], errors="ignore")

        # st.dataframe은 셀 선택 이벤트를 받을 수 있다.
        # 사용자가 성명, byte, 생성 문구 등 어떤 셀을 클릭하면 해당 셀의 행 번호를 읽어
        # 아래 큰 수정창의 대상 학생으로 연결한다.
        # row 선택 모드를 함께 켜면 왼쪽 선택 체크박스가 생기므로 cell 선택만 사용한다.
        # st.data_editor는 셀 클릭 이벤트와 표 편집을 동시에 안정적으로 제공하지 않으므로,
        # 메인 표는 선택 전용으로 두고 직접 수정은 아래 큰 수정창 또는 접힌 일괄 수정 표에서 처리한다.
        selection_event = st.dataframe(
            display_result_df,
            use_container_width=True,
            height=360,
            hide_index=True,
            key="generation_result_selector",
            on_select="rerun",
            selection_mode="single-cell",
            column_config={
                "학년": st.column_config.TextColumn("학년", width="small"),
                "반": st.column_config.TextColumn("반", width="small"),
                "번호": st.column_config.TextColumn("번호", width="small"),
                "성명": st.column_config.TextColumn("성명", width="medium"),
                "생성/수정 문구": st.column_config.TextColumn("생성/수정 문구", width="large"),
                "byte": st.column_config.NumberColumn("byte", width="small"),
                "생성일시": st.column_config.TextColumn("생성일시", width="medium"),
            },
        )

        selected_row_index = None
        selected_cells = []
        selected_row_indices = []

        try:
            selected_cells = list(selection_event.selection.cells)
        except Exception:
            try:
                selected_cells = list(selection_event.get("selection", {}).get("cells", []))
            except Exception:
                selected_cells = []

        if selected_cells:
            first_cell = selected_cells[-1]
            try:
                selected_row_index = int(first_cell[0])
            except Exception:
                selected_row_index = None

        if selected_row_index is None:
            try:
                selected_row_indices = list(selection_event.selection.rows)
            except Exception:
                try:
                    selected_row_indices = list(selection_event.get("selection", {}).get("rows", []))
                except Exception:
                    selected_row_indices = []

            if selected_row_indices:
                try:
                    selected_row_index = int(selected_row_indices[-1])
                except Exception:
                    selected_row_index = None

        if selected_row_index is not None and 0 <= selected_row_index < len(result_df):
            current_selected_sid = clean_text(result_df.iloc[selected_row_index].get("student_id", current_selected_sid))
            st.session_state.selected_generation_student_id = current_selected_sid

        with st.expander("표 형태로 여러 학생 문구 한꺼번에 수정", expanded=False):
            st.caption(
                "여러 학생의 문구를 표에서 한꺼번에 고칠 때만 사용하세요. "
                "학생 선택은 위 표의 성명, byte, 생성 문구 등 아무 셀이나 클릭해서 합니다."
            )

            edited_result_df = st.data_editor(
                result_df,
                use_container_width=True,
                height=360,
                hide_index=True,
                key="generation_result_bulk_editor",
                disabled=["student_id", "학년", "반", "번호", "성명", "byte", "생성일시"],
                column_config={
                    "student_id": None,
                    "학년": st.column_config.TextColumn("학년", width="small"),
                    "반": st.column_config.TextColumn("반", width="small"),
                    "번호": st.column_config.TextColumn("번호", width="small"),
                    "성명": st.column_config.TextColumn("성명", width="medium"),
                    "생성/수정 문구": st.column_config.TextColumn("생성/수정 문구", width="large"),
                    "byte": st.column_config.NumberColumn("byte", width="small"),
                    "생성일시": st.column_config.TextColumn("생성일시", width="medium"),
                },
            )

            col_save_table, col_table_info = st.columns([1.5, 4])
            with col_save_table:
                if st.button("표에서 수정한 문구 저장", type="primary", use_container_width=True):
                    student_map = {
                        row["student_id"]: row
                        for _, row in students.iterrows()
                    }
                    saved_count = 0
                    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    for _, row in edited_result_df.iterrows():
                        sid = clean_text(row.get("student_id", ""))
                        if not sid or sid not in student_map:
                            continue

                        edited_text = clean_text(row.get("생성/수정 문구", ""))
                        existing_result = st.session_state.results.get(sid, {})

                        # 이미 생성된 결과가 있거나, 표에서 새 문구를 직접 입력한 경우 저장한다.
                        if existing_result or edited_text:
                            existing_result["edited"] = edited_text
                            existing_result["bytes"] = byte_count(edited_text)
                            if not existing_result.get("generated"):
                                existing_result["generated"] = edited_text
                            if not existing_result.get("material"):
                                existing_result["material"] = build_student_material(student_map[sid])
                            if not existing_result.get("created_at"):
                                existing_result["created_at"] = now_text
                            st.session_state.results[sid] = existing_result
                            saved_count += 1

                    st.success(f"표에서 수정한 문구 {saved_count}건을 저장했습니다.")
                    st.rerun()

            with col_table_info:
                st.caption("표의 byte는 저장 후 다시 계산되어 반영됩니다. 긴 문장은 아래 큰 수정창에서 고치면 현재 byte를 바로 확인하기 쉽습니다.")

        st.divider()

        selected_match = students[students["student_id"] == current_selected_sid]
        if selected_match.empty:
            selected_detail_student = selected_student
            current_selected_sid = selected_student["student_id"]
        else:
            selected_detail_student = selected_match.iloc[0]

        selected_detail_label = (
            f"{selected_detail_student.get('반', '')}반 "
            f"{selected_detail_student.get('번호', '')}번 "
            f"{selected_detail_student.get('성명', '')}"
        )
        result = st.session_state.results.get(current_selected_sid, {})
        initial_text = safe_result_text(result)

        st.markdown(f"#### 표에서 선택한 학생 수정: {selected_detail_label}")
        if not result:
            st.info("아직 이 학생의 생기부 문구가 생성되지 않았습니다. 직접 입력하거나, 위에서 선택 학생 생기부 생성을 먼저 실행하세요.")

        editor_key = f"generation_detail_text_{current_selected_sid}"
        if editor_key not in st.session_state:
            st.session_state[editor_key] = initial_text

        edited = st.text_area(
            "교사 수정 문구",
            key=editor_key,
            height=220,
            help="입력한 내용의 byte를 아래에서 바로 확인할 수 있습니다.",
        )
        current_bytes = byte_count(edited)

        st.metric("현재 byte", current_bytes)

        if st.button("수정 문구 저장", type="primary"):
            result = st.session_state.results.get(current_selected_sid, {})
            result["edited"] = edited
            result["bytes"] = current_bytes
            if not result.get("generated"):
                result["generated"] = edited
            if not result.get("material"):
                result["material"] = build_student_material(selected_detail_student)
            if not result.get("created_at"):
                result["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.session_state.results[current_selected_sid] = result
            st.success("수정 문구를 저장했습니다.")
            st.rerun()

        st.divider()
        st.markdown("#### 최종 결과 다운로드")
        st.caption("마지막 단계입니다. 수정 내용을 확인한 뒤 결과 엑셀을 내려받으세요.")
        st.markdown(
            """
            <style>
            section.main div[data-testid="stDownloadButton"] > button {
                min-height: 58px !important;
                font-size: 1.14rem !important;
                font-weight: 900 !important;
                border-radius: 14px !important;
                box-shadow: 0 4px 12px rgba(217, 45, 32, 0.18) !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        excel_file = export_excel()
        st.download_button(
            "📥 결과 엑셀 다운로드",
            data=excel_file,
            file_name=f"BigHoneySangkibu_result_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

